# app.py
import os
import json
import glob
import sqlite3
from typing import List, Optional, Any, Dict
from datetime import datetime, date

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from fastapi.responses import JSONResponse, HTMLResponse, Response, StreamingResponse, RedirectResponse
from fastapi import Cookie
from itsdangerous import URLSafeSerializer, BadSignature
import hashlib
# from fastapi.staticfiles import StaticFiles  # 前端分離部署，不需要

# ========= 環境變數 =========
DB_PATH = os.getenv("DB_PATH", "/data/three_agents_system.db")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
KNOWLEDGE_TXT_PATH = os.getenv("KNOWLEDGE_TXT_PATH", "/data/kb.txt")
GLOBAL_KB_TEXT = ""
SESSION_SECRET = os.getenv("SESSION_SECRET", "change-me-session-secret")
session_signer = URLSafeSerializer(SESSION_SECRET, salt="session")
admin_session_signer = URLSafeSerializer(SESSION_SECRET, salt="admin_session")

# Admin 帳號（請以環境變數設定）
ADMIN_USER = os.getenv("ADMIN_USER")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")

# Google OAuth2（可選）
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")
OAUTH_REDIRECT_URI = os.getenv("OAUTH_REDIRECT_URI", "https://aijobvideobackend.zeabur.app/auth/google/callback")
try:
    from authlib.integrations.starlette_client import OAuth
    oauth = OAuth()
    if GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET:
        oauth.register(
            name="google",
            server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
            client_id=GOOGLE_CLIENT_ID,
            client_secret=GOOGLE_CLIENT_SECRET,
            client_kwargs={"scope": "openid email profile"},
        )
    _OAUTH_READY = True
except Exception as _e:
    print("[OAuth] not enabled:", _e)
    oauth = None
    _OAUTH_READY = False

# ========= App 與 CORS =========
app = FastAPI(title="Three AI Agents System with Long-term Memory")

# 掛載串流聊天路由
from chat_stream import router as chat_stream_router
app.include_router(chat_stream_router)

# 掛載點數系統
from points_integration import integrate_points_system
integrate_points_system(app)

# 動態設定 CORS：若需要帶 Cookie 就不能使用 "*"
# 預設白名單包含 GitHub Pages、Zeabur 前端子網域、正式站子網域與本機
ORIGINS = os.getenv(
    "ALLOWED_ORIGINS",
    "https://jacky6658.github.io,https://jacky6658.github.io/Altest/,http://localhost:3000,https://video.aijob.com.tw,https://aijobvideo.zeabur.app,https://aijob.com.tw"
).split(",")
ORIGINS = [o.strip().rstrip("/") for o in ORIGINS if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ORIGINS,
    allow_credentials=True,
    allow_methods=["POST", "OPTIONS", "GET", "PUT", "DELETE"],
    allow_headers=["*"],
)

# OAuth 需要 Starlette SessionMiddleware；使用獨立 cookie 名稱避免與本系統 session 混淆
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    session_cookie="oauth_session",
    same_site="none",
    https_only=True,
)

# 前端分離部署，不需要靜態文件服務

# ========= 引導式問答狀態（記憶體暫存） =========
QA_SESSIONS: Dict[str, Dict[str, Any]] = {}  # key: session_id
QA_QUESTIONS = [
    {"key":"structure","q":"【Q1】請選擇腳本結構（A 三段式 / B 問題解決 / C Before-After / D 教學 / E 敘事 / F 爆點連發）"},
    {"key":"duration","q":"【Q2】影片時長（30 或 60 秒）"},
    {"key":"topic","q":"【Q3】請輸入主題或產品名稱"},
    {"key":"goal","q":"【Q4】主要目標（吸流量 / 教育 / 轉單 / 品牌）"},
    {"key":"audience","q":"【Q5】目標受眾（年齡/性別/特質/痛點）"},
    {"key":"hook","q":"【Q6】開場鉤子類型（問句/反差/同理/數字）＋想放的關鍵詞"},
    {"key":"cta","q":"【Q7】CTA（關注/收藏 / 留言/私訊 / 購買連結）"}
]

def qa_reset(session_id: str):
    QA_SESSIONS[session_id] = {"step": 0, "answers": {}}

def qa_next_question(session_id: str) -> Optional[str]:
    st = QA_SESSIONS.get(session_id)
    if not st: return None
    step = st["step"]
    if step < len(QA_QUESTIONS):
        return QA_QUESTIONS[step]["q"]
    return None

def qa_record_answer(session_id: str, user_text: str):
    st = QA_SESSIONS.get(session_id)
    if not st: return
    step = st["step"]
    if step < len(QA_QUESTIONS):
        key = QA_QUESTIONS[step]["key"]
        st["answers"][key] = user_text
        st["step"] = step + 1

def compose_brief_from_answers(ans: Dict[str,str]) -> str:
    labels = {
        "structure":"結構","duration":"時長","topic":"主題","goal":"目標","audience":"受眾",
        "hook":"鉤子","cta":"CTA"
    }
    lines = []
    for it in QA_QUESTIONS:
        k = it["key"]
        if k in ans:
            lines.append(f"{labels.get(k,k)}：{ans[k]}")
    return "；".join(lines)

# ========= 簡易 KB 檢索 =========
def load_kb_text() -> str:
    path = KNOWLEDGE_TXT_PATH
    try:
        if os.path.exists(path):
            with open(path,"r",encoding="utf-8") as f:
                return f.read()
    except Exception:
        pass
    return ""

def retrieve_context(query: str, max_chars: int = 1200) -> str:
    text = GLOBAL_KB_TEXT or ""
    if not text: 
        return ""
    import re
    toks = [t for t in re.split(r'[\s，。；、,.:?!\-\/\[\]()]+', (query or "")) if len(t)>=1]
    toks = list(dict.fromkeys(toks))
    lines = text.splitlines()
    scored = []
    for i, line in enumerate(lines):
        score = sum(1 for t in toks if t and t in line)
        if score>0:
            scored.append((score, i, line))
    scored.sort(key=lambda x:(-x[0], x[1]))
    selected=[]
    total=0
    for _, _, ln in scored:
        if not ln.strip(): 
            continue
        take = ln.strip()
        if total + len(take) + 1 > max_chars:
            break
        selected.append(take)
        total += len(take) + 1
    if not selected:
        return text[:max_chars]
    return "\n".join(selected)

# ========= DB =========
def _ensure_db_dir(path: str):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

def get_conn() -> sqlite3.Connection:
    _ensure_db_dir(DB_PATH)
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    _ensure_db_dir(DB_PATH)
    conn = get_conn()
    cur = conn.cursor()
    
    # 原有表格
    cur.execute("""
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            user_id TEXT,
            user_input TEXT,
            mode TEXT,
            messages_json TEXT,
            previous_segments_json TEXT,
            response_json TEXT
        )
    """)
    try:
        cur.execute("ALTER TABLE requests ADD COLUMN user_id TEXT")
    except Exception:
        pass
    
    # 新增：三智能體系統表格
    # 1. 用戶基本資訊表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            email TEXT,
            name TEXT,
            platform_preferences TEXT,
            language_preference TEXT DEFAULT 'zh-TW',
            timezone TEXT DEFAULT 'Asia/Taipei',
            status TEXT DEFAULT 'active'
        )
    """)

    # 新增：Email/帳號登入表（本地帳號）
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users_auth (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            username TEXT UNIQUE,
            email TEXT,
            phone TEXT,
            password_hash TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(user_id)
        )
    """)
    
    # 2. 用戶定位檔案表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            business_type TEXT,
            target_audience TEXT,
            brand_voice TEXT,
            content_goals TEXT,
            primary_platform TEXT,
            secondary_platforms TEXT,
            posting_frequency TEXT,
            preferred_topics TEXT,
            content_styles TEXT,
            video_duration_preference TEXT,
            competitors TEXT,
            unique_value_proposition TEXT,
            current_followers INTEGER DEFAULT 0,
            engagement_rate REAL DEFAULT 0.0,
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            UNIQUE(user_id)
        )
    """)
    
    # 3. 會話記錄表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            session_id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            agent_type TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'active',
            context_summary TEXT,
            key_insights TEXT,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    """)
    
    # 4. 對話記錄表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            metadata TEXT,
            FOREIGN KEY (session_id) REFERENCES sessions(session_id)
        )
    """)
    
    # 5. 智能體記憶表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS agent_memories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            agent_type TEXT NOT NULL,
            memory_type TEXT NOT NULL,
            content TEXT NOT NULL,
            importance_score INTEGER DEFAULT 5,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            last_accessed DATETIME DEFAULT CURRENT_TIMESTAMP,
            access_count INTEGER DEFAULT 1,
            tags TEXT,
            related_memories TEXT,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    """)
    
    # 6. 選題建議表
    cur.execute("""
        CREATE TABLE IF NOT EXISTS topic_suggestions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            suggested_date DATE NOT NULL,
            topics TEXT NOT NULL,
            reasoning TEXT,
            user_feedback TEXT,
            used_count INTEGER DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            UNIQUE(user_id, suggested_date)
        )
    """)
    
    # 建立索引
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sessions_user_agent ON sessions(user_id, agent_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_session ON messages(session_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_memories_user_agent ON agent_memories(user_id, agent_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_memories_importance ON agent_memories(importance_score DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_topic_suggestions_user_date ON topic_suggestions(user_id, suggested_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_requests_user_time ON requests(user_id, created_at DESC)")

    # 用戶點數與訂單（簡化）
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS user_credits (
            user_id TEXT PRIMARY KEY,
            balance INTEGER DEFAULT 0,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            order_type TEXT NOT NULL,
            amount INTEGER DEFAULT 0,
            plan TEXT,
            status TEXT DEFAULT 'paid',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            paid_at DATETIME
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_user_time ON orders(user_id, created_at DESC)")

    # 訂閱方案表
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS subscription_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            price INTEGER NOT NULL,
            credits INTEGER NOT NULL,
            duration_days INTEGER DEFAULT 30,
            is_active BOOLEAN DEFAULT 1,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    # 用戶訂閱記錄
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS user_subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            plan_id INTEGER NOT NULL,
            status TEXT DEFAULT 'active',
            start_date DATETIME DEFAULT CURRENT_TIMESTAMP,
            end_date DATETIME,
            auto_renew BOOLEAN DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (plan_id) REFERENCES subscription_plans (id)
        )
        """
    )

    # 管理操作稽核表
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS admin_audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            action TEXT NOT NULL,
            admin_token_hash TEXT,
            target_user_id TEXT,
            details TEXT
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_action_time ON admin_audit_logs(action, created_at DESC)")
    
    conn.commit()
    conn.close()

@app.on_event("startup")
def on_startup():
    try:
        init_db()
        global GLOBAL_KB_TEXT
        GLOBAL_KB_TEXT = load_kb_text()
        print(f"[BOOT] KB loaded from {KNOWLEDGE_TXT_PATH} len={len(GLOBAL_KB_TEXT)}")
        print(f"[BOOT] DB ready at {DB_PATH}")
    except Exception as e:
        print("[BOOT] DB init failed:", e)

@app.get("/healthz")
def healthz(): return {"ok": True}

@app.get("/favicon.ico")
def favicon(): return Response(status_code=204)

@app.get("/", response_class=HTMLResponse)
def api_info():
    """API 資訊頁面"""
    return """
    <html><body style="font-family: Arial, sans-serif; max-width: 800px; margin: 50px auto; padding: 20px;">
      <h1>🎯 三智能體長期記憶系統</h1>
      <p>後端 API 服務已啟動！前端請訪問：<a href="https://jacky6658.github.io/Altest/" target="_blank">https://jacky6658.github.io/Altest/</a></p>
      
      <h2>📋 API 端點列表</h2>
      
      <h3>原有功能：</h3>
      <ul>
        <li><code>POST /chat_generate</code> - 腳本/文案二合一生成</li>
        <li><code>POST /generate_script</code> - 舊流程保留</li>
        <li><code>POST /chat_qa</code> - 引導式問答</li>
        <li><code>POST /export/xlsx</code> - Excel 匯出</li>
      </ul>
      
      <h3>新增三智能體功能：</h3>
      <ul>
        <li><strong>定位智能體</strong></li>
        <ul>
          <li><code>POST /agent/positioning/analyze</code> - 分析用戶定位</li>
          <li><code>PUT /agent/positioning/profile</code> - 更新定位檔案</li>
        </ul>
        <li><strong>選題智能體</strong></li>
        <ul>
          <li><code>POST /agent/topics/suggest</code> - 獲取選題建議</li>
          <li><code>GET /agent/topics/history</code> - 選題歷史</li>
        </ul>
        <li><strong>腳本文案智能體</strong></li>
        <ul>
          <li><code>POST /agent/content/generate</code> - 生成腳本/文案（增強版）</li>
        </ul>
        <li><strong>記憶系統</strong></li>
        <ul>
          <li><code>GET /memory/user/{user_id}</code> - 獲取用戶記憶</li>
          <li><code>POST /memory/add</code> - 添加記憶</li>
        </ul>
      </ul>
      
      <h2>🔧 系統狀態</h2>
      <p>✅ 資料庫：已初始化</p>
      <p>✅ 知識庫：已載入</p>
      <p>✅ 三智能體：已啟動</p>
      <p>✅ 長期記憶：已啟用</p>
    </body></html>
    """

# ========= Email/帳號 註冊 / 登入 / 會話 =========
from fastapi import Body
from fastapi import Form

@app.post("/auth/signup")
async def auth_signup(req: Request):
    data = await req.json()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not all([email, username, password, phone]):
        raise HTTPException(status_code=400, detail="missing_fields")

    user_id = f"u_{hashlib.md5((email+username).encode()).hexdigest()[:12]}"

    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        # 建立 users（若不存在）
        u = conn.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,)).fetchone()
        if not u:
            conn.execute(
                "INSERT INTO users (user_id, email, name) VALUES (?, ?, ?)",
                (user_id, email, username)
            )
        # 建立 users_auth
        conn.execute(
            "INSERT INTO users_auth (user_id, username, email, phone, password_hash) VALUES (?, ?, ?, ?, ?)",
            (user_id, username, email, phone, hash_password(password))
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=409, detail="user_exists")
    conn.close()
    return {"ok": True, "user_id": user_id}

@app.post("/auth/login")
async def auth_login(req: Request):
    data = await req.json()
    identifier = (data.get("identifier") or "").strip()
    password = (data.get("password") or "").strip()
    if not identifier or not password:
        raise HTTPException(status_code=400, detail="missing_fields")
    conn = get_conn(); conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT user_id, password_hash FROM users_auth WHERE username=? OR email=?",
        (identifier, identifier)
    ).fetchone()
    conn.close()
    if not row or row["password_hash"] != hash_password(password):
        raise HTTPException(status_code=401, detail="invalid_credentials")

    # 設置 Session Cookie
    token = create_session_cookie(row["user_id"])
    resp = JSONResponse({"ok": True})
    resp.set_cookie(
        "session", token,
        httponly=True, secure=True, samesite="none", max_age=7*24*3600
    )
    return resp

# 前端跨網域第三方 Cookie 可能被瀏覽器阻擋，提供彈窗版登入：
# 以第一方情境設定 session，最後回傳頁面自動通知 opener 並關閉
@app.post("/auth/login_popup", response_class=HTMLResponse)
async def auth_login_popup(identifier: str = Form(...), password: str = Form(...)):
    if not identifier or not password:
        return HTMLResponse("<p>missing fields</p>", status_code=400)
    conn = get_conn(); conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT user_id, password_hash FROM users_auth WHERE username=? OR email=?",
        (identifier, identifier)
    ).fetchone()
    conn.close()
    if not row or row["password_hash"] != hash_password(password):
        return HTMLResponse("<script>window.close()</script>", status_code=401)
    token = create_session_cookie(row["user_id"])
    html = f"""
<!DOCTYPE html><meta charset=\"utf-8\" />
<script>
  try {{
    document.cookie = "session={token}; Path=/; SameSite=None; Secure; HttpOnly";
  }} catch (_e) {{}}
  try {{ if (window.opener) window.opener.postMessage('login_ok','*'); }} catch(_e) {{}}
  window.close();
</script>
"""
    return HTMLResponse(html)

@app.get("/me")
def me(session: str | None = Cookie(default=None)):
    uid = verify_session_cookie(session) if session else None
    if not uid:
        return JSONResponse(status_code=401, content={"error": "unauthorized"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    u = conn.execute("SELECT user_id, email, name FROM users WHERE user_id=?", (uid,)).fetchone()
    conn.close()
    if not u:
        return JSONResponse(status_code=404, content={"error": "user_not_found"})
    return {"id": u["user_id"], "email": u["email"], "name": u["name"]}

@app.post("/logout")
def logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("session")
    return resp

# ========= Auth Helpers =========
def hash_password(password: str) -> str:
    return hashlib.sha256((password or "").encode("utf-8")).hexdigest()

def create_session_cookie(user_id: str) -> str:
    return session_signer.dumps({"uid": user_id, "ts": int(datetime.now().timestamp())})

def verify_session_cookie(cookie_val: str) -> str | None:
    try:
        data = session_signer.loads(cookie_val)
        return data.get("uid")
    except BadSignature:
        return None

def create_admin_session_cookie(username: str) -> str:
    return admin_session_signer.dumps({"adm": username, "ts": int(datetime.now().timestamp())})

def verify_admin_session_cookie(cookie_val: str) -> str | None:
    try:
        data = admin_session_signer.loads(cookie_val)
        return data.get("adm")
    except BadSignature:
        return None

# ========= Google OAuth2 Endpoints =========
@app.get("/auth/google/start")
async def google_start(request: Request, next: str | None = "/"):
    if not _OAUTH_READY:
        return JSONResponse(status_code=501, content={"error": "oauth_not_configured"})
    resp = await oauth.google.authorize_redirect(
        request,
        redirect_uri=OAUTH_REDIRECT_URI,
        state=next or "/",
        prompt="consent select_account",
        max_age=0,
    )
    # 確保不被站內既有 session 影響，先清除使用者 session cookie
    try:
        resp.delete_cookie("session")
    except Exception:
        pass
    return resp

@app.get("/auth/google/callback")
async def google_callback(request: Request):
    if not _OAUTH_READY:
        return JSONResponse(status_code=501, content={"error": "oauth_not_configured"})
    try:
        token = await oauth.google.authorize_access_token(request)
        try:
            print("[OAuth] token keys:", sorted(list((token or {}).keys())))
        except Exception:
            pass
        # 主要路徑：使用 id_token 解析
        idinfo = None
        try:
            idinfo = await oauth.google.parse_id_token(request, token)
            try:
                print("[OAuth] id_token parsed, keys:", sorted(list((idinfo or {}).keys())))
            except Exception:
                pass
        except Exception:
            idinfo = None
        # 後備路徑：若無 id_token，改呼叫 userinfo 端點
        if not idinfo or not idinfo.get("sub"):
            try:
                resp = await oauth.google.get("userinfo", token=token)
                idinfo = resp.json() if resp is not None else {}
                try:
                    print("[OAuth] userinfo keys:", sorted(list((idinfo or {}).keys())))
                except Exception:
                    pass
            except Exception:
                idinfo = {}
            # 進一步後備：直接以 access_token 打 OIDC userinfo 端點
            try:
                import httpx
                at = (token or {}).get("access_token")
                if at and (not idinfo or not idinfo.get("sub")):
                    async with httpx.AsyncClient(timeout=8) as client:
                        r = await client.get(
                            "https://openidconnect.googleapis.com/v1/userinfo",
                            headers={"Authorization": f"Bearer {at}"},
                        )
                        try:
                            print("[OAuth] direct userinfo status:", r.status_code)
                        except Exception:
                            pass
                        if r.status_code == 200:
                            idinfo = r.json()
                            try:
                                print("[OAuth] direct userinfo keys:", sorted(list((idinfo or {}).keys())))
                            except Exception:
                                pass
            except Exception as _e:
                try:
                    print("[OAuth] direct userinfo error:", _e)
                except Exception:
                    pass
        sub = (idinfo or {}).get("sub"); email = (idinfo or {}).get("email"); name = (idinfo or {}).get("name") or (email.split("@")[0] if email else "user")
        # 最後備援：若缺少 sub 但有 email，使用 email 雜湊生成穩定 ID
        if not sub and email:
            import hashlib
            sub = hashlib.sha256(email.encode("utf-8")).hexdigest()[:24]
        if not sub:
            try:
                print("[OAuth] missing sub/email. idinfo keys:", sorted(list((idinfo or {}).keys())))
            except Exception:
                pass
            return JSONResponse(status_code=400, content={"error": "invalid_google_response"})
        user_id = f"g_{sub}"
        create_or_get_user(user_id, email=email, name=name)
        token_val = create_session_cookie(user_id)
        resp = RedirectResponse(url=request.query_params.get("state") or "/")
        resp.set_cookie("session", token_val, httponly=True, secure=True, samesite="none", max_age=7*24*3600)
        return resp
    except Exception as e:
        print("[OAuth Callback Error]", e)
        return JSONResponse(status_code=500, content={"error": "oauth_failed"})

# ========= 內建知識庫 =========
BUILTIN_KB_SCRIPT = """
【短影音腳本原則（濃縮）】
1) Hook(0-5s) → Value → CTA。60s 版可拆 5~6 段，節奏清楚。
2) 每段輸出：type/start_sec/end_sec/camera/dialog/visual/cta。
3) Hook 用痛點/反差/數據鉤子 + 快節奏 B-roll；Value 拆重點；CTA 動詞+利益+下一步。
4) 語氣口語、短句、有節奏，避免空話。
"""

BUILTIN_KB_COPY = """
【社群文案原則（濃縮）】
1) 結構：吸睛開頭 → 主體賣點/故事 → CTA → Hashtags。
2) 風格：貼近受眾、短句、可搭 emoji、結尾有動作。
3) Hashtags：主關鍵字 1-3、延伸 5-8。
4) 欄位：main_copy / alternates / hashtags / cta / image_ideas（平台化圖片建議）。
"""

def load_extra_kb(max_chars=2500) -> str:
    chunks, total = [], 0
    if GLOBAL_KB_TEXT:
        seg = GLOBAL_KB_TEXT[:max_chars]
        chunks.append(f"\n[KB:global]\n{seg}")
        total += len(seg)
    else:
        paths = glob.glob("/data/kb*.txt") + glob.glob("/data/*.kb.txt") + glob.glob("/data/knowledge*.txt")
        for p in paths:
            try:
                t = open(p, "r", encoding="utf-8").read().strip()
                if not t: continue
                take = (max_chars - total)
                seg = t[:take]
                if seg:
                    chunks.append(f"\n[KB:{os.path.basename(p)}]\n{seg}")
                    total += len(seg)
                if total >= max_chars: break
            except Exception:
                pass
    return "\n".join(chunks)

EXTRA_KB = load_extra_kb()

# ========= 提示字 & 工具 =========
SHORT_HINT_SCRIPT = "內容有點太短了 🙏 請提供：行業/平台/時長(秒)/目標/主題（例如：『電商｜Reels｜60秒｜購買｜夏季新品開箱』），我就能生成完整腳本。"
SHORT_HINT_COPY   = "內容有點太短了 🙏 請提供：平台/受眾/語氣/主題/CTA（例如：『IG｜男生視角｜活力回歸｜CTA：點連結』），我就能生成完整貼文。"

def _ensure_json_block(text: str) -> str:
    if not text: raise ValueError("empty model output")
    t = text.strip()
    if t.startswith("```"):
        parts = t.split("```")
        if len(parts) >= 3: t = parts[1]
    i = min([x for x in (t.find("{"), t.find("[")) if x >= 0], default=-1)
    if i < 0: return t
    j = max(t.rfind("}"), t.rfind("]"))
    if j > i: return t[i:j+1]
    return t

def detect_mode(messages: List[Dict[str, str]], explicit: Optional[str]) -> str:
    """優先使用 explicit；否則用關鍵字判斷。"""
    if explicit in ("script", "copy"):
        return explicit
    last = ""
    for m in reversed(messages or []):
        if m.get("role") == "user":
            last = (m.get("content") or "").lower()
            break
    copy_keys = [
        "文案","貼文","copy","hashtag","hashtags",
        "ig","facebook","fb","linkedin","小紅書","x（twitter）","x/twitter","抖音文案"
    ]
    if any(k in last for k in copy_keys):
        return "copy"
    return "script"

def parse_segments(json_text: str) -> List[Dict[str, Any]]:
    data = json.loads(json_text)
    if isinstance(data, dict) and "segments" in data: data = data["segments"]
    if not isinstance(data, list): raise ValueError("segments must be a list")
    segs = []
    for it in data:
        segs.append({
            "type": it.get("type") or it.get("label") or "場景",
            "start_sec": it.get("start_sec", None),
            "end_sec": it.get("end_sec", None),
            "camera": it.get("camera", ""),
            "dialog": it.get("dialog", ""),
            "visual": it.get("visual", ""),
            "cta": it.get("cta", "")
        })
    return segs

def parse_copy(json_text: str) -> Dict[str, Any]:
    data = json.loads(json_text)
    if isinstance(data, list): data = data[0] if data else {}
    return {
        "main_copy":   data.get("main_copy", ""),
        "alternates":  data.get("alternates", []) or data.get("openers", []),
        "hashtags":    data.get("hashtags", []),
        "cta":         data.get("cta", ""),
        "image_ideas": data.get("image_ideas", [])
    }

# === NEW: 模板/時長/模式說明 ===
TEMPLATE_GUIDE = {
    "A": "三段式：Hook → Value → CTA。重點清楚、節奏明快，適合廣泛情境。",
    "B": "問題解決：痛點 → 解法 → 證據/示例 → CTA。適合教育與導購。",
    "C": "Before-After：改變前後對比，強調差異與收益 → CTA。適合案例/見證。",
    "D": "教學：步驟化教學（1-2-3）+ 注意事項 → CTA。適合技巧分享。",
    "E": "敘事：小故事鋪陳 → 轉折亮點 → CTA。適合品牌情緒/人物敘事。",
    "F": "爆點連發：連續強 Hook/金句/反差點，最後收斂 → CTA。適合抓注意力。"
}

def _duration_plan(duration: Optional[int]) -> Dict[str, Any]:
    """
    回傳分段建議與 fewshot JSON。30s 走 3 段；60s 走 6 段（每段~10s）。
    """
    if int(duration or 0) == 60:
        fewshot = """
{"segments":[
  {"type":"hook","start_sec":0,"end_sec":10,"camera":"CU","dialog":"...","visual":"...","cta":""},
  {"type":"value1","start_sec":10,"end_sec":20,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"value2","start_sec":20,"end_sec":30,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"value3","start_sec":30,"end_sec":40,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"value4","start_sec":40,"end_sec":50,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"cta","start_sec":50,"end_sec":60,"camera":"WS","dialog":"...","visual":"...","cta":"..."}
]}
"""
        return {"fewshot": fewshot, "note": "請以 60 秒約 6 段輸出，段與段間節奏分明。"}
    # default 30s
    fewshot = """
{"segments":[
  {"type":"hook","start_sec":0,"end_sec":5,"camera":"CU","dialog":"...","visual":"...","cta":""},
  {"type":"value","start_sec":5,"end_sec":25,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"cta","start_sec":25,"end_sec":30,"camera":"WS","dialog":"...","visual":"...","cta":"..."}
]}
"""
    return {"fewshot": fewshot, "note": "請以 30 秒 3 段輸出，Hook 要強、CTA 明確。"}

def build_script_prompt(
    user_input: str,
    previous_segments: List[Dict[str, Any]],
    template_type: Optional[str] = None,
    duration: Optional[int] = None,
    dialogue_mode: Optional[str] = None,
    knowledge_hint: Optional[str] = None,
) -> str:
    plan = _duration_plan(duration)
    fewshot = plan["fewshot"]
    duration_note = plan["note"]
    tmpl = (template_type or "").strip().upper()
    tmpl_text = TEMPLATE_GUIDE.get(tmpl, "未指定模板時由你判斷最合適的結構。")

    kb = (BUILTIN_KB_SCRIPT + "\n" + (EXTRA_KB or "")).strip()
    # 動態 KB 擷取：合併使用者輸入與可選提示
    q = user_input
    if knowledge_hint:
        q = f"{knowledge_hint}\n{user_input}"
    try:
        kb_ctx_dynamic = retrieve_context(q)
    except Exception:
        kb_ctx_dynamic = ""

    prev = json.dumps(previous_segments or [], ensure_ascii=False)

    mode_line = ""
    if (dialogue_mode or "").lower() == "free":
        mode_line = "語氣更自由、可主動提出精煉建議與反問以完善腳本；"
    elif (dialogue_mode or "").lower() == "guide":
        mode_line = "語氣偏引導，逐步釐清要素後直接給出完整分段；"

    return f"""
根據使用者輸入生成短影音腳本。{mode_line}

🎯 腳本參數：
• 模板：{tmpl or "（未指定）"} - {tmpl_text}
• 時長：{int(duration) if duration else "（未指定，預設 30）"} 秒
• 平台：Instagram Reels、TikTok、YouTube Shorts、Facebook Reels

📚 知識庫：
{kb}

【KB輔助摘錄】（若空白代表無）
{kb_ctx_dynamic[:1000]}

💡 台灣市場特色：
• 內容風格：生活化、親切、實用
• 節奏要求：2-3秒換畫面，節奏緊湊
• Hook原則：0-5秒直給結論，用大字卡與強情緒表情
• 語氣：堅定、直給結論，避免口癖贅字

使用者輸入：
{user_input}

已接受段落：
{prev}

直接輸出JSON格式，不要任何開場白或說明文字：
{fewshot}
"""

def build_copy_prompt(user_input: str, topic: Optional[str]) -> str:
    topic_line = f"\n【主題】{topic}" if topic else ""
    fewshot = """
{
  "main_copy":"主貼文（含換行與 emoji）",
  "alternates":["備選開頭A","備選開頭B","備選開頭C"],
  "hashtags":["#關鍵字1","#關鍵字2","#延伸3","#延伸4"],
  "cta":"行動呼籲一句話",
  "image_ideas":["配圖/照片/示意圖建議1","建議2","建議3"]
}
"""
    kb = (BUILTIN_KB_COPY + "\n" + (EXTRA_KB or "")).strip()
    return f"""
你是社群文案顧問。請依「使用者輸入」與可選的主題輸出**JSON**，包含主貼文、備選開頭、Hashtags、CTA，並加入 image_ideas（平台導向的圖片/拍法/視覺建議）。語氣可口語並適度使用 emoji。

{kb}

使用者輸入：
{user_input}{topic_line}

只回傳 JSON（單一物件，不要 markdown fence）：
{fewshot}
"""

# ========= Gemini =========
def use_gemini() -> bool: return bool(GEMINI_API_KEY)

def gemini_generate_text(prompt: str) -> str:
    import google.generativeai as genai
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel(GEMINI_MODEL)
    res = model.generate_content(prompt)
    return (res.text or "").strip()

# ========= Fallback =========
def fallback_segments(user_input: str, prev_len: int, duration: Optional[int]=None) -> List[Dict[str, Any]]:
    d = int(duration or 30)
    if d >= 60:
        # 粗略 60s 六段
        labels = ["hook","value1","value2","value3","value4","cta"]
        segs=[]
        start=0
        for i,l in enumerate(labels):
            end = 10*(i+1)
            if i==len(labels)-1: end = 60
            cam = "CU" if i==0 else ("WS" if i==len(labels)-1 else "MS")
            segs.append({
                "type": l, "start_sec": start, "end_sec": end, "camera": cam,
                "dialog": f"（模擬）{user_input[:36]}…",
                "visual": "（模擬）快切 B-roll / 大字卡",
                "cta": "點連結領取 🔗" if l=="cta" else ""
            })
            start = end
        return segs
    # 預設 30s 三段
    step = prev_len
    return [{
        "type": "hook" if step == 0 else ("cta" if step >= 2 else "value"),
        "start_sec": 0 if step == 0 else 5 if step == 1 else 25,
        "end_sec":   5 if step == 0 else 25 if step == 1 else 30,
        "camera": "CU" if step == 0 else "MS" if step == 1 else "WS",
        "dialog": f"（模擬）{user_input[:36]}…",
        "visual": "（模擬）快切 B-roll / 大字卡",
        "cta": "點連結領取 🔗" if step >= 2 else ""
    }]

def fallback_copy(user_input: str, topic: Optional[str]) -> Dict[str, Any]:
    t = f"（主題：{topic}）" if topic else ""
    return {
        "main_copy":  f"（模擬）IG 貼文：{user_input} {t}\n精神回歸、效率回升！⚡️\n今天就行動吧！",
        "alternates": ["🔥 今天就開始","💡 其實只要這樣做","👉 你也可以"],
        "hashtags":   ["#行銷","#AI","#文案","#社群經營"],
        "cta":        "立即點連結 🔗",
        "image_ideas":["產品近拍 + 生活情境","品牌色背景大字卡","步驟流程示意圖"]
    }

# ========= 三智能體系統核心功能 =========

# 用戶管理
def create_or_get_user(user_id: str, email: str = None, name: str = None) -> Dict:
    """創建或獲取用戶"""
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    
    user = conn.execute(
        "SELECT * FROM users WHERE user_id = ?", (user_id,)
    ).fetchone()
    
    if not user:
        conn.execute(
            "INSERT INTO users (user_id, email, name) VALUES (?, ?, ?)",
            (user_id, email, name)
        )
        conn.commit()
        user = conn.execute(
            "SELECT * FROM users WHERE user_id = ?", (user_id,)
        ).fetchone()
    
    conn.close()
    return dict(user) if user else None

def get_user_profile(user_id: str) -> Optional[Dict]:
    """獲取用戶定位檔案"""
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    profile = conn.execute(
        "SELECT * FROM user_profiles WHERE user_id = ?", (user_id,)
    ).fetchone()
    conn.close()
    return dict(profile) if profile else None

def update_user_profile(user_id: str, profile_data: Dict) -> bool:
    """更新用戶定位檔案"""
    conn = get_conn()
    
    existing = conn.execute(
        "SELECT id FROM user_profiles WHERE user_id = ?", (user_id,)
    ).fetchone()
    
    if existing:
        update_fields = []
        values = []
        for key, value in profile_data.items():
            if key != 'user_id' and value is not None:
                update_fields.append(f"{key} = ?")
                values.append(json.dumps(value) if isinstance(value, (list, dict)) else str(value))
        
        if update_fields:
            values.append(user_id)
            sql = f"UPDATE user_profiles SET {', '.join(update_fields)}, updated_at = CURRENT_TIMESTAMP WHERE user_id = ?"
            conn.execute(sql, values)
    else:
        profile_data['user_id'] = user_id
        fields = list(profile_data.keys())
        placeholders = ['?' for _ in fields]
        values = [json.dumps(v) if isinstance(v, (list, dict)) else str(v) for v in profile_data.values()]
        
        sql = f"INSERT INTO user_profiles ({', '.join(fields)}) VALUES ({', '.join(placeholders)})"
        conn.execute(sql, values)
    
    conn.commit()
    conn.close()
    return True

# 會話管理
def create_session(user_id: str, agent_type: str) -> str:
    """創建新會話"""
    session_id = f"{user_id}_{agent_type}_{int(datetime.now().timestamp())}"
    conn = get_conn()
    
    conn.execute(
        "INSERT INTO sessions (session_id, user_id, agent_type) VALUES (?, ?, ?)",
        (session_id, user_id, agent_type)
    )
    conn.commit()
    conn.close()
    
    return session_id

def add_message(session_id: str, role: str, content: str, metadata: Dict = None):
    """添加對話記錄"""
    conn = get_conn()
    conn.execute(
        "INSERT INTO messages (session_id, role, content, metadata) VALUES (?, ?, ?, ?)",
        (session_id, role, content, json.dumps(metadata) if metadata else None)
    )
    conn.commit()
    conn.close()

# 記憶系統
def add_memory(user_id: str, agent_type: str, memory_type: str, content: str, 
               importance_score: int = 5, tags: List[str] = None) -> int:
    """添加記憶"""
    conn = get_conn()
    
    cursor = conn.execute(
        """INSERT INTO agent_memories 
           (user_id, agent_type, memory_type, content, importance_score, tags) 
           VALUES (?, ?, ?, ?, ?, ?)""",
        (user_id, agent_type, memory_type, content, importance_score, 
         json.dumps(tags) if tags else None)
    )
    memory_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return memory_id

def get_user_memories(user_id: str, agent_type: str = None, memory_type: str = None, 
                     limit: int = 20) -> List[Dict]:
    """獲取用戶記憶"""
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    
    conditions = ["user_id = ?"]
    params = [user_id]
    
    if agent_type:
        conditions.append("agent_type = ?")
        params.append(agent_type)
    
    if memory_type:
        conditions.append("memory_type = ?")
        params.append(memory_type)
    
    params.append(limit)
    
    sql = f"""SELECT * FROM agent_memories 
              WHERE {' AND '.join(conditions)} 
              ORDER BY importance_score DESC, last_accessed DESC 
              LIMIT ?"""
    
    memories = conn.execute(sql, params).fetchall()
    conn.close()
    
    return [dict(memory) for memory in memories]

# 定位智能體
def positioning_agent_analyze(user_input: str, user_profile: Dict = None, memories: List[Dict] = None) -> str:
    """定位智能體分析 - 提供結構化定位選項"""
    context = "你是專業的短影音定位顧問，專門服務台灣市場，幫助用戶快速建立清晰的帳號定位。\n\n"
    
    # 加入知識庫內容
    kb_context = retrieve_context(user_input) or ""
    if kb_context:
        context += f"【知識庫參考】\n{kb_context}\n\n"
    
    if user_profile:
        context += f"用戶現有檔案：{json.dumps(user_profile, ensure_ascii=False)}\n\n"
    
    if memories:
        context += f"相關記憶：\n"
        for memory in memories[:5]:
            context += f"- {memory['content']}\n"
        context += "\n"
    
    context += f"用戶輸入：{user_input}\n\n"
    
    # 檢查哪些欄位還需要填寫
    missing_fields = []
    if not user_profile or not user_profile.get('business_type'):
        missing_fields.append("業務類型")
    if not user_profile or not user_profile.get('target_audience'):
        missing_fields.append("目標受眾")
    if not user_profile or not user_profile.get('brand_voice'):
        missing_fields.append("品牌語氣")
    if not user_profile or not user_profile.get('primary_platform'):
        missing_fields.append("主要平台")
    if not user_profile or not user_profile.get('content_goals'):
        missing_fields.append("內容目標")
    if not user_profile or not user_profile.get('posting_frequency'):
        missing_fields.append("發文頻率")
    
    context += """【重要】請基於知識庫內容，以結構化方式回應，提供具體的定位選項供用戶選擇：

📋 回應格式要求：
• 使用emoji作為分點符號，讓內容更易讀
• 段落分明，重點突出
• 提供具體實作方式
• 在回覆中明確標示「業務類型：」「目標受眾：」等欄位，方便系統自動提取
• 基於知識庫的流量/轉換邏輯、平台策略、內容結構等專業建議

🎯 分析步驟：
1️⃣ 先分析用戶的業務/產品/服務
2️⃣ 提供 2-3 個具體的定位方向選項
3️⃣ 每個選項包含完整6個欄位
4️⃣ 平台推薦專注於台灣用戶常用平台：Instagram Reels、TikTok、YouTube Shorts、Facebook Reels
5️⃣ 提供具體實作建議（基於知識庫的拍攝、剪輯、內容策略）
6️⃣ 最後提供 1-2 個後續問題引導

📝 格式範例：
【🎯 定位選項 A】
📊 業務類型：XXX
👥 目標受眾：XXX  
🎭 品牌語氣：XXX
📱 主要平台：Instagram Reels（台灣用戶最活躍）
🎯 內容目標：XXX
⏰ 發文頻率：XXX

💡 實作建議：
• 具體的內容策略（基於知識庫的流量型/轉換型配比）
• 平台操作要點（拍攝技巧、剪輯節奏、標題鉤子）
• 預期效果

【🎯 定位選項 B】
...

🤔 接下來你可以：
1️⃣ 選擇最適合的定位方向（A/B/C），我會幫你完善細節
2️⃣ 告訴我你的品牌想要傳達什麼形象和語氣？
3️⃣ 你還有其他想了解的定位問題嗎？"""
    
    return context

# 選題智能體
def topic_selection_agent_generate(user_profile: Dict, memories: List[Dict] = None) -> str:
    """選題智能體生成建議"""
    context = f"你是專業的內容選題顧問，為用戶提供每日靈感建議。\n\n"
    
    if user_profile:
        context += f"用戶檔案：\n"
        context += f"- 業務類型：{user_profile.get('business_type', '未設定')}\n"
        context += f"- 目標受眾：{user_profile.get('target_audience', '未設定')}\n"
        context += f"- 品牌語氣：{user_profile.get('brand_voice', '未設定')}\n"
        context += f"- 主要平台：{user_profile.get('primary_platform', '未設定')}\n\n"
    
    if memories:
        context += f"相關洞察：\n"
        for memory in memories[:3]:
            context += f"- {memory['content']}\n"
        context += "\n"
    
    context += """提供5個具體的內容選題建議，每個選題包含：

📝 選題結構：
1️⃣ 標題/主題
2️⃣ 為什麼適合這個用戶
3️⃣ 預期效果
4️⃣ 創作建議
5️⃣ 相關熱門標籤

💡 實作要點：
• 考慮當前熱點、季節性、用戶興趣和平台特性
• 提供具體的拍攝建議
• 包含Hook、Value、CTA結構
• 適合台灣用戶的內容風格

直接輸出選題建議，不要任何開場白或說明文字。"""
    
    return context

def extract_key_insights(text: str, agent_type: str) -> List[str]:
    """從AI回應中提取關鍵洞察"""
    insights = []
    lines = text.split('\n')
    
    for line in lines:
        line = line.strip()
        if len(line) > 20 and any(keyword in line for keyword in ['建議', '應該', '可以', '重點', '關鍵']):
            insights.append(line)
    
    return insights[:3]

# === NEW: 粗略從文字中擷取定位欄位 ===
def extract_profile_fields(text: str) -> Dict[str, Any]:
    """智能擷取定位欄位，從用戶敘述或 AI 回應中抓取定位資訊。"""
    if not text:
        return {}
    t = text.strip()
    import re
    fields: Dict[str, Any] = {}

    # 業務類型 - 更廣泛的匹配
    business_patterns = [
        r"(?:業務類型|行業|產業|做|經營|從事)[:：]\s*([^\n，。,；;]{2,50})",
        r"(?:我是|我們是|公司是|帳號是|專注於|主要做)\s*([^\n，。,；;]{2,50})",
        r"(?:AI智能體|AI自動化|短影音|電商|教育|科技|行銷|內容創作|知識分享)",
    ]
    for pattern in business_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["business_type"] = m.group(1).strip()
            break

    # 目標受眾 - 更智能的匹配
    audience_patterns = [
        r"(?:目標受眾|受眾|TA|觀眾|粉絲)[:：]\s*([^\n]{2,100})",
        r"(?:效率控|職場打工人|科技好奇寶寶|未來生活嚮往者|年輕人|學生|上班族|新手爸媽)",
        r"(?:年齡|性別|職業|興趣|痛點)[:：]\s*([^\n]{2,80})",
    ]
    for pattern in audience_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["target_audience"] = m.group(1).strip()
            break

    # 品牌語氣 - 更廣泛的匹配
    voice_patterns = [
        r"(?:品牌語氣|語氣|口吻|風格)[:：]\s*([^\n，。,；;]{2,50})",
        r"(?:幽默|俏皮|專業|親切|活潑|嚴肅|輕鬆|正式|口語|白話)",
        r"(?:像.*朋友|酷朋友|自然|有記憶點|有共鳴)",
    ]
    for pattern in voice_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["brand_voice"] = m.group(1).strip()
            break

    # 主要平台 - 更智能的匹配
    platform_patterns = [
        r"(?:主要平台|核心平台|平台|在哪裡經營)[:：]\s*([^\n，。,；;]{2,50})",
        r"(?:抖音|小紅書|IG|Instagram|YouTube|Facebook|TikTok|微博|B站)",
    ]
    for pattern in platform_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["primary_platform"] = m.group(1).strip()
            break

    # 內容目標 - 更廣泛的匹配
    goals_patterns = [
        r"(?:內容目標|目標|目的|想要)[:：]\s*([^\n]{2,100})",
        r"(?:轉單|曝光|名單|教育|品牌|流量|粉絲|互動|銷售|推廣)",
    ]
    for pattern in goals_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["content_goals"] = m.group(1).strip()
            break

    # 發文頻率 - 更智能的匹配
    frequency_patterns = [
        r"(?:發文頻率|頻率|多久發|更新)[:：]\s*([^\n，。,；;]{2,30})",
        r"(?:每天|每週|每月|不定期|固定|經常|偶爾)",
    ]
    for pattern in frequency_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["posting_frequency"] = m.group(1).strip()
            break

    return fields

# === NEW: 無模型時的自然回覆（參考資料庫） ===
def natural_fallback_positioning(user_input: str, user_profile: Optional[Dict], memories: List[Dict]) -> str:
    """在沒有外部模型時，根據用戶檔案與記憶，生成比較自然的建議文本。"""
    bp = user_profile or {}
    biz = bp.get("business_type") or "（未設定）"
    aud = bp.get("target_audience") or "（未設定）"
    voice = bp.get("brand_voice") or "（未設定）"
    platform = bp.get("primary_platform") or "（未設定）"

    insights_lines = []
    for m in (memories or [])[:3]:
        insights_lines.append(f"- {m.get('content','').strip()}")
    insights_block = "\n".join(insights_lines) if insights_lines else "（暫無）"

    return (
        "🔍 初步分析（根據已知檔案與你的描述）\n\n"
        f"1) 業務類型：{biz}\n"
        f"2) 目標受眾：{aud}\n"
        f"3) 品牌語氣建議：{voice if voice!='（未設定）' else '先以清晰、可信、口語為主，後續再微調'}\n"
        f"4) 平台策略：優先耕耘 {platform if platform!='（未設定）' else '你最熟悉且受眾集中的平台'}，再輔以次要平台做導流。\n"
        "5) 內容方向：以痛點切入 + 案例/示範 + 明確 CTA。每週固定欄目（例如：教學/開箱/QA/案例）。\n\n"
        "🧠 近期洞察：\n"
        f"{insights_block}\n\n"
        "✅ 下一步：\n"
        "- 告訴我你的產品/服務一句話＋主要受眾＋希望達成的目標（例如：轉單/曝光/名單）\n"
        "- 我會據此補齊定位檔案並給你 2 版內容策略草案"
    )

# ========= 引導式問答 API =========
@app.post("/chat_qa")
async def chat_qa(req: Request):
    data = await req.json()
    session_id = (data.get("session_id") or "qa").strip() or "qa"
    user_msg = (data.get("message") or "").strip()

    # 初次進入：建立 session 並送歡迎 + Q1
    if session_id not in QA_SESSIONS:
        qa_reset(session_id)
        q = qa_next_question(session_id)
        return {
            "session_id": session_id,
            "assistant_message": "嗨👋 讓我們一步步生成你的短影音腳本！\n" + (q or ""),
            "segments": [],
            "done": False,
            "error": None
        }

    # 正常流程：記錄上一題的回答
    qa_record_answer(session_id, user_msg)
    next_q = qa_next_question(session_id)
    if next_q:
        return {
            "session_id": session_id,
            "assistant_message": next_q,
            "segments": [],
            "done": False,
            "error": None
        }

    # 問答完成 → 組合描述 + 取 KB context → 走原有 build_script_prompt
    ans = QA_SESSIONS.get(session_id, {}).get("answers", {})
    brief = compose_brief_from_answers(ans)
    kb_ctx = retrieve_context(brief) or ""
    # 將 QA 選到的 structure/duration 帶入
    template_type = (ans.get("structure") or "").strip()[:1].upper() or None
    try:
        duration = int((ans.get("duration") or "").strip())
    except Exception:
        duration = 30

    user_input = f"{brief}\n\n【KB輔助摘錄】\n{kb_ctx}"

    previous_segments = []
    prompt = build_script_prompt(
        user_input,
        previous_segments,
        template_type=template_type,
        duration=duration,
        dialogue_mode="guide",
    )
    try:
        if use_gemini():
            out = gemini_generate_text(prompt)
            j = _ensure_json_block(out)
            segments = parse_segments(j)
        else:
            segments = fallback_segments(user_input, 0, duration=duration)
    except Exception as e:
        print("[chat_qa] error:", e)
        segments = []

    # 清除 session
    QA_SESSIONS.pop(session_id, None)

    return {
        "session_id": session_id,
        "assistant_message": "我已根據你的回答生成第一版腳本（可再調整）。",
        "segments": segments,
        "done": True,
        "error": None
    }

# ========= /chat_generate =========
@app.post("/chat_generate")
async def chat_generate(req: Request):
    """
    body: {
      user_id?: str,
      session_id?: str,
      messages: [{role, content}],
      previous_segments?: [segment...],
      remember?: bool,
      mode?: "script" | "copy",          # ← 保留既有：腳本/文案
      topic?: str,                        # ← 文案主題（可選）
      dialogue_mode?: "guide" | "free",   # ← 新增：引導/自由 對話風格（可選）
      template_type?: "A"|"B"|"C"|"D"|"E"|"F",  # ← 新增
      duration?: 30|60,                   # ← 新增：腳本時長
      knowledge_hint?: str                # ← 新增：檢索提示詞（可選）
    }
    """
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_id = (data.get("user_id") or "").strip() or get_anon_user_id(req)
    messages = data.get("messages") or []
    previous_segments = data.get("previous_segments") or []
    topic = (data.get("topic") or "").strip() or None

    explicit_mode = (data.get("mode") or "").strip().lower() or None
    mode = detect_mode(messages, explicit=explicit_mode)

    # NEW: 讀取新參數（後端若沒收到也不影響舊行為）
    dialogue_mode = (data.get("dialogue_mode") or "").strip().lower() or None
    template_type = (data.get("template_type") or "").strip().upper() or None
    try:
        duration = int(data.get("duration")) if data.get("duration") is not None else None
    except Exception:
        duration = None
    knowledge_hint = (data.get("knowledge_hint") or "").strip() or None

    user_input = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            user_input = (m.get("content") or "").strip()
            break

    # 若輸入過短，也直接嘗試生成（避免制式提示打斷對話）
    hint = SHORT_HINT_COPY if mode == "copy" else SHORT_HINT_SCRIPT
    if len(user_input) < 6:
        user_input = f"（使用者提示較短）請主動追問必要資訊並先給出初步建議。\n提示：{user_input or '請先幫我開始'}"

    try:
        if mode == "copy":
            prompt = build_copy_prompt(user_input, topic)
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                copy = parse_copy(j)
            else:
                copy = fallback_copy(user_input, topic)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "我先給你第一版完整貼文（可再加要求，我會幫你改得更貼近風格）。",
                "segments": [],
                "copy": copy,
                "error": None
            }

        else:  # script
            prompt = build_script_prompt(
                user_input,
                previous_segments,
                template_type=template_type,
                duration=duration,
                dialogue_mode=dialogue_mode,
                knowledge_hint=knowledge_hint,
            )
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                segments = parse_segments(j)
            else:
                segments = fallback_segments(user_input, len(previous_segments or []), duration=duration)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "我先給你第一版完整腳本（可再加要求，我會幫你改得更貼近風格）。",
                "segments": segments,
                "copy": None,
                "error": None
            }

        # DB 紀錄（保留原行為）
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO requests (user_id, user_input, mode, messages_json, previous_segments_json, response_json) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    user_id,
                    user_input, mode,
                    json.dumps(messages, ensure_ascii=False),
                    json.dumps(previous_segments, ensure_ascii=False),
                    json.dumps(resp, ensure_ascii=False),
                ),
            )
            conn.commit()
            conn.close()
        except Exception as e:
            print("[DB] insert failed:", e)

        return resp

    except Exception as e:
        print("[chat_generate] error:", e)
        return JSONResponse(status_code=500, content={
            "session_id": data.get("session_id") or "s",
            "assistant_message": "伺服器忙碌，稍後再試",
            "segments": [],
            "copy": None,
            "error": "internal_server_error"
        })

# ========= 舊流程：/generate_script =========
@app.post("/generate_script")
async def generate_script(req: Request):
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_input = (data.get("user_input") or "").strip()
    previous_segments = data.get("previous_segments") or []

    # 向下相容：舊端點若想支援 60s/模板，也可帶入這兩個欄位（可選）
    template_type = (data.get("template_type") or "").strip().upper() or None
    try:
        duration = int(data.get("duration")) if data.get("duration") is not None else None
    except Exception:
        duration = None

    if len(user_input) < 6:
        return {"segments": [], "error": SHORT_HINT_SCRIPT}

    try:
        prompt = build_script_prompt(
            user_input,
            previous_segments,
            template_type=template_type,
            duration=duration
        )
        if use_gemini():
            out = gemini_generate_text(prompt)
            j = _ensure_json_block(out)
            segments = parse_segments(j)
        else:
            segments = fallback_segments(user_input, len(previous_segments or []), duration=duration)
        return {"segments": segments, "error": None}
    except Exception as e:
        print("[generate_script] error:", e)
        return JSONResponse(status_code=500, content={"segments": [], "error": "internal_server_error"})

# ========= 匯出：Word 暫停 / Excel 保留 =========
@app.post("/export/docx")
async def export_docx_disabled():
    return JSONResponse(status_code=501, content={"error": "docx_export_disabled"})

def _ensure_xlsx():
    try:
        import openpyxl  # noqa
        return True
    except Exception:
        return False

@app.post("/export/xlsx")
async def export_xlsx(req: Request):
    if not _ensure_xlsx():
        return JSONResponse(status_code=501, content={"error": "xlsx_not_available"})
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = await req.json()
    segments = data.get("segments") or []
    copy = data.get("copy") or None

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "腳本分段"
    ws1.append(["#","type","start_sec","end_sec","camera","dialog","visual","cta"])
    for i, s in enumerate(segments, 1):
        ws1.append([i, s.get("type"), s.get("start_sec"), s.get("end_sec"),
                    s.get("camera"), s.get("dialog"), s.get("visual"), s.get("cta")])

    ws2 = wb.create_sheet("文案")
    ws2.append(["主貼文"]); ws2.append([copy.get("main_copy") if copy else ""])
    ws2.append([]); ws2.append(["備選開頭"])
    for a in (copy.get("alternates") if copy else []) or []: ws2.append([a])
    ws2.append([]); ws2.append(["Hashtags"])
    ws2.append([" ".join(copy.get("hashtags") if copy else [])])
    ws2.append([]); ws2.append(["CTA"])
    ws2.append([copy.get("cta") if copy else ""])
    ws2.append([]); ws2.append(["圖片建議"])
    for idea in (copy.get("image_ideas") if copy else []) or []: ws2.append([idea])

    for ws in (ws1, ws2):
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value else 0 for c in col) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 80)

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="export.xlsx"'}
    )

# ========= CSV 下載 & Google Sheet 連動 =========
import csv
import json
from fastapi.responses import FileResponse, Response
from io import StringIO

@app.get("/download/requests_export.csv")
def download_requests_csv():
    export_path = "/data/requests_export.csv"
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM requests ORDER BY id DESC")
    rows = cur.fetchall()
    headers = [desc[0] for desc in cur.description]
    conn.close()

    with open(export_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return FileResponse(
        export_path,
        media_type="text/csv",
        filename="requests_export.csv",
    )


@app.get("/export/google-sheet")
def export_for_google_sheet(limit: int = 100):
    try:
        limit = int(limit)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 2000))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"SELECT id, created_at, user_input, mode FROM requests ORDER BY id DESC LIMIT {limit}"
    )
    rows = cur.fetchall()
    conn.close()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "created_at", "user_input", "mode"])
    for row in rows:
        writer.writerow(row)

    return Response(content=output.getvalue(), media_type="text/csv")


@app.get("/export/google-sheet-flat")
def export_google_sheet_flat(limit: int = 200):
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT id, created_at, user_input, mode, response_json
        FROM requests
        ORDER BY id DESC
        LIMIT {limit}
        """
    )
    rows = cur.fetchall()
    conn.close()

    out = StringIO()
    writer = csv.writer(out)

    headers = [
        "id", "created_at", "mode", "user_input",
        "assistant_message",
        "copy_main_copy",
        "copy_cta",
        "copy_hashtags",
        "copy_alternates_joined",
        "segments_count",
        "seg1_type", "seg1_start_sec", "seg1_end_sec", "seg1_dialog", "seg1_visual", "seg1_cta",
        "seg2_type", "seg2_start_sec", "seg2_end_sec", "seg2_dialog", "seg2_visual", "seg2_cta",
        "seg3_type", "seg3_start_sec", "seg3_end_sec", "seg3_dialog", "seg3_visual", "seg3_cta",
    ]
    writer.writerow(headers)

    for _id, created_at, user_input, mode, resp_json in rows:
        assistant_message = ""
        copy_main = ""
        copy_cta = ""
        copy_hashtags = ""
        copy_alternates_joined = ""
        segments_count = 0

        def empty_seg():
            return ["", "", "", "", "", ""]
        seg1 = empty_seg()
        seg2 = empty_seg()
        seg3 = empty_seg()

        try:
            data = json.loads(resp_json or "{}")
            assistant_message = (data.get("assistant_message") or "")[:500]

            c = data.get("copy") or {}
            if isinstance(c, dict):
                copy_main = c.get("main_copy") or ""
                copy_cta = c.get("cta") or ""
                tags = c.get("hashtags") or []
                if isinstance(tags, list):
                    copy_hashtags = " ".join(map(str, tags))
                alts = c.get("alternates") or c.get("openers") or []
                if isinstance(alts, list):
                    copy_alternates_joined = " | ".join(map(str, alts))

            segs = data.get("segments") or []
            if isinstance(segs, list):
                segments_count = len(segs)

                def to_seg(s):
                    return [
                        s.get("type", ""),
                        s.get("start_sec", ""),
                        s.get("end_sec", ""),
                        s.get("dialog", ""),
                        s.get("visual", ""),
                        s.get("cta", ""),
                    ]

                if len(segs) >= 1: seg1 = to_seg(segs[0])
                if len(segs) >= 2: seg2 = to_seg(segs[1])
                if len(segs) >= 3: seg3 = to_seg(segs[2])

        except Exception as e:
            assistant_message = f"[JSON parse error] {str(e)}"

        writer.writerow([
            _id, created_at, mode, user_input,
            assistant_message,
            copy_main,
            copy_cta,
            copy_hashtags,
            copy_alternates_joined,
            segments_count,
            *seg1, *seg2, *seg3,
        ])

    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "inline; filename=export_flat.csv"},
    )

# ========= Google Sheet 扁平化（v2） =========
import csv
import json
from io import StringIO
from fastapi.responses import Response

@app.get("/export/google-sheet-flat-v2")
def export_google_sheet_flat_v2(limit: int = 200):
    """
    扁平化 CSV（含 copy 與前 3 個 segments），禁用快取。
    在 Google Sheets 使用：
      =IMPORTDATA("https://aijobvideobackend.zeabur.app/export/google-sheet-flat-v2?limit=500")
    """
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT id, created_at, user_input, mode, response_json
        FROM requests
        ORDER BY id DESC
        LIMIT {limit}
        """
    )
    rows = cur.fetchall()
    conn.close()

    out = StringIO()
    writer = csv.writer(out)

    headers = [
        "id", "created_at", "mode", "user_input",
        "assistant_message",
        "copy_main_copy", "copy_cta", "copy_hashtags", "copy_alternates_joined",
        "segments_count",
        "seg1_type", "seg1_start_sec", "seg1_end_sec", "seg1_dialog", "seg1_visual", "seg1_cta",
        "seg2_type", "seg2_start_sec", "seg2_end_sec", "seg2_dialog", "seg2_visual", "seg2_cta",
        "seg3_type", "seg3_start_sec", "seg3_end_sec", "seg3_dialog", "seg3_visual", "seg3_cta",
    ]
    writer.writerow(headers)

    def empty_seg():
        return ["", "", "", "", "", ""]

    for _id, created_at, user_input, mode, resp_json in rows:
        assistant_message = ""
        copy_main = ""
        copy_cta = ""
        copy_hashtags = ""
        copy_alternates = ""
        segments_count = 0
        seg1 = empty_seg()
        seg2 = empty_seg()
        seg3 = empty_seg()

        try:
            data = json.loads(resp_json or "{}")
            assistant_message = (data.get("assistant_message") or "")[:500]

            c = data.get("copy") or {}
            if isinstance(c, dict):
                copy_main = c.get("main_copy") or ""
                copy_cta = c.get("cta") or ""
                tags = c.get("hashtags") or []
                if isinstance(tags, list):
                    copy_hashtags = " ".join(map(str, tags))
                alts = c.get("alternates") or c.get("openers") or []
                if isinstance(alts, list):
                    copy_alternates = " | ".join(map(str, alts))

            segs = data.get("segments") or []
            if isinstance(segs, list):
                segments_count = len(segs)

                def to_seg(s):
                    return [
                        s.get("type", ""),
                        s.get("start_sec", ""),
                        s.get("end_sec", ""),
                        s.get("dialog", ""),
                        s.get("visual", ""),
                        s.get("cta", ""),
                    ]

                if len(segs) >= 1: seg1 = to_seg(segs[0])
                if len(segs) >= 2: seg2 = to_seg(segs[1])
                if len(segs) >= 3: seg3 = to_seg(segs[2])

        except Exception as e:
            assistant_message = f"[JSON parse error] {str(e)}"

        writer.writerow([
            _id, created_at, mode, user_input,
            assistant_message,
            copy_main, copy_cta, copy_hashtags, copy_alternates,
            segments_count,
            *seg1, *seg2, *seg3,
        ])

    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": "inline; filename=export_flat_v2.csv",
            "Cache-Control": "no-store, max-age=0, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )

# ========= Admin APIs（簡易狀態與用戶列表） =========
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN")  # 可選，若未設置則不驗證

def _check_admin(req: Request):
    # 先看是否有有效 admin session cookie
    adm_cookie = req.cookies.get("admin_session")
    if adm_cookie and verify_admin_session_cookie(adm_cookie):
        return True
    # 其次允許 token（自動化工具/備援）
    tok = req.headers.get("x-admin-token") or req.query_params.get("token")
    if ADMIN_TOKEN and tok == ADMIN_TOKEN:
        return True
    return False

@app.get("/admin/users")
async def admin_users(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    users = conn.execute("SELECT user_id, email, name, created_at, updated_at, status FROM users ORDER BY created_at DESC LIMIT 500").fetchall()
    auths = conn.execute(
        """
        SELECT user_id, username, email, phone, created_at,
               CASE WHEN password_hash IS NOT NULL AND length(password_hash)>0 THEN 1 ELSE 0 END AS has_password
        FROM users_auth ORDER BY created_at DESC LIMIT 500
        """
    ).fetchall()
    conn.close()
    return {
        "users": [dict(u) for u in users],
        "users_auth": [dict(a) for a in auths],
    }

@app.get("/admin/users_full")
async def admin_users_full(req: Request, limit: int = 500):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 500
    limit = max(1, min(limit, 2000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    users = conn.execute("SELECT * FROM users ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
    auths = conn.execute(
        """
        SELECT user_id, username, email, phone, created_at,
               CASE WHEN password_hash IS NOT NULL AND length(password_hash)>0 THEN 1 ELSE 0 END AS has_password
        FROM users_auth ORDER BY created_at DESC LIMIT ?
        """,
        (limit,)
    ).fetchall()
    credits = conn.execute("SELECT * FROM user_credits").fetchall()
    orders = conn.execute("SELECT * FROM orders ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return {
        "users": [dict(u) for u in users],
        "users_auth": [dict(a) for a in auths],
        "credits": [dict(c) for c in credits],
        "orders": [dict(o) for o in orders],
    }

@app.get("/admin/usage")
async def admin_usage(req: Request, limit: int = 30):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 30
    limit = max(1, min(limit, 500))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    total_requests = conn.execute("SELECT COUNT(1) AS c FROM requests").fetchone()["c"]
    latest = conn.execute(
        f"SELECT id, created_at, user_input, mode FROM requests ORDER BY id DESC LIMIT {limit}"
    ).fetchall()
    conn.close()
    return {
        "total_requests": total_requests,
        "latest": [dict(r) for r in latest],
    }

@app.get("/admin/users.csv")
async def admin_users_csv(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    users = conn.execute("SELECT user_id, email, name, created_at, updated_at, status FROM users ORDER BY created_at DESC").fetchall()
    auths = conn.execute("SELECT user_id, username, email AS auth_email, phone, created_at AS auth_created FROM users_auth ORDER BY created_at DESC").fetchall()
    conn.close()

    from io import StringIO
    s = StringIO()
    import csv
    w = csv.writer(s)
    w.writerow(["user_id","email","name","created_at","updated_at","status","username","auth_email","phone","auth_created"])
    # 以 user_id 關聯（此處簡化：逐筆合併，若無對應則留空）
    auth_map = {a["user_id"]: a for a in auths}
    for u in users:
        a = auth_map.get(u["user_id"]) or {}
        w.writerow([
            u["user_id"], u["email"], u["name"], u["created_at"], u["updated_at"], u["status"],
            a.get("username",""), a.get("auth_email",""), a.get("phone",""), a.get("auth_created",""),
        ])
    return Response(content=s.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=users.csv"})

@app.get("/admin/usage.csv")
async def admin_usage_csv(req: Request, limit: int = 1000):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 1000
    limit = max(1, min(limit, 5000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"SELECT id, created_at, user_input, mode FROM requests ORDER BY id DESC LIMIT {limit}"
    ).fetchall()
    conn.close()

    from io import StringIO
    s = StringIO()
    import csv
    w = csv.writer(s)
    w.writerow(["id","created_at","mode","user_input"])
    for r in rows:
        w.writerow([r["id"], r["created_at"], r["mode"], r["user_input"]])
    return Response(content=s.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=usage.csv"})

@app.get("/admin/users_auth")
async def admin_users_auth(req: Request, limit: int = 1000):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 1000
    limit = max(1, min(limit, 5000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"SELECT id, user_id, username, email, phone, created_at, updated_at FROM users_auth ORDER BY created_at DESC LIMIT {limit}"
    ).fetchall()
    conn.close()
    return {"users_auth": [dict(r) for r in rows]}

@app.post("/admin/user/reset_password")
async def admin_reset_password(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    data = await req.json()
    user_id = (data.get("user_id") or "").strip()
    identifier = (data.get("identifier") or "").strip()  # username 或 email
    new_password = (data.get("new_password") or "").strip()
    if not new_password or len(new_password) < 6:
        return JSONResponse(status_code=400, content={"error": "weak_password", "message": "密碼至少 6 碼"})
    if not user_id and not identifier:
        return JSONResponse(status_code=400, content={"error": "missing_identifier"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        if user_id:
            row = conn.execute("SELECT id, user_id, username, email FROM users_auth WHERE user_id=?", (user_id,)).fetchone()
        else:
            row = conn.execute("SELECT id, user_id, username, email FROM users_auth WHERE username=? OR email=?", (identifier, identifier)).fetchone()
        if not row:
            conn.close()
            return JSONResponse(status_code=404, content={"error": "user_not_found"})
        conn.execute("UPDATE users_auth SET password_hash=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (hash_password(new_password), row["id"]))
        # 稽核記錄
        try:
            _tok = req.headers.get("x-admin-token") or req.query_params.get("token") or ""
            admin_hash = hashlib.sha256(_tok.encode("utf-8")).hexdigest() if _tok else None
            conn.execute(
                "INSERT INTO admin_audit_logs (action, admin_token_hash, target_user_id, details) VALUES (?, ?, ?, ?)",
                ("reset_password", admin_hash, row["user_id"], json.dumps({"username": row["username"], "email": row["email"]}, ensure_ascii=False))
            )
        except Exception as _e:
            print("[audit] write failed:", _e)
        conn.commit(); conn.close()
        return {"ok": True}
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.post("/admin/user/add_credits")
async def admin_add_credits(req: Request):
    """管理員為用戶充值點數"""
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    
    data = await req.json()
    user_id = (data.get("user_id") or "").strip()
    identifier = (data.get("identifier") or "").strip()  # username 或 email
    credits = data.get("credits", 0)
    reason = (data.get("reason") or "管理員充值").strip()

    if not identifier and not user_id:
        return JSONResponse(status_code=400, content={"error": "missing_fields", "message": "請提供用戶ID或email"})
    
    if credits <= 0:
        return JSONResponse(status_code=400, content={"error": "invalid_credits", "message": "充值點數必須大於0"})

    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        # 查找用戶 - 先查 users_auth，再查 users
        if user_id:
            # 直接使用user_id查找
            row = conn.execute(
                "SELECT user_id, username, email FROM users_auth WHERE user_id = ?",
                (user_id,)
            ).fetchone()
            
            # 如果在 users_auth 中找不到，嘗試在 users 表中查找
            if not row:
                row = conn.execute(
                    "SELECT user_id, name as username, email FROM users WHERE user_id = ?",
                    (user_id,)
                ).fetchone()
        else:
            # 使用identifier查找
            row = conn.execute(
                "SELECT user_id, username, email FROM users_auth WHERE username = ? OR email = ?",
                (identifier, identifier)
            ).fetchone()
            
            # 如果在 users_auth 中找不到，嘗試在 users 表中查找
            if not row:
                row = conn.execute(
                    "SELECT user_id, name as username, email FROM users WHERE name = ? OR email = ?",
                    (identifier, identifier)
                ).fetchone()
        
        if not row:
            return JSONResponse(status_code=404, content={"error": "user_not_found", "message": "找不到指定的用戶"})

        user_id = row["user_id"]
        
        # 獲取當前點數餘額
        credit_row = conn.execute(
            "SELECT balance FROM user_credits WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        
        current_balance = credit_row["balance"] if credit_row else 0
        new_balance = current_balance + credits

        # 更新或插入點數記錄
        conn.execute(
            """INSERT INTO user_credits (user_id, balance, updated_at) 
               VALUES (?, ?, CURRENT_TIMESTAMP) 
               ON CONFLICT(user_id) DO UPDATE SET 
               balance = ?, updated_at = CURRENT_TIMESTAMP""",
            (user_id, new_balance, new_balance)
        )

        # 記錄訂單
        conn.execute(
            """INSERT INTO orders (user_id, order_type, amount, status, description, created_at) 
               VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)""",
            (user_id, "admin_credit", credits, "completed", reason)
        )

        # 稽核記錄
        try:
            _tok = req.headers.get("x-admin-token") or req.query_params.get("token") or ""
            admin_hash = hashlib.sha256(_tok.encode("utf-8")).hexdigest() if _tok else None
            conn.execute(
                "INSERT INTO admin_audit_logs (action, admin_token_hash, target_user_id, details) VALUES (?, ?, ?, ?)",
                ("add_credits", admin_hash, user_id, json.dumps({
                    "username": row["username"], 
                    "email": row["email"],
                    "credits_added": credits,
                    "old_balance": current_balance,
                    "new_balance": new_balance,
                    "reason": reason
                }, ensure_ascii=False))
            )
        except Exception as _e:
            print("[audit] write failed:", _e)
        
        conn.commit(); conn.close()
        return {
            "ok": True, 
            "user_id": user_id,
            "credits_added": credits,
            "old_balance": current_balance,
            "new_balance": new_balance
        }

    except Exception as e:
        conn.close()
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/admin/user/{user_id}/credits")
async def admin_get_user_credits(user_id: str, req: Request):
    """查看用戶點數餘額"""
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        # 獲取用戶信息
        user_row = conn.execute(
            "SELECT user_id, username, email FROM users_auth WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        
        if not user_row:
            return JSONResponse(status_code=404, content={"error": "user_not_found"})

        # 獲取點數餘額
        credit_row = conn.execute(
            "SELECT balance, updated_at FROM user_credits WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        
        # 獲取訂單記錄
        orders = conn.execute(
            "SELECT * FROM orders WHERE user_id = ? ORDER BY created_at DESC LIMIT 20",
            (user_id,)
        ).fetchall()
        
        conn.close()
        
        return {
            "user_id": user_id,
            "username": user_row["username"],
            "email": user_row["email"],
            "balance": credit_row["balance"] if credit_row else 0,
            "updated_at": credit_row["updated_at"] if credit_row else None,
            "orders": [dict(o) for o in orders]
        }

    except Exception as e:
        conn.close()
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/api/plans")
async def get_subscription_plans():
    """獲取訂閱方案列表"""
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        plans = conn.execute(
            "SELECT * FROM subscription_plans WHERE is_active = 1 ORDER BY price ASC"
        ).fetchall()
        conn.close()
        return {"plans": [dict(p) for p in plans]}
    except Exception as e:
        conn.close()
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.post("/api/purchase")
async def purchase_plan(req: Request):
    """購買訂閱方案"""
    try:
        data = await req.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "invalid_json"})
    
    user_id = (data.get("user_id") or "").strip()
    plan_id = data.get("plan_id")
    payment_method = (data.get("payment_method") or "manual").strip()
    
    if not user_id or not plan_id:
        return JSONResponse(status_code=400, content={"error": "missing_fields"})
    
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        # 獲取方案信息
        plan = conn.execute(
            "SELECT * FROM subscription_plans WHERE id = ? AND is_active = 1",
            (plan_id,)
        ).fetchone()
        
        if not plan:
            return JSONResponse(status_code=404, content={"error": "plan_not_found"})
        
        # 創建訂單
        order_id = conn.execute(
            """INSERT INTO orders (user_id, order_type, amount, plan, status, payment_method, created_at, paid_at) 
               VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)""",
            (user_id, "subscription", plan["credits"], plan["name"], "paid", payment_method)
        ).lastrowid
        
        # 充值點數
        credit_row = conn.execute(
            "SELECT balance FROM user_credits WHERE user_id = ?",
            (user_id,)
        ).fetchone()
        
        current_balance = credit_row["balance"] if credit_row else 0
        new_balance = current_balance + plan["credits"]
        
        conn.execute(
            """INSERT INTO user_credits (user_id, balance, updated_at) 
               VALUES (?, ?, CURRENT_TIMESTAMP) 
               ON CONFLICT(user_id) DO UPDATE SET 
               balance = ?, updated_at = CURRENT_TIMESTAMP""",
            (user_id, new_balance, new_balance)
        )
        
        # 記錄訂閱
        conn.execute(
            """INSERT INTO user_subscriptions (user_id, plan_id, status, start_date, end_date) 
               VALUES (?, ?, ?, CURRENT_TIMESTAMP, datetime('now', '+{} days'))""",
            (user_id, plan_id, "active", plan["duration_days"])
        )
        
        conn.commit(); conn.close()
        
        return {
            "ok": True,
            "order_id": order_id,
            "plan_name": plan["name"],
            "credits_added": plan["credits"],
            "new_balance": new_balance
        }
        
    except Exception as e:
        conn.close()
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/admin/requests_full")
async def admin_requests_full(req: Request, limit: int = 200, user_id: str | None = None, mode: str | None = None, date_from: str | None = None, date_to: str | None = None):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    conditions = []
    params = []
    if user_id:
        conditions.append("user_id = ?"); params.append(user_id)
    if mode:
        conditions.append("mode = ?"); params.append(mode)
    if date_from:
        conditions.append("date(created_at) >= date(?)"); params.append(date_from)
    if date_to:
        conditions.append("date(created_at) <= date(?)"); params.append(date_to)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = conn.execute(
        f"SELECT id, created_at, user_id, mode, user_input, response_json FROM requests {where} ORDER BY id DESC LIMIT {limit}",
        params,
    ).fetchall()
    conn.close()
    return {"items": [dict(r) for r in rows]}

@app.get("/admin/messages")
async def admin_messages(req: Request, user_id: str | None = None, session_id: str | None = None, limit: int = 200):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        conditions = []
        params = []
        if user_id:
            conditions.append("s.user_id = ?"); params.append(user_id)
        if session_id:
            conditions.append("m.session_id = ?"); params.append(session_id)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        rows = conn.execute(
            f"""
            SELECT m.id, m.session_id, s.user_id, s.agent_type, m.role, m.content, m.timestamp
            FROM messages m
            LEFT JOIN sessions s ON s.session_id = m.session_id
            {where}
            ORDER BY m.id DESC
            LIMIT {limit}
            """,
            params,
        ).fetchall()
        conn.close()
        return {"items": [dict(r) for r in rows]}
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/admin/analytics")
async def admin_analytics(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        total_users = conn.execute("SELECT COUNT(1) AS c FROM users").fetchone()["c"]
        total_requests = conn.execute("SELECT COUNT(1) AS c FROM requests").fetchone()["c"]
        # 今日請求
        today = conn.execute("SELECT COUNT(1) AS c FROM requests WHERE date(created_at) = date('now','localtime')").fetchone()["c"]
        # 近7日
        last7 = conn.execute(
            """
            SELECT strftime('%Y-%m-%d', created_at) AS d, COUNT(1) AS c
            FROM requests
            WHERE date(created_at) >= date('now','localtime','-6 day')
            GROUP BY d ORDER BY d ASC
            """
        ).fetchall()
        last7d = [{"date": r["d"], "count": r["c"]} for r in last7]
        # 模式分佈
        by_mode_rows = conn.execute("SELECT COALESCE(mode,'') AS mode, COUNT(1) AS c FROM requests GROUP BY COALESCE(mode,'')").fetchall()
        by_mode = { (r["mode"] or ""): r["c"] for r in by_mode_rows }
        # agent 分佈（sessions）
        by_agent_rows = conn.execute("SELECT agent_type, COUNT(1) AS c FROM sessions GROUP BY agent_type").fetchall()
        by_agent = { r["agent_type"]: r["c"] for r in by_agent_rows }
        # 近7日 agent 使用次數（依 sessions.created_at）
        agent_daily_rows = conn.execute(
            """
            SELECT strftime('%Y-%m-%d', created_at) AS d, agent_type, COUNT(1) AS c
            FROM sessions
            WHERE date(created_at) >= date('now','localtime','-6 day')
            GROUP BY d, agent_type
            ORDER BY d ASC
            """
        ).fetchall()
        agent_daily = {}
        for r in agent_daily_rows:
            agent_daily.setdefault(r["d"], {})[r["agent_type"]] = r["c"]
        # 訊息總數/今日
        total_messages = conn.execute("SELECT COUNT(1) AS c FROM messages").fetchone()["c"]
        today_messages = conn.execute("SELECT COUNT(1) AS c FROM messages WHERE date(timestamp) = date('now','localtime')").fetchone()["c"]
        conn.close()
        return {
            "total_users": total_users,
            "total_requests": total_requests,
            "today_requests": today,
            "last7d": last7d,
            "by_mode": by_mode,
            "by_agent": by_agent,
            "total_messages": total_messages,
            "today_messages": today_messages,
            "agent_daily": agent_daily,
        }
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(admin_session: str | None = Cookie(default=None)):
    if not (admin_session and verify_admin_session_cookie(admin_session)):
        # 簡易登入頁
        return HTMLResponse(content="""
<!DOCTYPE html>
<html lang=\"zh-Hant\"><head><meta charset=\"utf-8\"/><meta name=\"viewport\" content=\"width=device-width, initial-scale=1\"/><title>AIJob 管理登入</title>
<style>body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,'Noto Sans TC',sans-serif;margin:40px;color:#111;background:#f6f7fb}
.card{max-width:360px;margin:0 auto;border:1px solid #e5e7eb;border-radius:12px;background:#fff;padding:16px}
input,button{width:100%;padding:10px;border:1px solid #cbd5e1;border-radius:8px;margin-top:10px}
button{background:#111;color:#fff}
.muted{color:#6b7280;font-size:12px;margin-top:8px}
</style></head><body>
<div class=\"card\"><h2>AIJob 管理登入</h2>
<input id=\"u\" placeholder=\"帳號\"><input id=\"p\" placeholder=\"密碼\" type=\"password\">
<button onclick=\"login()\">登入</button>
<div class=\"muted\">僅限管理者使用。登入後將建立安全的管理 Session。</div></div>
<script>
async function login(){
  const r = await fetch('/admin/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:document.getElementById('u').value,password:document.getElementById('p').value})});
  const j = await r.json(); if(j&&j.ok){ location.href='/admin'; } else { alert(j.message||'登入失敗'); }
}
</script></body></html>
""", status_code=200)
    
    # 讀取 admin.html 檔案
    try:
        # 在 Docker 容器中，admin 資料夾在 /app/admin/
        admin_html_path = '/app/admin/admin.html'
        with open(admin_html_path, 'r', encoding='utf-8') as f:
            admin_html_content = f.read()
        return HTMLResponse(content=admin_html_content)
    except Exception as e:
        # 如果檔案讀取失敗，返回錯誤頁面
        return HTMLResponse(content=f"""
<!DOCTYPE html>
<html lang=\"zh-Hant\">
<head><meta charset=\"utf-8\"/><title>管理後台錯誤</title></head>
<body>
<h1>管理後台載入錯誤</h1>
<p>無法載入管理後台檔案: {str(e)}</p>
<p>請檢查 admin.html 檔案是否存在於正確位置。</p>
</body>
</html>
""", status_code=500)

# === Admin Login/Logout ===
@app.post("/admin/login")
async def admin_login(req: Request):
    try:
        data = await req.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "bad_request"})
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not ADMIN_USER or not ADMIN_PASSWORD:
        return JSONResponse(status_code=500, content={"error": "admin_not_configured", "message": "尚未設定 ADMIN_USER/ADMIN_PASSWORD"})
    if username != ADMIN_USER or password != ADMIN_PASSWORD:
        return JSONResponse(status_code=401, content={"error": "invalid_credentials", "message": "帳號或密碼錯誤"})
    token = create_admin_session_cookie(username)
    resp = JSONResponse({"ok": True})
    # Cookie 屬性：HttpOnly+Secure+SameSite=None，存活 5 小時
    resp.set_cookie("admin_session", token, httponly=True, secure=True, samesite="none", max_age=5*3600)
    return resp

@app.post("/admin/logout")
async def admin_logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("admin_session")
    return resp

@app.get("/admin/healthz")
async def admin_healthz(req: Request):
    has_session = False
    try:
        c = req.cookies.get("admin_session")
        has_session = bool(c and verify_admin_session_cookie(c))
    except Exception:
        has_session = False
    return {
        "ok": True,
        "admin_ready": bool(ADMIN_USER and ADMIN_PASSWORD),
        "oauth_ready": bool('_OAUTH_READY' in globals() and _OAUTH_READY),
        "has_admin_session": has_session,
    }

# ========= 三智能體 API 端點 =========
# 統一聊天端點（自然對談 + KB + 記憶 + 人設）
AGENT_PERSONAS = {
    "positioning": (
        "你是專業的短影音定位顧問。所有回覆必須優先結合已知知識庫(KB)與用戶檔案，避免空泛內容。"
        "與用戶對談請採『少量輸出 + 反問引導』的節奏，一次只推進 1~2 個重點，"
        "並聚焦在：業務類型定位、目標受眾、品牌形象定位、平台策略建議、內容目標設定、發文頻率。"
        "回覆須具體、可執行、含金量高。"
    ),
    "topics": (
        "你是專業的爆款短影音選題顧問。優先根據 KB 與用戶定位，提供可直接實作的選題建議，"
        "避免大眾化冗長清單，必要時反問 1 個關鍵條件再給 3~5 條具體選題。"
    ),
    "script": (
        "你是專業的短影音腳本寫手。優先根據 KB 與用戶檔案產出可拍攝的分段腳本，"
        "不足時先以 1~2 句反問補足關鍵條件再生成，保持精煉、可落地。"
    ),
}

def _mem_agent_key(agent_type: str) -> str:
    if agent_type == "positioning":
        return "positioning"
    if agent_type == "topics":
        return "topic_selection"
    return "script_copy"

@app.post("/chat")
async def chat(req: Request):
    """統一聊天：自然對談，帶入用戶檔案/記憶/知識庫。"""
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_id = (data.get("user_id") or "").strip()
    agent_type = (data.get("agent_type") or "script").strip()
    messages = data.get("messages") or []
    template_type = (data.get("template_type") or "").strip().upper() or None
    duration = data.get("duration")

    if not user_id:
        raise HTTPException(status_code=400, detail="user_id is required")

    # 確保用戶存在
    create_or_get_user(user_id)

    # 讀取檔案與記憶
    user_profile = get_user_profile(user_id)
    memories_all = get_user_memories(user_id, agent_type=_mem_agent_key(agent_type), limit=20)

    # 建會話
    session_id = data.get("session_id") or create_session(user_id, agent_type)

    # 將最近一則 user 訊息加入訊息表
    last_user = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            last_user = (m.get("content") or "").strip()
            break
    if last_user:
        add_message(session_id, "user", last_user)

    # 人設與 KB ground
    persona = AGENT_PERSONAS.get(agent_type, AGENT_PERSONAS["script"])
    kb_ctx = retrieve_context(last_user) if last_user else ""
    kb_all = (EXTRA_KB or "").strip()

    # 可選：把模板/時長附加到上下文
    script_hint = ""
    if agent_type == "script":
        if template_type:
            script_hint += f"\n【指定模板】{template_type}"
        if duration:
            try:
                script_hint += f"\n【指定時長】{int(duration)} 秒"
            except Exception:
                pass

    system_ctx = (
        f"{persona}\n請以自然中文對談，不用制式清單。若能從知識庫或用戶檔案得到答案，請優先結合。\n\n"
        f"【重要格式要求】\n"
        f"• 使用emoji作為分點符號，讓內容更易讀；每次最多 5~8 行\n"
        f"• 優先給出可執行建議，若條件不足先反問 1~2 個關鍵問題\n"
        f"• 基於知識庫內容提供專業建議\n"
        f"• 回應結構：📝 主要觀點 → 💡 具體建議 → ✨ 實作要點 → 🎯 行動指引\n\n"
        f"【用戶檔案（若空代表未設定）】\n{json.dumps(user_profile or {}, ensure_ascii=False)}\n\n"
        f"【相關記憶（節選）】\n" + "\n".join([f"- {m.get('content','')}" for m in memories_all[:5]]) + "\n\n"
        f"【全域知識摘要（截斷）】\n{kb_all[:1200]}\n\n"
        f"【KB動態擷取】\n{(kb_ctx or '')[:800]}\n" 
        f"{script_hint}\n"
    )

    # 產生回覆
    if use_gemini():
        prompt = (
            system_ctx + "\n---\n" + (last_user or "") + "\n\n請以對談形式回覆，避免重覆使用相同句型。使用emoji分段，讓內容更易讀。"
        )
        ai_response = gemini_generate_text(prompt)
    else:
        # 無模型的自然回覆（較快）
        if agent_type == "positioning":
            ai_response = natural_fallback_positioning(last_user, user_profile, memories_all)
        elif agent_type == "topics":
            base = last_user or "請提供今日的選題靈感"
            ai_response = (
                "以下是依你的定位與近期洞察給的選題方向（可回我要哪個展開）：\n\n"
                "1) 熱點＋你產品的關聯切入\n"
                "2) 受眾常見痛點的快速解法\n"
                "3) 使用前/後對比案例\n"
                "4) 30 秒微教學 + 行動呼籲\n"
                "5) 迷你訪談/QA 回覆留言\n\n"
                f"你剛提到：{base[:80]}… 我建議先從 2) 或 4) 開始。"
            )
        else:  # script
            ai_response = (
                "了解，我會用自然口吻陪你討論腳本。先說明你的主題、平台與目標，我再給你第一版結構與開場。"
            )

    add_message(session_id, "assistant", ai_response)

    # 嘗試抽取並更新定位檔案（只針對定位）
    if agent_type == "positioning":
        try:
            draft = {}
            draft.update(extract_profile_fields(last_user))
            draft.update(extract_profile_fields(ai_response))
            draft = {k:v for k,v in draft.items() if v}
            if draft:
                update_user_profile(user_id, draft)
                user_profile = get_user_profile(user_id)
        except Exception as e:
            print("[/chat] profile extract failed:", e)

    result_obj = {
        "session_id": session_id,
        "assistant_message": ai_response,
        "user_profile": user_profile if agent_type == "positioning" else None,
        "error": None
    }
    if agent_type == "positioning" and 'positioning_summary_text' in locals() and positioning_summary_text:
        result_obj["positioning_summary"] = positioning_summary_text
    return result_obj

# === NEW: 流式聊天端點 ===
from fastapi import BackgroundTasks

@app.post("/chat_stream")
async def chat_stream(req: Request):
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_id = (data.get("user_id") or "").strip() or get_anon_user_id(req)
    agent_type = (data.get("agent_type") or "script").strip()
    messages = data.get("messages") or []
    template_type = (data.get("template_type") or "").strip().upper() or None
    duration = data.get("duration")

    if not user_id:
        raise HTTPException(status_code=400, detail="user_id is required")

    create_or_get_user(user_id)
    user_profile = get_user_profile(user_id)
    memories_all = get_user_memories(user_id, agent_type=_mem_agent_key(agent_type), limit=20)

    session_id = data.get("session_id") or create_session(user_id, agent_type)

    last_user = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            last_user = (m.get("content") or "").strip()
            break
    if last_user:
        add_message(session_id, "user", last_user)

    persona = AGENT_PERSONAS.get(agent_type, AGENT_PERSONAS["script"])
    kb_ctx = retrieve_context(last_user) if last_user else ""
    kb_all = (EXTRA_KB or "").strip()
    script_hint = ""
    if agent_type == "script":
        if template_type:
            script_hint += f"\n【指定模板】{template_type}"
        if duration:
            try:
                script_hint += f"\n【指定時長】{int(duration)} 秒"
            except Exception:
                pass

    system_ctx = (
        f"{persona}\n請以自然中文對談，不用制式清單。若能從知識庫或用戶檔案得到答案，請優先結合。\n" 
        f"【用戶檔案（若空代表未設定）】\n{json.dumps(user_profile or {}, ensure_ascii=False)}\n\n"
        f"【相關記憶（節選）】\n" + "\n".join([f"- {m.get('content','')}" for m in memories_all[:5]]) + "\n\n"
        f"【全域知識摘要（截斷）】\n{kb_all[:1200]}\n\n"
        f"【KB動態擷取】\n{(kb_ctx or '')[:800]}\n" 
        f"{script_hint}\n"
    )

    # 取得最近對話以增強上下文連貫
    def get_recent_messages(session_id: str, limit: int = 8):
        try:
            conn = get_conn()
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT role, content FROM messages WHERE session_id = ? ORDER BY id DESC LIMIT ?",
                (session_id, limit),
            ).fetchall()
            conn.close()
            return list(reversed([dict(r) for r in rows]))
        except Exception:
            return []

    recent_msgs = get_recent_messages(session_id, 8)

    async def gen():
        # 簡易切片流：若有模型可逐段送出，否則一次送出自然回覆
        if use_gemini():
            convo = "\n".join([f"{m['role']}: {m['content']}" for m in recent_msgs])
            full = gemini_generate_text(system_ctx + "\n---\n" + (convo or (last_user or "")))
        else:
            if agent_type == "positioning":
                full = natural_fallback_positioning(last_user, user_profile, memories_all)
            elif agent_type == "topics":
                base = last_user or "請提供今日的選題靈感"
                full = (
                    "以下是依你的定位與近期洞察給的選題方向（可回我要哪個展開）：\n\n"
                    "1) 熱點＋你產品的關聯切入\n"
                    "2) 受眾常見痛點的快速解法\n"
                    "3) 使用前/後對比案例\n"
                    "4) 30 秒微教學 + 行動呼籲\n"
                    "5) 迷你訪談/QA 回覆留言\n\n"
                    f"你剛提到：{base[:80]}… 我建議先從 2) 或 4) 開始。"
                )
            else:
                full = "了解，我會用自然口吻陪你討論腳本。先說明你的主題、平台與目標，我再給你第一版結構與開場。"

        # 逐段輸出
        chunk_size = 60
        for i in range(0, len(full), chunk_size):
            yield full[i:i+chunk_size]
        # 完成後寫入訊息
        add_message(session_id, "assistant", full)

        # 定位：嘗試更新檔案並把回覆摘要存成筆記
        if agent_type == "positioning":
            try:
                draft = {}
                draft.update(extract_profile_fields(last_user))
                draft.update(extract_profile_fields(full))
                draft = {k:v for k,v in draft.items() if v}
                if draft:
                    update_user_profile(user_id, draft)
                # 存成「note」型記憶，供前端右側筆記本顯示
                note = (full or "").strip()
                if note:
                    add_memory(user_id, "positioning", "note", note[:800], importance_score=6)
            except Exception:
                pass
        # 選題：把回覆存成筆記並保存選題建議
        elif agent_type == "topics":
            try:
                note = (full or "").strip()
                if note:
                    add_memory(user_id, "topic_selection", "note", note[:800], importance_score=6)
                    
                    # 保存選題建議到資料庫
                    from datetime import date
                    conn = get_conn()
                    conn.execute(
                        """INSERT OR REPLACE INTO topic_suggestions 
                           (user_id, suggested_date, topics, reasoning) 
                           VALUES (?, ?, ?, ?)""",
                        (user_id, date.today().isoformat(), json.dumps({"suggestions": note}), note)
                    )
                    conn.commit()
                    conn.close()
            except Exception as e:
                print(f"[Topics Save Error] {e}")
                pass
        # 腳本：把回覆存成筆記
        elif agent_type == "script":
            try:
                note = (full or "").strip()
                if note:
                    add_memory(user_id, "script_copy", "note", note[:800], importance_score=6)
            except Exception:
                pass

    return StreamingResponse(gen(), media_type="text/plain")


# 定位智能體
@app.post("/agent/positioning/analyze")
async def positioning_analyze(req: Request):
    """定位智能體分析用戶定位"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        user_input = data.get("user_input", "")
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # 確保用戶存在
        create_or_get_user(user_id)
        
        # 獲取用戶檔案和相關記憶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="positioning", limit=10)
        
        # 創建會話
        session_id = create_session(user_id, "positioning")
        add_message(session_id, "user", user_input)
        
        # 生成分析
        analysis_context = positioning_agent_analyze(user_input, user_profile, memories)
        
        # 調用 AI 生成回應（無模型時提供自然回覆）
        if use_gemini():
            ai_response = gemini_generate_text(analysis_context)
        else:
            ai_response = natural_fallback_positioning(user_input, user_profile, memories)
        
        add_message(session_id, "assistant", ai_response)
        
        # 提取關鍵洞察並保存為記憶
        if ai_response and len(ai_response) > 50:
            key_insights = extract_key_insights(ai_response, "positioning")
            for insight in key_insights:
                add_memory(user_id, "positioning", "insight", insight, importance_score=7)

        # 從用戶輸入與 AI 回應中嘗試擷取定位欄位，更新檔案（草稿）
        try:
            draft_fields = {}
            draft_fields.update(extract_profile_fields(user_input))
            draft_fields.update(extract_profile_fields(ai_response))
            # 過濾空值
            draft_fields = {k:v for k,v in draft_fields.items() if v}
            if draft_fields:
                update_user_profile(user_id, draft_fields)
                # 重新讀取最新檔案
                user_profile = get_user_profile(user_id)
        except Exception as _e:
            print("[Positioning] extract_profile_fields failed:", _e)
        
        # 生成結構化的定位摘要（包含執行建議）
        positioning_summary = ""
        tone_guidelines = ""
        execution_suggestions = ""
        
        if ai_response:
            # 簡單解析AI回應，提取關鍵信息
            lines = ai_response.split('\n')
            for line in lines:
                line = line.strip()
                if '業務類型：' in line or '目標受眾：' in line or '品牌語氣：' in line:
                    positioning_summary += line + "\n"
                elif '語氣' in line and ('專業' in line or '親切' in line or '幽默' in line or '權威' in line):
                    tone_guidelines = line
                elif '實作建議' in line or '執行' in line or '建議' in line:
                    execution_suggestions += line + "\n"
        
        # 如果沒有提取到足夠信息，使用默認值
        if not positioning_summary:
            positioning_summary = "基於您的描述，建議建立專業的短影音定位策略。"
        if not tone_guidelines:
            tone_guidelines = "使用專業術語，保持客觀理性，強調數據和事實。"
        if not execution_suggestions:
            execution_suggestions = "建議採用流量型與轉換型內容 7:3 配比，每週發布 3-5 次，專注於 Instagram Reels 平台。"
        
        return {
            "session_id": session_id,
            "response": ai_response,
            "user_profile": user_profile,
            "positioning_summary": positioning_summary,
            "tone_guidelines": tone_guidelines,
            "execution_suggestions": execution_suggestions,
            "error": None
        }
        
    except Exception as e:
        print(f"[Positioning Agent Error] {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={
                "error": "internal_server_error", 
                "message": "伺服器內部錯誤，請稍後再試",
                "details": str(e) if "DEBUG" in os.environ else "Internal server error"
            }
        )

@app.put("/agent/positioning/profile")
async def update_positioning_profile(req: Request):
    """更新用戶定位檔案"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        profile_data = data.get("profile_data", {})
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # 確保用戶存在
        create_or_get_user(user_id)
        
        # 更新檔案
        success = update_user_profile(user_id, profile_data)
        
        if success:
            # 保存檔案更新為記憶
            add_memory(user_id, "positioning", "profile_update", 
                      f"用戶檔案已更新：{json.dumps(profile_data, ensure_ascii=False)}", 
                      importance_score=8)
        
        return {
            "success": success,
            "message": "檔案更新成功" if success else "檔案更新失敗",
            "error": None
        }
        
    except Exception as e:
        print(f"[Profile Update Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# 新增：取得用戶定位檔案與筆記（供前端右側同步顯示）
@app.get("/agent/positioning/profile")
async def get_positioning_profile(user_id: str, notes_limit: int = 10):
    try:
        profile = get_user_profile(user_id)
        notes = get_user_memories(user_id, agent_type="positioning", memory_type="note", limit=notes_limit)
        return {
            "user_id": user_id,
            "profile": profile or {},
            "notes": notes,
            "error": None
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

# 新增：通用筆記查詢 API
@app.get("/agent/notes")
async def get_agent_notes(user_id: str, agent_type: str, memory_type: str = "note", limit: int = 10):
    try:
        notes = get_user_memories(user_id, agent_type=agent_type, memory_type=memory_type, limit=limit)
        return {
            "user_id": user_id,
            "agent_type": agent_type,
            "notes": notes,
            "error": None
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

# 新增：一鍵生成定位功能
@app.post("/agent/positioning/generate")
async def generate_positioning(req: Request):
    """一鍵生成完整定位檔案"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        theme = data.get("theme", "")  # 用戶提供的主題/產品/服務
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        if not theme.strip():
            raise HTTPException(status_code=400, detail="theme is required")
        
        # 確保用戶存在
        create_or_get_user(user_id)
        
        # 獲取現有檔案和記憶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="positioning", limit=5)
        
        # 創建會話
        session_id = create_session(user_id, "positioning")
        add_message(session_id, "user", f"一鍵生成定位：{theme}")
        
        # 構建一鍵生成提示詞
        context = f"""你是專業的短影音定位顧問，專門服務台灣市場，請根據用戶提供的主題「{theme}」生成完整的定位檔案。

請分析這個主題並提供：
1. 業務類型：具體的行業分類
2. 目標受眾：明確的台灣受眾畫像（年齡、職業、痛點、需求）
3. 品牌語氣：適合台灣用戶的溝通風格
4. 主要平台：台灣最適合的短影音平台（推薦：Instagram Reels、TikTok、YouTube Shorts、Facebook Reels）
5. 內容目標：具體要達成的效果
6. 發文頻率：建議的更新頻率

【重要】平台推薦請專注於台灣用戶常用的平台，避免推薦B站、小紅書等大陸平台。

請以結構化格式回應，每個欄位都要具體明確，便於系統自動提取。

格式：
業務類型：[具體分類]
目標受眾：[詳細描述]
品牌語氣：[風格特點]
主要平台：[平台名稱]
內容目標：[具體目標]
發文頻率：[頻率建議]"""
        
        # 調用 AI 生成定位
        if use_gemini():
            ai_response = gemini_generate_text(context)
        else:
            # 無模型時的範例回覆
            ai_response = f"""根據「{theme}」主題，我為你生成以下定位：

業務類型：{theme}相關服務
目標受眾：對{theme}有興趣的台灣潛在客戶
品牌語氣：專業親切
主要平台：Instagram Reels
內容目標：建立專業形象，吸引潛在客戶
發文頻率：每週2-3次"""
        
        add_message(session_id, "assistant", ai_response)
        
        # 提取定位欄位並更新檔案
        extracted_fields = extract_profile_fields(ai_response)
        if extracted_fields:
            update_user_profile(user_id, extracted_fields)
            # 重新讀取最新檔案
            user_profile = get_user_profile(user_id)
        
        # 保存 AI 回應為筆記
        if ai_response and len(ai_response) > 50:
            add_memory(user_id, "positioning", "note", ai_response, importance_score=8)
        
        return {
            "session_id": session_id,
            "response": ai_response,
            "user_profile": user_profile,
            "extracted_fields": extracted_fields,
            "error": None
        }
        
    except Exception as e:
        print(f"[Generate Positioning Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# 選題智能體
@app.post("/agent/topics/suggest")
async def topic_suggest(req: Request):
    """獲取選題建議"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        target_date = data.get("target_date")  # YYYY-MM-DD 格式
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # 確保用戶存在
        create_or_get_user(user_id)
        
        # 解析日期
        if target_date:
            try:
                from datetime import datetime
                target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
            except ValueError:
                from datetime import date
                target_date = date.today()
        else:
            from datetime import date
            target_date = date.today()
        
        # 獲取用戶檔案和相關記憶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="topic_selection", limit=5)
        
        # 生成選題建議
        suggestion_context = topic_selection_agent_generate(user_profile, memories)
        
        # 調用 AI 生成選題
        if use_gemini():
            ai_response = gemini_generate_text(suggestion_context)
        else:
            ai_response = "AI服務暫時不可用，請稍後再試。"
        
        # 保存選題建議
        conn = get_conn()
        conn.execute(
            """INSERT OR REPLACE INTO topic_suggestions 
               (user_id, suggested_date, topics, reasoning) 
               VALUES (?, ?, ?, ?)""",
            (user_id, target_date.isoformat(), json.dumps({"suggestions": ai_response}), ai_response)
        )
        conn.commit()
        conn.close()
        
        return {
            "user_id": user_id,
            "suggested_date": target_date.isoformat(),
            "suggestions": ai_response,
            "reasoning": ai_response,
            "error": None
        }
        
    except Exception as e:
        print(f"[Topic Selection Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

@app.get("/agent/topics/history")
async def topic_history(user_id: str, limit: int = 10):
    """獲取選題歷史"""
    try:
        conn = get_conn()
        conn.row_factory = sqlite3.Row
        suggestions = conn.execute(
            "SELECT * FROM topic_suggestions WHERE user_id = ? ORDER BY suggested_date DESC LIMIT ?",
            (user_id, limit)
        ).fetchall()
        conn.close()
        
        return {
            "user_id": user_id,
            "suggestions": [dict(s) for s in suggestions],
            "error": None
        }
        
    except Exception as e:
        print(f"[Topic History Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# 腳本文案智能體（增強版）
@app.post("/agent/content/generate")
async def content_generate(req: Request):
    """生成腳本或文案（增強版，整合記憶系統）"""
    try:
        data = await req.json()
        user_id = data.get("user_id") or get_anon_user_id(req)
        user_input = data.get("user_input", "")
        mode = data.get("mode", "script")  # "script" 或 "copy"
        template_type = data.get("template_type")
        duration = data.get("duration")
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # 確保用戶存在
        create_or_get_user(user_id)
        
        # 獲取用戶檔案和相關記憶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="script_copy", limit=10)
        
        # 創建會話
        session_id = create_session(user_id, "script_copy")
        add_message(session_id, "user", user_input)
        
        # 構建增強的提示詞（整合用戶檔案和記憶）
        enhanced_input = user_input
        
        if user_profile:
            profile_context = f"""
【用戶定位檔案】
- 業務類型：{user_profile.get('business_type', '未設定')}
- 目標受眾：{user_profile.get('target_audience', '未設定')}
- 品牌語氣：{user_profile.get('brand_voice', '未設定')}
- 主要平台：{user_profile.get('primary_platform', '未設定')}
"""
            enhanced_input = f"{profile_context}\n\n用戶需求：{user_input}"
        
        if memories:
            memory_context = "\n【相關記憶】\n"
            for memory in memories[:3]:
                memory_context += f"- {memory['content']}\n"
            enhanced_input = f"{enhanced_input}\n\n{memory_context}"
        
        # 使用現有的 chat_generate 邏輯，但傳入增強後的輸入
        enhanced_data = {
            "user_id": user_id,
            "session_id": session_id,
            "messages": [{"role": "user", "content": enhanced_input}],
            "mode": mode,
            "template_type": template_type,
            "duration": duration
        }
        
        # 調用現有的生成邏輯
        result = await chat_generate_internal(enhanced_data)
        
        # 添加記憶
        if result.get("assistant_message"):
            add_memory(user_id, "script_copy", "generation", 
                      f"生成{mode}：{user_input[:100]}...", 
                      importance_score=6)
        
        return result
        
    except Exception as e:
        print(f"[Content Generation Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# 記憶系統 API
@app.get("/memory/user/{user_id}")
async def get_user_memory(user_id: str, agent_type: str = None, memory_type: str = None, limit: int = 20):
    """獲取用戶記憶"""
    try:
        memories = get_user_memories(user_id, agent_type, memory_type, limit)
        
        return {
            "user_id": user_id,
            "memories": memories,
            "count": len(memories),
            "error": None
        }
        
    except Exception as e:
        print(f"[Memory Retrieval Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

@app.post("/memory/add")
async def add_memory_endpoint(req: Request):
    """添加記憶"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        agent_type = data.get("agent_type")
        memory_type = data.get("memory_type")
        content = data.get("content")
        importance_score = data.get("importance_score", 5)
        tags = data.get("tags", [])
        
        if not all([user_id, agent_type, memory_type, content]):
            raise HTTPException(status_code=400, detail="Missing required fields")
        
        memory_id = add_memory(user_id, agent_type, memory_type, content, importance_score, tags)
        
        return {
            "memory_id": memory_id,
            "message": "記憶添加成功",
            "error": None
        }
        
    except Exception as e:
        print(f"[Memory Addition Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# 新增：一鍵生成腳本功能
@app.post("/agent/script/generate")
async def generate_script_one_click(req: Request):
    """一鍵生成腳本功能"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        theme = data.get("theme", "")  # 用戶提供的主題/文字
        template_type = data.get("template_type", "A")  # 預設三段式
        duration = data.get("duration", 30)  # 預設30秒
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        if not theme.strip():
            raise HTTPException(status_code=400, detail="theme is required")
        
        # 確保用戶存在
        create_or_get_user(user_id)
        
        # 獲取用戶檔案和相關記憶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="script_copy", limit=10)
        
        # 創建會話
        session_id = create_session(user_id, "script_copy")
        add_message(session_id, "user", f"一鍵生成腳本：{theme}")
        
        # 構建一鍵生成提示詞
        context = f"""根據主題「{theme}」生成短影音腳本。

🎯 腳本參數：
• 模板：{template_type} - {TEMPLATE_GUIDE.get(template_type, "三段式")}
• 時長：{duration} 秒
• 平台：Instagram Reels、TikTok、YouTube Shorts、Facebook Reels

📚 知識庫：
{BUILTIN_KB_SCRIPT}

💡 台灣市場特色：
• 內容風格：生活化、親切、實用
• 節奏要求：2-3秒換畫面，節奏緊湊
• Hook原則：0-5秒直給結論，用大字卡與強情緒表情
• 語氣：堅定、直給結論，避免口癖贅字

直接輸出JSON格式，不要任何開場白或說明文字：

{{
  "segments":[
    {{"type":"hook","start_sec":0,"end_sec":5,"camera":"CU","dialog":"...","visual":"...","cta":""}},
    {{"type":"value","start_sec":5,"end_sec":25,"camera":"MS","dialog":"...","visual":"...","cta":""}},
    {{"type":"cta","start_sec":25,"end_sec":30,"camera":"WS","dialog":"...","visual":"...","cta":"..."}}
  ]
}}"""
        
        # 調用 AI 生成腳本
        if use_gemini():
            ai_response = gemini_generate_text(context)
        else:
            # 無模型時的範例回覆
            ai_response = f"""{{
  "segments":[
    {{"type":"hook","start_sec":0,"end_sec":5,"camera":"CU","dialog":"你知道{theme}的秘密嗎？","visual":"大字卡+驚訝表情","cta":""}},
    {{"type":"value","start_sec":5,"end_sec":25,"camera":"MS","dialog":"今天我要分享{theme}的實用技巧，讓你輕鬆掌握！","visual":"示範畫面","cta":""}},
    {{"type":"cta","start_sec":25,"end_sec":30,"camera":"WS","dialog":"想要更多{theme}技巧，記得關注我！","visual":"關注按鈕","cta":"點關注"}}
  ]
}}"""
        
        add_message(session_id, "assistant", ai_response)
        
        # 解析腳本
        try:
            if use_gemini():
                segments = parse_segments(ai_response)
            else:
                # 解析範例回覆
                import json
                data = json.loads(ai_response)
                segments = data.get("segments", [])
        except Exception as e:
            print(f"[Script Parse Error] {e}")
            segments = []
        
        # 保存腳本生成為筆記
        if ai_response and len(ai_response) > 50:
            add_memory(user_id, "script_copy", "note", ai_response, importance_score=8)
        
        return {
            "session_id": session_id,
            "assistant_message": "🚀 一鍵生成完成！我為你生成了完整的腳本。",
            "segments": segments,
            "error": None
        }
        
    except Exception as e:
        print(f"[One-Click Script Generation Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# 內部函數：chat_generate 的內部邏輯（供 content_generate 調用）
async def chat_generate_internal(data: dict):
    """chat_generate 的內部邏輯，供其他函數調用"""
    user_id = (data.get("user_id") or "").strip() or "web"
    messages = data.get("messages") or []
    previous_segments = data.get("previous_segments") or []
    topic = (data.get("topic") or "").strip() or None

    explicit_mode = (data.get("mode") or "").strip().lower() or None
    mode = detect_mode(messages, explicit=explicit_mode)

    dialogue_mode = (data.get("dialogue_mode") or "").strip().lower() or None
    template_type = (data.get("template_type") or "").strip().upper() or None
    try:
        duration = int(data.get("duration")) if data.get("duration") is not None else None
    except Exception:
        duration = None
    knowledge_hint = (data.get("knowledge_hint") or "").strip() or None

    user_input = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            user_input = (m.get("content") or "").strip()
            break

    # 輸入過短時，仍持續對話而非回傳制式提示
    hint = SHORT_HINT_COPY if mode == "copy" else SHORT_HINT_SCRIPT
    if len(user_input) < 6:
        user_input = f"（使用者提示較短）請主動追問關鍵條件並先給出方向性建議。\n提示：{user_input or '開始'}"

    try:
        if mode == "copy":
            prompt = build_copy_prompt(user_input, topic)
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                copy = parse_copy(j)
            else:
                copy = fallback_copy(user_input, topic)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "我先給你第一版完整貼文（可再加要求，我會幫你改得更貼近風格）。",
                "segments": [],
                "copy": copy,
                "error": None
            }

        else:  # script
            prompt = build_script_prompt(
                user_input,
                previous_segments,
                template_type=template_type,
                duration=duration,
                dialogue_mode=dialogue_mode,
                knowledge_hint=knowledge_hint,
            )
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                segments = parse_segments(j)
            else:
                segments = fallback_segments(user_input, len(previous_segments or []), duration=duration)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "我先給你第一版完整腳本（可再加要求，我會幫你改得更貼近風格）。",
                "segments": segments,
                "copy": None,
                "error": None
            }

        return resp

    except Exception as e:
        print("[chat_generate_internal] error:", e)
        return {
            "session_id": data.get("session_id") or "s",
            "assistant_message": "伺服器忙碌，稍後再試",
            "segments": [],
            "copy": None,
            "error": "internal_server_error"
        }

# 匿名用戶 ID（未登入時避免跨裝置/跨 IP 互相污染記憶）
def get_anon_user_id(req: Request) -> str:
    try:
        ip = (req.client.host if req and req.client else '0.0.0.0')
        ua = (req.headers.get('user-agent') or 'ua')[:40]
        h = hashlib.sha256(f"{ip}|{ua}".encode('utf-8')).hexdigest()[:16]
        from datetime import date
        d = date.today().isoformat()
        return f"anon_{h}_{d}"
    except Exception:
        return "anon_web"

# 依目前問題挑選最相關記憶，避免回覆偏離當下上下文
def select_relevant_memories(query: str, memories: list[dict], k: int = 5) -> list[dict]:
    try:
        if not memories:
            return []
        q = (query or '').strip()
        if not q:
            return memories[:k]
        import re
        toks = [t for t in re.split(r"[\s，。；、,.:?!\-\/\[\]()]+", q) if len(t) >= 2]
        toks = list(dict.fromkeys(toks))
        scored = []
        for m in memories:
            txt = (m.get('content') or '').lower()
            score = 0
            for t in toks:
                if t and t.lower() in txt:
                    score += 1
            # 額外加權：較新的/較重要的
            score = score * 10 + int(m.get('importance_score') or 0)
            scored.append((score, m))
        scored.sort(key=lambda x: -x[0])
        return [m for _, m in scored[:k]]
    except Exception:
        return memories[:k]

# ========= 點數系統整合 =========
try:
    from .points_integration import integrate_points_system
    integrate_points_system(app)
    print("✅ AI Points System integrated successfully")
except ImportError as e:
    print(f"⚠️  AI Points System not available: {e}")
except Exception as e:
    print(f"❌ Failed to integrate AI Points System: {e}")

# 啟動服務器
if __name__ == "__main__":
    import uvicorn
    print("🚀 啟動三智能體系統...")
    print("📍 本地訪問：http://localhost:8080")
    print("📋 API 文檔：http://localhost:8080/docs")
    uvicorn.run(app, host="0.0.0.0", port=8080, log_level="info")