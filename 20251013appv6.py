# app.py
import os
import json
import glob
import sqlite3
from typing import List, Optional, Any, Dict
from datetime import datetime, date

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from fastapi.responses import JSONResponse, HTMLResponse, Response, StreamingResponse, RedirectResponse
from fastapi import Cookie
from itsdangerous import URLSafeSerializer, BadSignature
import hashlib
# from fastapi.staticfiles import StaticFiles  # å‰ç«¯åˆ†é›¢éƒ¨ç½²ï¼Œä¸éœ€è¦

# ========= ç’°å¢ƒè®Šæ•¸ =========
DB_PATH = os.getenv("DB_PATH", "/data/three_agents_system.db")
GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY") or os.getenv("GOOGLE_API_KEY")
KNOWLEDGE_TXT_PATH = os.getenv("KNOWLEDGE_TXT_PATH", "/data/data/kb.txt")
GLOBAL_KB_TEXT = ""
SESSION_SECRET = os.getenv("SESSION_SECRET", "change-me-session-secret")
session_signer = URLSafeSerializer(SESSION_SECRET, salt="session")
admin_session_signer = URLSafeSerializer(SESSION_SECRET, salt="admin_session")

# Admin å¸³è™Ÿï¼ˆè«‹ä»¥ç’°å¢ƒè®Šæ•¸è¨­å®šï¼‰
ADMIN_USER = os.getenv("ADMIN_USER")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD")

# Google OAuth2ï¼ˆå¯é¸ï¼‰
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET")
OAUTH_REDIRECT_URI = os.getenv("OAUTH_REDIRECT_URI", "https://aijobvideobackend.zeabur.app/auth/google/callback")
try:
    from authlib.integrations.starlette_client import OAuth
    oauth = OAuth()
    if GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET:
        oauth.register(
            name="google",
            server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
            client_id=GOOGLE_CLIENT_ID,
            client_secret=GOOGLE_CLIENT_SECRET,
            client_kwargs={"scope": "openid email profile"},
        )
    _OAUTH_READY = True
except Exception as _e:
    print("[OAuth] not enabled:", _e)
    oauth = None
    _OAUTH_READY = False

# ========= App èˆ‡ CORS =========
app = FastAPI(title="Three AI Agents System with Long-term Memory")

# å‹•æ…‹è¨­å®š CORSï¼šè‹¥éœ€è¦å¸¶ Cookie å°±ä¸èƒ½ä½¿ç”¨ "*"
# é è¨­ç™½åå–®åŒ…å« GitHub Pagesã€Zeabur å‰ç«¯å­ç¶²åŸŸã€æ­£å¼ç«™å­ç¶²åŸŸèˆ‡æœ¬æ©Ÿ
ORIGINS = os.getenv(
    "ALLOWED_ORIGINS",
    "https://jacky6658.github.io,https://jacky6658.github.io/Altest/,http://localhost:3000,https://video.aijob.com.tw,https://aijobvideo.zeabur.app,https://aijob.com.tw"
).split(",")
ORIGINS = [o.strip().rstrip("/") for o in ORIGINS if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ORIGINS,
    allow_credentials=True,
    allow_methods=["POST", "OPTIONS", "GET", "PUT", "DELETE"],
    allow_headers=["*"],
)

# OAuth éœ€è¦ Starlette SessionMiddlewareï¼›ä½¿ç”¨ç¨ç«‹ cookie åç¨±é¿å…èˆ‡æœ¬ç³»çµ± session æ··æ·†
app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    session_cookie="oauth_session",
    same_site="none",
    https_only=True,
)

# å‰ç«¯åˆ†é›¢éƒ¨ç½²ï¼Œä¸éœ€è¦éœæ…‹æ–‡ä»¶æœå‹™

# ========= å¼•å°å¼å•ç­”ç‹€æ…‹ï¼ˆè¨˜æ†¶é«”æš«å­˜ï¼‰ =========
QA_SESSIONS: Dict[str, Dict[str, Any]] = {}  # key: session_id
QA_QUESTIONS = [
    {"key":"structure","q":"ã€Q1ã€‘è«‹é¸æ“‡è…³æœ¬çµæ§‹ï¼ˆA ä¸‰æ®µå¼ / B å•é¡Œè§£æ±º / C Before-After / D æ•™å­¸ / E æ•˜äº‹ / F çˆ†é»é€£ç™¼ï¼‰"},
    {"key":"duration","q":"ã€Q2ã€‘å½±ç‰‡æ™‚é•·ï¼ˆ30 æˆ– 60 ç§’ï¼‰"},
    {"key":"topic","q":"ã€Q3ã€‘è«‹è¼¸å…¥ä¸»é¡Œæˆ–ç”¢å“åç¨±"},
    {"key":"goal","q":"ã€Q4ã€‘ä¸»è¦ç›®æ¨™ï¼ˆå¸æµé‡ / æ•™è‚² / è½‰å–® / å“ç‰Œï¼‰"},
    {"key":"audience","q":"ã€Q5ã€‘ç›®æ¨™å—çœ¾ï¼ˆå¹´é½¡/æ€§åˆ¥/ç‰¹è³ª/ç—›é»ï¼‰"},
    {"key":"hook","q":"ã€Q6ã€‘é–‹å ´é‰¤å­é¡å‹ï¼ˆå•å¥/åå·®/åŒç†/æ•¸å­—ï¼‰ï¼‹æƒ³æ”¾çš„é—œéµè©"},
    {"key":"cta","q":"ã€Q7ã€‘CTAï¼ˆé—œæ³¨/æ”¶è— / ç•™è¨€/ç§è¨Š / è³¼è²·é€£çµï¼‰"}
]

def qa_reset(session_id: str):
    QA_SESSIONS[session_id] = {"step": 0, "answers": {}}

def qa_next_question(session_id: str) -> Optional[str]:
    st = QA_SESSIONS.get(session_id)
    if not st: return None
    step = st["step"]
    if step < len(QA_QUESTIONS):
        return QA_QUESTIONS[step]["q"]
    return None

def qa_record_answer(session_id: str, user_text: str):
    st = QA_SESSIONS.get(session_id)
    if not st: return
    step = st["step"]
    if step < len(QA_QUESTIONS):
        key = QA_QUESTIONS[step]["key"]
        st["answers"][key] = user_text
        st["step"] = step + 1

def compose_brief_from_answers(ans: Dict[str,str]) -> str:
    labels = {
        "structure":"çµæ§‹","duration":"æ™‚é•·","topic":"ä¸»é¡Œ","goal":"ç›®æ¨™","audience":"å—çœ¾",
        "hook":"é‰¤å­","cta":"CTA"
    }
    lines = []
    for it in QA_QUESTIONS:
        k = it["key"]
        if k in ans:
            lines.append(f"{labels.get(k,k)}ï¼š{ans[k]}")
    return "ï¼›".join(lines)

# ========= ç°¡æ˜“ KB æª¢ç´¢ =========
def load_kb_text() -> str:
    path = KNOWLEDGE_TXT_PATH
    try:
        if os.path.exists(path):
            with open(path,"r",encoding="utf-8") as f:
                return f.read()
    except Exception:
        pass
    return ""

def retrieve_context(query: str, max_chars: int = 1200) -> str:
    text = GLOBAL_KB_TEXT or ""
    if not text: 
        return ""
    import re
    toks = [t for t in re.split(r'[\sï¼Œã€‚ï¼›ã€,.:?!\-\/\[\]()]+', (query or "")) if len(t)>=1]
    toks = list(dict.fromkeys(toks))
    lines = text.splitlines()
    scored = []
    for i, line in enumerate(lines):
        score = sum(1 for t in toks if t and t in line)
        if score>0:
            scored.append((score, i, line))
    scored.sort(key=lambda x:(-x[0], x[1]))
    selected=[]
    total=0
    for _, _, ln in scored:
        if not ln.strip(): 
            continue
        take = ln.strip()
        if total + len(take) + 1 > max_chars:
            break
        selected.append(take)
        total += len(take) + 1
    if not selected:
        return text[:max_chars]
    return "\n".join(selected)

# ========= DB =========
def _ensure_db_dir(path: str):
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

def get_conn() -> sqlite3.Connection:
    _ensure_db_dir(DB_PATH)
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    _ensure_db_dir(DB_PATH)
    conn = get_conn()
    cur = conn.cursor()
    
    # åŸæœ‰è¡¨æ ¼
    cur.execute("""
        CREATE TABLE IF NOT EXISTS requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            user_id TEXT,
            user_input TEXT,
            mode TEXT,
            messages_json TEXT,
            previous_segments_json TEXT,
            response_json TEXT
        )
    """)
    try:
        cur.execute("ALTER TABLE requests ADD COLUMN user_id TEXT")
    except Exception:
        pass
    
    # æ–°å¢ï¼šä¸‰æ™ºèƒ½é«”ç³»çµ±è¡¨æ ¼
    # 1. ç”¨æˆ¶åŸºæœ¬è³‡è¨Šè¡¨
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id TEXT PRIMARY KEY,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            email TEXT,
            name TEXT,
            platform_preferences TEXT,
            language_preference TEXT DEFAULT 'zh-TW',
            timezone TEXT DEFAULT 'Asia/Taipei',
            status TEXT DEFAULT 'active'
        )
    """)

    # æ–°å¢ï¼šEmail/å¸³è™Ÿç™»å…¥è¡¨ï¼ˆæœ¬åœ°å¸³è™Ÿï¼‰
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users_auth (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            username TEXT UNIQUE,
            email TEXT,
            phone TEXT,
            password_hash TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(user_id)
        )
    """)
    
    # 2. ç”¨æˆ¶å®šä½æª”æ¡ˆè¡¨
    cur.execute("""
        CREATE TABLE IF NOT EXISTS user_profiles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            business_type TEXT,
            target_audience TEXT,
            brand_voice TEXT,
            content_goals TEXT,
            primary_platform TEXT,
            secondary_platforms TEXT,
            posting_frequency TEXT,
            preferred_topics TEXT,
            content_styles TEXT,
            video_duration_preference TEXT,
            competitors TEXT,
            unique_value_proposition TEXT,
            current_followers INTEGER DEFAULT 0,
            engagement_rate REAL DEFAULT 0.0,
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            UNIQUE(user_id)
        )
    """)
    
    # 3. æœƒè©±è¨˜éŒ„è¡¨
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            session_id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            agent_type TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'active',
            context_summary TEXT,
            key_insights TEXT,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    """)
    
    # 4. å°è©±è¨˜éŒ„è¡¨
    cur.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id TEXT NOT NULL,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            metadata TEXT,
            FOREIGN KEY (session_id) REFERENCES sessions(session_id)
        )
    """)
    
    # 5. æ™ºèƒ½é«”è¨˜æ†¶è¡¨
    cur.execute("""
        CREATE TABLE IF NOT EXISTS agent_memories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            agent_type TEXT NOT NULL,
            memory_type TEXT NOT NULL,
            content TEXT NOT NULL,
            importance_score INTEGER DEFAULT 5,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            last_accessed DATETIME DEFAULT CURRENT_TIMESTAMP,
            access_count INTEGER DEFAULT 1,
            tags TEXT,
            related_memories TEXT,
            FOREIGN KEY (user_id) REFERENCES users(user_id)
        )
    """)
    
    # 6. é¸é¡Œå»ºè­°è¡¨
    cur.execute("""
        CREATE TABLE IF NOT EXISTS topic_suggestions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            suggested_date DATE NOT NULL,
            topics TEXT NOT NULL,
            reasoning TEXT,
            user_feedback TEXT,
            used_count INTEGER DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(user_id),
            UNIQUE(user_id, suggested_date)
        )
    """)
    
    # å»ºç«‹ç´¢å¼•
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sessions_user_agent ON sessions(user_id, agent_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_messages_session ON messages(session_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_memories_user_agent ON agent_memories(user_id, agent_type)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_memories_importance ON agent_memories(importance_score DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_topic_suggestions_user_date ON topic_suggestions(user_id, suggested_date)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_requests_user_time ON requests(user_id, created_at DESC)")

    # ç”¨æˆ¶é»æ•¸èˆ‡è¨‚å–®ï¼ˆç°¡åŒ–ï¼‰
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS user_credits (
            user_id TEXT PRIMARY KEY,
            balance INTEGER DEFAULT 0,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            order_type TEXT NOT NULL,
            amount INTEGER DEFAULT 0,
            plan TEXT,
            status TEXT DEFAULT 'paid',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            paid_at DATETIME
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_orders_user_time ON orders(user_id, created_at DESC)")

    # ç®¡ç†æ“ä½œç¨½æ ¸è¡¨
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS admin_audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            action TEXT NOT NULL,
            admin_token_hash TEXT,
            target_user_id TEXT,
            details TEXT
        )
        """
    )
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_action_time ON admin_audit_logs(action, created_at DESC)")
    
    conn.commit()
    conn.close()

@app.on_event("startup")
def on_startup():
    try:
        init_db()
        global GLOBAL_KB_TEXT
        GLOBAL_KB_TEXT = load_kb_text()
        print(f"[BOOT] KB loaded from {KNOWLEDGE_TXT_PATH} len={len(GLOBAL_KB_TEXT)}")
        print(f"[BOOT] DB ready at {DB_PATH}")
    except Exception as e:
        print("[BOOT] DB init failed:", e)

@app.get("/healthz")
def healthz(): return {"ok": True}

@app.get("/favicon.ico")
def favicon(): return Response(status_code=204)

@app.get("/", response_class=HTMLResponse)
def api_info():
    """API è³‡è¨Šé é¢"""
    return """
    <html><body style="font-family: Arial, sans-serif; max-width: 800px; margin: 50px auto; padding: 20px;">
      <h1>ğŸ¯ ä¸‰æ™ºèƒ½é«”é•·æœŸè¨˜æ†¶ç³»çµ±</h1>
      <p>å¾Œç«¯ API æœå‹™å·²å•Ÿå‹•ï¼å‰ç«¯è«‹è¨ªå•ï¼š<a href="https://jacky6658.github.io/Altest/" target="_blank">https://jacky6658.github.io/Altest/</a></p>
      
      <h2>ğŸ“‹ API ç«¯é»åˆ—è¡¨</h2>
      
      <h3>åŸæœ‰åŠŸèƒ½ï¼š</h3>
      <ul>
        <li><code>POST /chat_generate</code> - è…³æœ¬/æ–‡æ¡ˆäºŒåˆä¸€ç”Ÿæˆ</li>
        <li><code>POST /generate_script</code> - èˆŠæµç¨‹ä¿ç•™</li>
        <li><code>POST /chat_qa</code> - å¼•å°å¼å•ç­”</li>
        <li><code>POST /export/xlsx</code> - Excel åŒ¯å‡º</li>
      </ul>
      
      <h3>æ–°å¢ä¸‰æ™ºèƒ½é«”åŠŸèƒ½ï¼š</h3>
      <ul>
        <li><strong>å®šä½æ™ºèƒ½é«”</strong></li>
        <ul>
          <li><code>POST /agent/positioning/analyze</code> - åˆ†æç”¨æˆ¶å®šä½</li>
          <li><code>PUT /agent/positioning/profile</code> - æ›´æ–°å®šä½æª”æ¡ˆ</li>
        </ul>
        <li><strong>é¸é¡Œæ™ºèƒ½é«”</strong></li>
        <ul>
          <li><code>POST /agent/topics/suggest</code> - ç²å–é¸é¡Œå»ºè­°</li>
          <li><code>GET /agent/topics/history</code> - é¸é¡Œæ­·å²</li>
        </ul>
        <li><strong>è…³æœ¬æ–‡æ¡ˆæ™ºèƒ½é«”</strong></li>
        <ul>
          <li><code>POST /agent/content/generate</code> - ç”Ÿæˆè…³æœ¬/æ–‡æ¡ˆï¼ˆå¢å¼·ç‰ˆï¼‰</li>
        </ul>
        <li><strong>è¨˜æ†¶ç³»çµ±</strong></li>
        <ul>
          <li><code>GET /memory/user/{user_id}</code> - ç²å–ç”¨æˆ¶è¨˜æ†¶</li>
          <li><code>POST /memory/add</code> - æ·»åŠ è¨˜æ†¶</li>
        </ul>
      </ul>
      
      <h2>ğŸ”§ ç³»çµ±ç‹€æ…‹</h2>
      <p>âœ… è³‡æ–™åº«ï¼šå·²åˆå§‹åŒ–</p>
      <p>âœ… çŸ¥è­˜åº«ï¼šå·²è¼‰å…¥</p>
      <p>âœ… ä¸‰æ™ºèƒ½é«”ï¼šå·²å•Ÿå‹•</p>
      <p>âœ… é•·æœŸè¨˜æ†¶ï¼šå·²å•Ÿç”¨</p>
    </body></html>
    """

# ========= Email/å¸³è™Ÿ è¨»å†Š / ç™»å…¥ / æœƒè©± =========
from fastapi import Body

@app.post("/auth/signup")
async def auth_signup(req: Request):
    data = await req.json()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not all([email, username, password, phone]):
        raise HTTPException(status_code=400, detail="missing_fields")

    user_id = f"u_{hashlib.md5((email+username).encode()).hexdigest()[:12]}"

    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        # å»ºç«‹ usersï¼ˆè‹¥ä¸å­˜åœ¨ï¼‰
        u = conn.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,)).fetchone()
        if not u:
            conn.execute(
                "INSERT INTO users (user_id, email, name) VALUES (?, ?, ?)",
                (user_id, email, username)
            )
        # å»ºç«‹ users_auth
        conn.execute(
            "INSERT INTO users_auth (user_id, username, email, phone, password_hash) VALUES (?, ?, ?, ?, ?)",
            (user_id, username, email, phone, hash_password(password))
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=409, detail="user_exists")
    conn.close()
    return {"ok": True, "user_id": user_id}

@app.post("/auth/login")
async def auth_login(req: Request):
    data = await req.json()
    identifier = (data.get("identifier") or "").strip()
    password = (data.get("password") or "").strip()
    if not identifier or not password:
        raise HTTPException(status_code=400, detail="missing_fields")
    conn = get_conn(); conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT user_id, password_hash FROM users_auth WHERE username=? OR email=?",
        (identifier, identifier)
    ).fetchone()
    conn.close()
    if not row or row["password_hash"] != hash_password(password):
        raise HTTPException(status_code=401, detail="invalid_credentials")

    # è¨­ç½® Session Cookie
    token = create_session_cookie(row["user_id"])
    resp = JSONResponse({"ok": True})
    resp.set_cookie(
        "session", token,
        httponly=True, secure=True, samesite="none", max_age=7*24*3600
    )
    return resp

@app.get("/me")
def me(session: str | None = Cookie(default=None)):
    uid = verify_session_cookie(session) if session else None
    if not uid:
        return JSONResponse(status_code=401, content={"error": "unauthorized"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    u = conn.execute("SELECT user_id, email, name FROM users WHERE user_id=?", (uid,)).fetchone()
    conn.close()
    if not u:
        return JSONResponse(status_code=404, content={"error": "user_not_found"})
    return {"id": u["user_id"], "email": u["email"], "name": u["name"]}

@app.post("/logout")
def logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("session")
    return resp

# ========= Auth Helpers =========
def hash_password(password: str) -> str:
    return hashlib.sha256((password or "").encode("utf-8")).hexdigest()

def create_session_cookie(user_id: str) -> str:
    return session_signer.dumps({"uid": user_id, "ts": int(datetime.now().timestamp())})

def verify_session_cookie(cookie_val: str) -> str | None:
    try:
        data = session_signer.loads(cookie_val)
        return data.get("uid")
    except BadSignature:
        return None

def create_admin_session_cookie(username: str) -> str:
    return admin_session_signer.dumps({"adm": username, "ts": int(datetime.now().timestamp())})

def verify_admin_session_cookie(cookie_val: str) -> str | None:
    try:
        data = admin_session_signer.loads(cookie_val)
        return data.get("adm")
    except BadSignature:
        return None

# ========= Google OAuth2 Endpoints =========
@app.get("/auth/google/start")
async def google_start(request: Request, next: str | None = "/"):
    if not _OAUTH_READY:
        return JSONResponse(status_code=501, content={"error": "oauth_not_configured"})
    return await oauth.google.authorize_redirect(request, redirect_uri=OAUTH_REDIRECT_URI, state=next or "/")

@app.get("/auth/google/callback")
async def google_callback(request: Request):
    if not _OAUTH_READY:
        return JSONResponse(status_code=501, content={"error": "oauth_not_configured"})
    try:
        token = await oauth.google.authorize_access_token(request)
        idinfo = await oauth.google.parse_id_token(request, token)
        sub = idinfo.get("sub"); email = idinfo.get("email"); name = idinfo.get("name") or (email.split("@")[0] if email else "user")
        if not sub:
            return JSONResponse(status_code=400, content={"error": "invalid_google_response"})
        user_id = f"g_{sub}"
        create_or_get_user(user_id, email=email, name=name)
        token_val = create_session_cookie(user_id)
        resp = RedirectResponse(url=request.query_params.get("state") or "/")
        resp.set_cookie("session", token_val, httponly=True, secure=True, samesite="lax", max_age=7*24*3600)
        return resp
    except Exception as e:
        print("[OAuth Callback Error]", e)
        return JSONResponse(status_code=500, content={"error": "oauth_failed"})

# ========= å…§å»ºçŸ¥è­˜åº« =========
BUILTIN_KB_SCRIPT = """
ã€çŸ­å½±éŸ³è…³æœ¬åŸå‰‡ï¼ˆæ¿ƒç¸®ï¼‰ã€‘
1) Hook(0-5s) â†’ Value â†’ CTAã€‚60s ç‰ˆå¯æ‹† 5~6 æ®µï¼Œç¯€å¥æ¸…æ¥šã€‚
2) æ¯æ®µè¼¸å‡ºï¼štype/start_sec/end_sec/camera/dialog/visual/ctaã€‚
3) Hook ç”¨ç—›é»/åå·®/æ•¸æ“šé‰¤å­ + å¿«ç¯€å¥ B-rollï¼›Value æ‹†é‡é»ï¼›CTA å‹•è©+åˆ©ç›Š+ä¸‹ä¸€æ­¥ã€‚
4) èªæ°£å£èªã€çŸ­å¥ã€æœ‰ç¯€å¥ï¼Œé¿å…ç©ºè©±ã€‚
"""

BUILTIN_KB_COPY = """
ã€ç¤¾ç¾¤æ–‡æ¡ˆåŸå‰‡ï¼ˆæ¿ƒç¸®ï¼‰ã€‘
1) çµæ§‹ï¼šå¸ç›é–‹é ­ â†’ ä¸»é«”è³£é»/æ•…äº‹ â†’ CTA â†’ Hashtagsã€‚
2) é¢¨æ ¼ï¼šè²¼è¿‘å—çœ¾ã€çŸ­å¥ã€å¯æ­ emojiã€çµå°¾æœ‰å‹•ä½œã€‚
3) Hashtagsï¼šä¸»é—œéµå­— 1-3ã€å»¶ä¼¸ 5-8ã€‚
4) æ¬„ä½ï¼šmain_copy / alternates / hashtags / cta / image_ideasï¼ˆå¹³å°åŒ–åœ–ç‰‡å»ºè­°ï¼‰ã€‚
"""

def load_extra_kb(max_chars=2500) -> str:
    chunks, total = [], 0
    if GLOBAL_KB_TEXT:
        seg = GLOBAL_KB_TEXT[:max_chars]
        chunks.append(f"\n[KB:global]\n{seg}")
        total += len(seg)
    else:
        paths = glob.glob("/data/kb*.txt") + glob.glob("/data/*.kb.txt") + glob.glob("/data/knowledge*.txt")
        for p in paths:
            try:
                t = open(p, "r", encoding="utf-8").read().strip()
                if not t: continue
                take = (max_chars - total)
                seg = t[:take]
                if seg:
                    chunks.append(f"\n[KB:{os.path.basename(p)}]\n{seg}")
                    total += len(seg)
                if total >= max_chars: break
            except Exception:
                pass
    return "\n".join(chunks)

EXTRA_KB = load_extra_kb()

# ========= æç¤ºå­— & å·¥å…· =========
SHORT_HINT_SCRIPT = "å…§å®¹æœ‰é»å¤ªçŸ­äº† ğŸ™ è«‹æä¾›ï¼šè¡Œæ¥­/å¹³å°/æ™‚é•·(ç§’)/ç›®æ¨™/ä¸»é¡Œï¼ˆä¾‹å¦‚ï¼šã€é›»å•†ï½œReelsï½œ60ç§’ï½œè³¼è²·ï½œå¤å­£æ–°å“é–‹ç®±ã€ï¼‰ï¼Œæˆ‘å°±èƒ½ç”Ÿæˆå®Œæ•´è…³æœ¬ã€‚"
SHORT_HINT_COPY   = "å…§å®¹æœ‰é»å¤ªçŸ­äº† ğŸ™ è«‹æä¾›ï¼šå¹³å°/å—çœ¾/èªæ°£/ä¸»é¡Œ/CTAï¼ˆä¾‹å¦‚ï¼šã€IGï½œç”·ç”Ÿè¦–è§’ï½œæ´»åŠ›å›æ­¸ï½œCTAï¼šé»é€£çµã€ï¼‰ï¼Œæˆ‘å°±èƒ½ç”Ÿæˆå®Œæ•´è²¼æ–‡ã€‚"

def _ensure_json_block(text: str) -> str:
    if not text: raise ValueError("empty model output")
    t = text.strip()
    if t.startswith("```"):
        parts = t.split("```")
        if len(parts) >= 3: t = parts[1]
    i = min([x for x in (t.find("{"), t.find("[")) if x >= 0], default=-1)
    if i < 0: return t
    j = max(t.rfind("}"), t.rfind("]"))
    if j > i: return t[i:j+1]
    return t

def detect_mode(messages: List[Dict[str, str]], explicit: Optional[str]) -> str:
    """å„ªå…ˆä½¿ç”¨ explicitï¼›å¦å‰‡ç”¨é—œéµå­—åˆ¤æ–·ã€‚"""
    if explicit in ("script", "copy"):
        return explicit
    last = ""
    for m in reversed(messages or []):
        if m.get("role") == "user":
            last = (m.get("content") or "").lower()
            break
    copy_keys = [
        "æ–‡æ¡ˆ","è²¼æ–‡","copy","hashtag","hashtags",
        "ig","facebook","fb","linkedin","å°ç´…æ›¸","xï¼ˆtwitterï¼‰","x/twitter","æŠ–éŸ³æ–‡æ¡ˆ"
    ]
    if any(k in last for k in copy_keys):
        return "copy"
    return "script"

def parse_segments(json_text: str) -> List[Dict[str, Any]]:
    data = json.loads(json_text)
    if isinstance(data, dict) and "segments" in data: data = data["segments"]
    if not isinstance(data, list): raise ValueError("segments must be a list")
    segs = []
    for it in data:
        segs.append({
            "type": it.get("type") or it.get("label") or "å ´æ™¯",
            "start_sec": it.get("start_sec", None),
            "end_sec": it.get("end_sec", None),
            "camera": it.get("camera", ""),
            "dialog": it.get("dialog", ""),
            "visual": it.get("visual", ""),
            "cta": it.get("cta", "")
        })
    return segs

def parse_copy(json_text: str) -> Dict[str, Any]:
    data = json.loads(json_text)
    if isinstance(data, list): data = data[0] if data else {}
    return {
        "main_copy":   data.get("main_copy", ""),
        "alternates":  data.get("alternates", []) or data.get("openers", []),
        "hashtags":    data.get("hashtags", []),
        "cta":         data.get("cta", ""),
        "image_ideas": data.get("image_ideas", [])
    }

# === NEW: æ¨¡æ¿/æ™‚é•·/æ¨¡å¼èªªæ˜ ===
TEMPLATE_GUIDE = {
    "A": "ä¸‰æ®µå¼ï¼šHook â†’ Value â†’ CTAã€‚é‡é»æ¸…æ¥šã€ç¯€å¥æ˜å¿«ï¼Œé©åˆå»£æ³›æƒ…å¢ƒã€‚",
    "B": "å•é¡Œè§£æ±ºï¼šç—›é» â†’ è§£æ³• â†’ è­‰æ“š/ç¤ºä¾‹ â†’ CTAã€‚é©åˆæ•™è‚²èˆ‡å°è³¼ã€‚",
    "C": "Before-Afterï¼šæ”¹è®Šå‰å¾Œå°æ¯”ï¼Œå¼·èª¿å·®ç•°èˆ‡æ”¶ç›Š â†’ CTAã€‚é©åˆæ¡ˆä¾‹/è¦‹è­‰ã€‚",
    "D": "æ•™å­¸ï¼šæ­¥é©ŸåŒ–æ•™å­¸ï¼ˆ1-2-3ï¼‰+ æ³¨æ„äº‹é … â†’ CTAã€‚é©åˆæŠ€å·§åˆ†äº«ã€‚",
    "E": "æ•˜äº‹ï¼šå°æ•…äº‹é‹ªé™³ â†’ è½‰æŠ˜äº®é» â†’ CTAã€‚é©åˆå“ç‰Œæƒ…ç·’/äººç‰©æ•˜äº‹ã€‚",
    "F": "çˆ†é»é€£ç™¼ï¼šé€£çºŒå¼· Hook/é‡‘å¥/åå·®é»ï¼Œæœ€å¾Œæ”¶æ–‚ â†’ CTAã€‚é©åˆæŠ“æ³¨æ„åŠ›ã€‚"
}

def _duration_plan(duration: Optional[int]) -> Dict[str, Any]:
    """
    å›å‚³åˆ†æ®µå»ºè­°èˆ‡ fewshot JSONã€‚30s èµ° 3 æ®µï¼›60s èµ° 6 æ®µï¼ˆæ¯æ®µ~10sï¼‰ã€‚
    """
    if int(duration or 0) == 60:
        fewshot = """
{"segments":[
  {"type":"hook","start_sec":0,"end_sec":10,"camera":"CU","dialog":"...","visual":"...","cta":""},
  {"type":"value1","start_sec":10,"end_sec":20,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"value2","start_sec":20,"end_sec":30,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"value3","start_sec":30,"end_sec":40,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"value4","start_sec":40,"end_sec":50,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"cta","start_sec":50,"end_sec":60,"camera":"WS","dialog":"...","visual":"...","cta":"..."}
]}
"""
        return {"fewshot": fewshot, "note": "è«‹ä»¥ 60 ç§’ç´„ 6 æ®µè¼¸å‡ºï¼Œæ®µèˆ‡æ®µé–“ç¯€å¥åˆ†æ˜ã€‚"}
    # default 30s
    fewshot = """
{"segments":[
  {"type":"hook","start_sec":0,"end_sec":5,"camera":"CU","dialog":"...","visual":"...","cta":""},
  {"type":"value","start_sec":5,"end_sec":25,"camera":"MS","dialog":"...","visual":"...","cta":""},
  {"type":"cta","start_sec":25,"end_sec":30,"camera":"WS","dialog":"...","visual":"...","cta":"..."}
]}
"""
    return {"fewshot": fewshot, "note": "è«‹ä»¥ 30 ç§’ 3 æ®µè¼¸å‡ºï¼ŒHook è¦å¼·ã€CTA æ˜ç¢ºã€‚"}

def build_script_prompt(
    user_input: str,
    previous_segments: List[Dict[str, Any]],
    template_type: Optional[str] = None,
    duration: Optional[int] = None,
    dialogue_mode: Optional[str] = None,
    knowledge_hint: Optional[str] = None,
) -> str:
    plan = _duration_plan(duration)
    fewshot = plan["fewshot"]
    duration_note = plan["note"]
    tmpl = (template_type or "").strip().upper()
    tmpl_text = TEMPLATE_GUIDE.get(tmpl, "æœªæŒ‡å®šæ¨¡æ¿æ™‚ç”±ä½ åˆ¤æ–·æœ€åˆé©çš„çµæ§‹ã€‚")

    kb = (BUILTIN_KB_SCRIPT + "\n" + (EXTRA_KB or "")).strip()
    # å‹•æ…‹ KB æ“·å–ï¼šåˆä½µä½¿ç”¨è€…è¼¸å…¥èˆ‡å¯é¸æç¤º
    q = user_input
    if knowledge_hint:
        q = f"{knowledge_hint}\n{user_input}"
    try:
        kb_ctx_dynamic = retrieve_context(q)
    except Exception:
        kb_ctx_dynamic = ""

    prev = json.dumps(previous_segments or [], ensure_ascii=False)

    mode_line = ""
    if (dialogue_mode or "").lower() == "free":
        mode_line = "èªæ°£æ›´è‡ªç”±ã€å¯ä¸»å‹•æå‡ºç²¾ç…‰å»ºè­°èˆ‡åå•ä»¥å®Œå–„è…³æœ¬ï¼›"
    elif (dialogue_mode or "").lower() == "guide":
        mode_line = "èªæ°£åå¼•å°ï¼Œé€æ­¥é‡æ¸…è¦ç´ å¾Œç›´æ¥çµ¦å‡ºå®Œæ•´åˆ†æ®µï¼›"

    return f"""
æ ¹æ“šä½¿ç”¨è€…è¼¸å…¥ç”ŸæˆçŸ­å½±éŸ³è…³æœ¬ã€‚{mode_line}

ğŸ¯ è…³æœ¬åƒæ•¸ï¼š
â€¢ æ¨¡æ¿ï¼š{tmpl or "ï¼ˆæœªæŒ‡å®šï¼‰"} - {tmpl_text}
â€¢ æ™‚é•·ï¼š{int(duration) if duration else "ï¼ˆæœªæŒ‡å®šï¼Œé è¨­ 30ï¼‰"} ç§’
â€¢ å¹³å°ï¼šInstagram Reelsã€TikTokã€YouTube Shortsã€Facebook Reels

ğŸ“š çŸ¥è­˜åº«ï¼š
{kb}

ã€KBè¼”åŠ©æ‘˜éŒ„ã€‘ï¼ˆè‹¥ç©ºç™½ä»£è¡¨ç„¡ï¼‰
{kb_ctx_dynamic[:1000]}

ğŸ’¡ å°ç£å¸‚å ´ç‰¹è‰²ï¼š
â€¢ å…§å®¹é¢¨æ ¼ï¼šç”Ÿæ´»åŒ–ã€è¦ªåˆ‡ã€å¯¦ç”¨
â€¢ ç¯€å¥è¦æ±‚ï¼š2-3ç§’æ›ç•«é¢ï¼Œç¯€å¥ç·Šæ¹Š
â€¢ HookåŸå‰‡ï¼š0-5ç§’ç›´çµ¦çµè«–ï¼Œç”¨å¤§å­—å¡èˆ‡å¼·æƒ…ç·’è¡¨æƒ…
â€¢ èªæ°£ï¼šå …å®šã€ç›´çµ¦çµè«–ï¼Œé¿å…å£ç™–è´…å­—

ä½¿ç”¨è€…è¼¸å…¥ï¼š
{user_input}

å·²æ¥å—æ®µè½ï¼š
{prev}

ç›´æ¥è¼¸å‡ºJSONæ ¼å¼ï¼Œä¸è¦ä»»ä½•é–‹å ´ç™½æˆ–èªªæ˜æ–‡å­—ï¼š
{fewshot}
"""

def build_copy_prompt(user_input: str, topic: Optional[str]) -> str:
    topic_line = f"\nã€ä¸»é¡Œã€‘{topic}" if topic else ""
    fewshot = """
{
  "main_copy":"ä¸»è²¼æ–‡ï¼ˆå«æ›è¡Œèˆ‡ emojiï¼‰",
  "alternates":["å‚™é¸é–‹é ­A","å‚™é¸é–‹é ­B","å‚™é¸é–‹é ­C"],
  "hashtags":["#é—œéµå­—1","#é—œéµå­—2","#å»¶ä¼¸3","#å»¶ä¼¸4"],
  "cta":"è¡Œå‹•å‘¼ç±²ä¸€å¥è©±",
  "image_ideas":["é…åœ–/ç…§ç‰‡/ç¤ºæ„åœ–å»ºè­°1","å»ºè­°2","å»ºè­°3"]
}
"""
    kb = (BUILTIN_KB_COPY + "\n" + (EXTRA_KB or "")).strip()
    return f"""
ä½ æ˜¯ç¤¾ç¾¤æ–‡æ¡ˆé¡§å•ã€‚è«‹ä¾ã€Œä½¿ç”¨è€…è¼¸å…¥ã€èˆ‡å¯é¸çš„ä¸»é¡Œè¼¸å‡º**JSON**ï¼ŒåŒ…å«ä¸»è²¼æ–‡ã€å‚™é¸é–‹é ­ã€Hashtagsã€CTAï¼Œä¸¦åŠ å…¥ image_ideasï¼ˆå¹³å°å°å‘çš„åœ–ç‰‡/æ‹æ³•/è¦–è¦ºå»ºè­°ï¼‰ã€‚èªæ°£å¯å£èªä¸¦é©åº¦ä½¿ç”¨ emojiã€‚

{kb}

ä½¿ç”¨è€…è¼¸å…¥ï¼š
{user_input}{topic_line}

åªå›å‚³ JSONï¼ˆå–®ä¸€ç‰©ä»¶ï¼Œä¸è¦ markdown fenceï¼‰ï¼š
{fewshot}
"""

# ========= Gemini =========
def use_gemini() -> bool: return bool(GEMINI_API_KEY)

def gemini_generate_text(prompt: str) -> str:
    import google.generativeai as genai
    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel(GEMINI_MODEL)
    res = model.generate_content(prompt)
    return (res.text or "").strip()

# ========= Fallback =========
def fallback_segments(user_input: str, prev_len: int, duration: Optional[int]=None) -> List[Dict[str, Any]]:
    d = int(duration or 30)
    if d >= 60:
        # ç²—ç•¥ 60s å…­æ®µ
        labels = ["hook","value1","value2","value3","value4","cta"]
        segs=[]
        start=0
        for i,l in enumerate(labels):
            end = 10*(i+1)
            if i==len(labels)-1: end = 60
            cam = "CU" if i==0 else ("WS" if i==len(labels)-1 else "MS")
            segs.append({
                "type": l, "start_sec": start, "end_sec": end, "camera": cam,
                "dialog": f"ï¼ˆæ¨¡æ“¬ï¼‰{user_input[:36]}â€¦",
                "visual": "ï¼ˆæ¨¡æ“¬ï¼‰å¿«åˆ‡ B-roll / å¤§å­—å¡",
                "cta": "é»é€£çµé ˜å– ğŸ”—" if l=="cta" else ""
            })
            start = end
        return segs
    # é è¨­ 30s ä¸‰æ®µ
    step = prev_len
    return [{
        "type": "hook" if step == 0 else ("cta" if step >= 2 else "value"),
        "start_sec": 0 if step == 0 else 5 if step == 1 else 25,
        "end_sec":   5 if step == 0 else 25 if step == 1 else 30,
        "camera": "CU" if step == 0 else "MS" if step == 1 else "WS",
        "dialog": f"ï¼ˆæ¨¡æ“¬ï¼‰{user_input[:36]}â€¦",
        "visual": "ï¼ˆæ¨¡æ“¬ï¼‰å¿«åˆ‡ B-roll / å¤§å­—å¡",
        "cta": "é»é€£çµé ˜å– ğŸ”—" if step >= 2 else ""
    }]

def fallback_copy(user_input: str, topic: Optional[str]) -> Dict[str, Any]:
    t = f"ï¼ˆä¸»é¡Œï¼š{topic}ï¼‰" if topic else ""
    return {
        "main_copy":  f"ï¼ˆæ¨¡æ“¬ï¼‰IG è²¼æ–‡ï¼š{user_input} {t}\nç²¾ç¥å›æ­¸ã€æ•ˆç‡å›å‡ï¼âš¡ï¸\nä»Šå¤©å°±è¡Œå‹•å§ï¼",
        "alternates": ["ğŸ”¥ ä»Šå¤©å°±é–‹å§‹","ğŸ’¡ å…¶å¯¦åªè¦é€™æ¨£åš","ğŸ‘‰ ä½ ä¹Ÿå¯ä»¥"],
        "hashtags":   ["#è¡ŒéŠ·","#AI","#æ–‡æ¡ˆ","#ç¤¾ç¾¤ç¶“ç‡Ÿ"],
        "cta":        "ç«‹å³é»é€£çµ ğŸ”—",
        "image_ideas":["ç”¢å“è¿‘æ‹ + ç”Ÿæ´»æƒ…å¢ƒ","å“ç‰Œè‰²èƒŒæ™¯å¤§å­—å¡","æ­¥é©Ÿæµç¨‹ç¤ºæ„åœ–"]
    }

# ========= ä¸‰æ™ºèƒ½é«”ç³»çµ±æ ¸å¿ƒåŠŸèƒ½ =========

# ç”¨æˆ¶ç®¡ç†
def create_or_get_user(user_id: str, email: str = None, name: str = None) -> Dict:
    """å‰µå»ºæˆ–ç²å–ç”¨æˆ¶"""
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    
    user = conn.execute(
        "SELECT * FROM users WHERE user_id = ?", (user_id,)
    ).fetchone()
    
    if not user:
        conn.execute(
            "INSERT INTO users (user_id, email, name) VALUES (?, ?, ?)",
            (user_id, email, name)
        )
        conn.commit()
        user = conn.execute(
            "SELECT * FROM users WHERE user_id = ?", (user_id,)
        ).fetchone()
    
    conn.close()
    return dict(user) if user else None

def get_user_profile(user_id: str) -> Optional[Dict]:
    """ç²å–ç”¨æˆ¶å®šä½æª”æ¡ˆ"""
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    profile = conn.execute(
        "SELECT * FROM user_profiles WHERE user_id = ?", (user_id,)
    ).fetchone()
    conn.close()
    return dict(profile) if profile else None

def update_user_profile(user_id: str, profile_data: Dict) -> bool:
    """æ›´æ–°ç”¨æˆ¶å®šä½æª”æ¡ˆ"""
    conn = get_conn()
    
    existing = conn.execute(
        "SELECT id FROM user_profiles WHERE user_id = ?", (user_id,)
    ).fetchone()
    
    if existing:
        update_fields = []
        values = []
        for key, value in profile_data.items():
            if key != 'user_id' and value is not None:
                update_fields.append(f"{key} = ?")
                values.append(json.dumps(value) if isinstance(value, (list, dict)) else str(value))
        
        if update_fields:
            values.append(user_id)
            sql = f"UPDATE user_profiles SET {', '.join(update_fields)}, updated_at = CURRENT_TIMESTAMP WHERE user_id = ?"
            conn.execute(sql, values)
    else:
        profile_data['user_id'] = user_id
        fields = list(profile_data.keys())
        placeholders = ['?' for _ in fields]
        values = [json.dumps(v) if isinstance(v, (list, dict)) else str(v) for v in profile_data.values()]
        
        sql = f"INSERT INTO user_profiles ({', '.join(fields)}) VALUES ({', '.join(placeholders)})"
        conn.execute(sql, values)
    
    conn.commit()
    conn.close()
    return True

# æœƒè©±ç®¡ç†
def create_session(user_id: str, agent_type: str) -> str:
    """å‰µå»ºæ–°æœƒè©±"""
    session_id = f"{user_id}_{agent_type}_{int(datetime.now().timestamp())}"
    conn = get_conn()
    
    conn.execute(
        "INSERT INTO sessions (session_id, user_id, agent_type) VALUES (?, ?, ?)",
        (session_id, user_id, agent_type)
    )
    conn.commit()
    conn.close()
    
    return session_id

def add_message(session_id: str, role: str, content: str, metadata: Dict = None):
    """æ·»åŠ å°è©±è¨˜éŒ„"""
    conn = get_conn()
    conn.execute(
        "INSERT INTO messages (session_id, role, content, metadata) VALUES (?, ?, ?, ?)",
        (session_id, role, content, json.dumps(metadata) if metadata else None)
    )
    conn.commit()
    conn.close()

# è¨˜æ†¶ç³»çµ±
def add_memory(user_id: str, agent_type: str, memory_type: str, content: str, 
               importance_score: int = 5, tags: List[str] = None) -> int:
    """æ·»åŠ è¨˜æ†¶"""
    conn = get_conn()
    
    cursor = conn.execute(
        """INSERT INTO agent_memories 
           (user_id, agent_type, memory_type, content, importance_score, tags) 
           VALUES (?, ?, ?, ?, ?, ?)""",
        (user_id, agent_type, memory_type, content, importance_score, 
         json.dumps(tags) if tags else None)
    )
    memory_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return memory_id

def get_user_memories(user_id: str, agent_type: str = None, memory_type: str = None, 
                     limit: int = 20) -> List[Dict]:
    """ç²å–ç”¨æˆ¶è¨˜æ†¶"""
    conn = get_conn()
    conn.row_factory = sqlite3.Row
    
    conditions = ["user_id = ?"]
    params = [user_id]
    
    if agent_type:
        conditions.append("agent_type = ?")
        params.append(agent_type)
    
    if memory_type:
        conditions.append("memory_type = ?")
        params.append(memory_type)
    
    params.append(limit)
    
    sql = f"""SELECT * FROM agent_memories 
              WHERE {' AND '.join(conditions)} 
              ORDER BY importance_score DESC, last_accessed DESC 
              LIMIT ?"""
    
    memories = conn.execute(sql, params).fetchall()
    conn.close()
    
    return [dict(memory) for memory in memories]

# å®šä½æ™ºèƒ½é«”
def positioning_agent_analyze(user_input: str, user_profile: Dict = None, memories: List[Dict] = None) -> str:
    """å®šä½æ™ºèƒ½é«”åˆ†æ - æä¾›çµæ§‹åŒ–å®šä½é¸é …"""
    context = "ä½ æ˜¯å°ˆæ¥­çš„çŸ­å½±éŸ³å®šä½é¡§å•ï¼Œå°ˆé–€æœå‹™å°ç£å¸‚å ´ï¼Œå¹«åŠ©ç”¨æˆ¶å¿«é€Ÿå»ºç«‹æ¸…æ™°çš„å¸³è™Ÿå®šä½ã€‚\n\n"
    
    # åŠ å…¥çŸ¥è­˜åº«å…§å®¹
    kb_context = retrieve_context(user_input) or ""
    if kb_context:
        context += f"ã€çŸ¥è­˜åº«åƒè€ƒã€‘\n{kb_context}\n\n"
    
    if user_profile:
        context += f"ç”¨æˆ¶ç¾æœ‰æª”æ¡ˆï¼š{json.dumps(user_profile, ensure_ascii=False)}\n\n"
    
    if memories:
        context += f"ç›¸é—œè¨˜æ†¶ï¼š\n"
        for memory in memories[:5]:
            context += f"- {memory['content']}\n"
        context += "\n"
    
    context += f"ç”¨æˆ¶è¼¸å…¥ï¼š{user_input}\n\n"
    
    # æª¢æŸ¥å“ªäº›æ¬„ä½é‚„éœ€è¦å¡«å¯«
    missing_fields = []
    if not user_profile or not user_profile.get('business_type'):
        missing_fields.append("æ¥­å‹™é¡å‹")
    if not user_profile or not user_profile.get('target_audience'):
        missing_fields.append("ç›®æ¨™å—çœ¾")
    if not user_profile or not user_profile.get('brand_voice'):
        missing_fields.append("å“ç‰Œèªæ°£")
    if not user_profile or not user_profile.get('primary_platform'):
        missing_fields.append("ä¸»è¦å¹³å°")
    if not user_profile or not user_profile.get('content_goals'):
        missing_fields.append("å…§å®¹ç›®æ¨™")
    if not user_profile or not user_profile.get('posting_frequency'):
        missing_fields.append("ç™¼æ–‡é »ç‡")
    
    context += """ã€é‡è¦ã€‘è«‹åŸºæ–¼çŸ¥è­˜åº«å…§å®¹ï¼Œä»¥çµæ§‹åŒ–æ–¹å¼å›æ‡‰ï¼Œæä¾›å…·é«”çš„å®šä½é¸é …ä¾›ç”¨æˆ¶é¸æ“‡ï¼š

ğŸ“‹ å›æ‡‰æ ¼å¼è¦æ±‚ï¼š
â€¢ ä½¿ç”¨emojiä½œç‚ºåˆ†é»ç¬¦è™Ÿï¼Œè®“å…§å®¹æ›´æ˜“è®€
â€¢ æ®µè½åˆ†æ˜ï¼Œé‡é»çªå‡º
â€¢ æä¾›å…·é«”å¯¦ä½œæ–¹å¼
â€¢ åœ¨å›è¦†ä¸­æ˜ç¢ºæ¨™ç¤ºã€Œæ¥­å‹™é¡å‹ï¼šã€ã€Œç›®æ¨™å—çœ¾ï¼šã€ç­‰æ¬„ä½ï¼Œæ–¹ä¾¿ç³»çµ±è‡ªå‹•æå–
â€¢ åŸºæ–¼çŸ¥è­˜åº«çš„æµé‡/è½‰æ›é‚è¼¯ã€å¹³å°ç­–ç•¥ã€å…§å®¹çµæ§‹ç­‰å°ˆæ¥­å»ºè­°

ğŸ¯ åˆ†ææ­¥é©Ÿï¼š
1ï¸âƒ£ å…ˆåˆ†æç”¨æˆ¶çš„æ¥­å‹™/ç”¢å“/æœå‹™
2ï¸âƒ£ æä¾› 2-3 å€‹å…·é«”çš„å®šä½æ–¹å‘é¸é …
3ï¸âƒ£ æ¯å€‹é¸é …åŒ…å«å®Œæ•´6å€‹æ¬„ä½
4ï¸âƒ£ å¹³å°æ¨è–¦å°ˆæ³¨æ–¼å°ç£ç”¨æˆ¶å¸¸ç”¨å¹³å°ï¼šInstagram Reelsã€TikTokã€YouTube Shortsã€Facebook Reels
5ï¸âƒ£ æä¾›å…·é«”å¯¦ä½œå»ºè­°ï¼ˆåŸºæ–¼çŸ¥è­˜åº«çš„æ‹æ”ã€å‰ªè¼¯ã€å…§å®¹ç­–ç•¥ï¼‰
6ï¸âƒ£ æœ€å¾Œæä¾› 1-2 å€‹å¾ŒçºŒå•é¡Œå¼•å°

ğŸ“ æ ¼å¼ç¯„ä¾‹ï¼š
ã€ğŸ¯ å®šä½é¸é … Aã€‘
ğŸ“Š æ¥­å‹™é¡å‹ï¼šXXX
ğŸ‘¥ ç›®æ¨™å—çœ¾ï¼šXXX  
ğŸ­ å“ç‰Œèªæ°£ï¼šXXX
ğŸ“± ä¸»è¦å¹³å°ï¼šInstagram Reelsï¼ˆå°ç£ç”¨æˆ¶æœ€æ´»èºï¼‰
ğŸ¯ å…§å®¹ç›®æ¨™ï¼šXXX
â° ç™¼æ–‡é »ç‡ï¼šXXX

ğŸ’¡ å¯¦ä½œå»ºè­°ï¼š
â€¢ å…·é«”çš„å…§å®¹ç­–ç•¥ï¼ˆåŸºæ–¼çŸ¥è­˜åº«çš„æµé‡å‹/è½‰æ›å‹é…æ¯”ï¼‰
â€¢ å¹³å°æ“ä½œè¦é»ï¼ˆæ‹æ”æŠ€å·§ã€å‰ªè¼¯ç¯€å¥ã€æ¨™é¡Œé‰¤å­ï¼‰
â€¢ é æœŸæ•ˆæœ

ã€ğŸ¯ å®šä½é¸é … Bã€‘
...

ğŸ¤” æ¥ä¸‹ä¾†ä½ å¯ä»¥ï¼š
1ï¸âƒ£ é¸æ“‡æœ€é©åˆçš„å®šä½æ–¹å‘ï¼ˆA/B/Cï¼‰ï¼Œæˆ‘æœƒå¹«ä½ å®Œå–„ç´°ç¯€
2ï¸âƒ£ å‘Šè¨´æˆ‘ä½ çš„å“ç‰Œæƒ³è¦å‚³é”ä»€éº¼å½¢è±¡å’Œèªæ°£ï¼Ÿ
3ï¸âƒ£ ä½ é‚„æœ‰å…¶ä»–æƒ³äº†è§£çš„å®šä½å•é¡Œå—ï¼Ÿ"""
    
    return context

# é¸é¡Œæ™ºèƒ½é«”
def topic_selection_agent_generate(user_profile: Dict, memories: List[Dict] = None) -> str:
    """é¸é¡Œæ™ºèƒ½é«”ç”Ÿæˆå»ºè­°"""
    context = f"ä½ æ˜¯å°ˆæ¥­çš„å…§å®¹é¸é¡Œé¡§å•ï¼Œç‚ºç”¨æˆ¶æä¾›æ¯æ—¥éˆæ„Ÿå»ºè­°ã€‚\n\n"
    
    if user_profile:
        context += f"ç”¨æˆ¶æª”æ¡ˆï¼š\n"
        context += f"- æ¥­å‹™é¡å‹ï¼š{user_profile.get('business_type', 'æœªè¨­å®š')}\n"
        context += f"- ç›®æ¨™å—çœ¾ï¼š{user_profile.get('target_audience', 'æœªè¨­å®š')}\n"
        context += f"- å“ç‰Œèªæ°£ï¼š{user_profile.get('brand_voice', 'æœªè¨­å®š')}\n"
        context += f"- ä¸»è¦å¹³å°ï¼š{user_profile.get('primary_platform', 'æœªè¨­å®š')}\n\n"
    
    if memories:
        context += f"ç›¸é—œæ´å¯Ÿï¼š\n"
        for memory in memories[:3]:
            context += f"- {memory['content']}\n"
        context += "\n"
    
    context += """æä¾›5å€‹å…·é«”çš„å…§å®¹é¸é¡Œå»ºè­°ï¼Œæ¯å€‹é¸é¡ŒåŒ…å«ï¼š

ğŸ“ é¸é¡Œçµæ§‹ï¼š
1ï¸âƒ£ æ¨™é¡Œ/ä¸»é¡Œ
2ï¸âƒ£ ç‚ºä»€éº¼é©åˆé€™å€‹ç”¨æˆ¶
3ï¸âƒ£ é æœŸæ•ˆæœ
4ï¸âƒ£ å‰µä½œå»ºè­°
5ï¸âƒ£ ç›¸é—œç†±é–€æ¨™ç±¤

ğŸ’¡ å¯¦ä½œè¦é»ï¼š
â€¢ è€ƒæ…®ç•¶å‰ç†±é»ã€å­£ç¯€æ€§ã€ç”¨æˆ¶èˆˆè¶£å’Œå¹³å°ç‰¹æ€§
â€¢ æä¾›å…·é«”çš„æ‹æ”å»ºè­°
â€¢ åŒ…å«Hookã€Valueã€CTAçµæ§‹
â€¢ é©åˆå°ç£ç”¨æˆ¶çš„å…§å®¹é¢¨æ ¼

ç›´æ¥è¼¸å‡ºé¸é¡Œå»ºè­°ï¼Œä¸è¦ä»»ä½•é–‹å ´ç™½æˆ–èªªæ˜æ–‡å­—ã€‚"""
    
    return context

def extract_key_insights(text: str, agent_type: str) -> List[str]:
    """å¾AIå›æ‡‰ä¸­æå–é—œéµæ´å¯Ÿ"""
    insights = []
    lines = text.split('\n')
    
    for line in lines:
        line = line.strip()
        if len(line) > 20 and any(keyword in line for keyword in ['å»ºè­°', 'æ‡‰è©²', 'å¯ä»¥', 'é‡é»', 'é—œéµ']):
            insights.append(line)
    
    return insights[:3]

# === NEW: ç²—ç•¥å¾æ–‡å­—ä¸­æ“·å–å®šä½æ¬„ä½ ===
def extract_profile_fields(text: str) -> Dict[str, Any]:
    """æ™ºèƒ½æ“·å–å®šä½æ¬„ä½ï¼Œå¾ç”¨æˆ¶æ•˜è¿°æˆ– AI å›æ‡‰ä¸­æŠ“å–å®šä½è³‡è¨Šã€‚"""
    if not text:
        return {}
    t = text.strip()
    import re
    fields: Dict[str, Any] = {}

    # æ¥­å‹™é¡å‹ - æ›´å»£æ³›çš„åŒ¹é…
    business_patterns = [
        r"(?:æ¥­å‹™é¡å‹|è¡Œæ¥­|ç”¢æ¥­|åš|ç¶“ç‡Ÿ|å¾äº‹)[:ï¼š]\s*([^\nï¼Œã€‚,ï¼›;]{2,50})",
        r"(?:æˆ‘æ˜¯|æˆ‘å€‘æ˜¯|å…¬å¸æ˜¯|å¸³è™Ÿæ˜¯|å°ˆæ³¨æ–¼|ä¸»è¦åš)\s*([^\nï¼Œã€‚,ï¼›;]{2,50})",
        r"(?:AIæ™ºèƒ½é«”|AIè‡ªå‹•åŒ–|çŸ­å½±éŸ³|é›»å•†|æ•™è‚²|ç§‘æŠ€|è¡ŒéŠ·|å…§å®¹å‰µä½œ|çŸ¥è­˜åˆ†äº«)",
    ]
    for pattern in business_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["business_type"] = m.group(1).strip()
            break

    # ç›®æ¨™å—çœ¾ - æ›´æ™ºèƒ½çš„åŒ¹é…
    audience_patterns = [
        r"(?:ç›®æ¨™å—çœ¾|å—çœ¾|TA|è§€çœ¾|ç²‰çµ²)[:ï¼š]\s*([^\n]{2,100})",
        r"(?:æ•ˆç‡æ§|è·å ´æ‰“å·¥äºº|ç§‘æŠ€å¥½å¥‡å¯¶å¯¶|æœªä¾†ç”Ÿæ´»åš®å¾€è€…|å¹´è¼•äºº|å­¸ç”Ÿ|ä¸Šç­æ—|æ–°æ‰‹çˆ¸åª½)",
        r"(?:å¹´é½¡|æ€§åˆ¥|è·æ¥­|èˆˆè¶£|ç—›é»)[:ï¼š]\s*([^\n]{2,80})",
    ]
    for pattern in audience_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["target_audience"] = m.group(1).strip()
            break

    # å“ç‰Œèªæ°£ - æ›´å»£æ³›çš„åŒ¹é…
    voice_patterns = [
        r"(?:å“ç‰Œèªæ°£|èªæ°£|å£å»|é¢¨æ ¼)[:ï¼š]\s*([^\nï¼Œã€‚,ï¼›;]{2,50})",
        r"(?:å¹½é»˜|ä¿çš®|å°ˆæ¥­|è¦ªåˆ‡|æ´»æ½‘|åš´è‚…|è¼•é¬†|æ­£å¼|å£èª|ç™½è©±)",
        r"(?:åƒ.*æœ‹å‹|é…·æœ‹å‹|è‡ªç„¶|æœ‰è¨˜æ†¶é»|æœ‰å…±é³´)",
    ]
    for pattern in voice_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["brand_voice"] = m.group(1).strip()
            break

    # ä¸»è¦å¹³å° - æ›´æ™ºèƒ½çš„åŒ¹é…
    platform_patterns = [
        r"(?:ä¸»è¦å¹³å°|æ ¸å¿ƒå¹³å°|å¹³å°|åœ¨å“ªè£¡ç¶“ç‡Ÿ)[:ï¼š]\s*([^\nï¼Œã€‚,ï¼›;]{2,50})",
        r"(?:æŠ–éŸ³|å°ç´…æ›¸|IG|Instagram|YouTube|Facebook|TikTok|å¾®åš|Bç«™)",
    ]
    for pattern in platform_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["primary_platform"] = m.group(1).strip()
            break

    # å…§å®¹ç›®æ¨™ - æ›´å»£æ³›çš„åŒ¹é…
    goals_patterns = [
        r"(?:å…§å®¹ç›®æ¨™|ç›®æ¨™|ç›®çš„|æƒ³è¦)[:ï¼š]\s*([^\n]{2,100})",
        r"(?:è½‰å–®|æ›å…‰|åå–®|æ•™è‚²|å“ç‰Œ|æµé‡|ç²‰çµ²|äº’å‹•|éŠ·å”®|æ¨å»£)",
    ]
    for pattern in goals_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["content_goals"] = m.group(1).strip()
            break

    # ç™¼æ–‡é »ç‡ - æ›´æ™ºèƒ½çš„åŒ¹é…
    frequency_patterns = [
        r"(?:ç™¼æ–‡é »ç‡|é »ç‡|å¤šä¹…ç™¼|æ›´æ–°)[:ï¼š]\s*([^\nï¼Œã€‚,ï¼›;]{2,30})",
        r"(?:æ¯å¤©|æ¯é€±|æ¯æœˆ|ä¸å®šæœŸ|å›ºå®š|ç¶“å¸¸|å¶çˆ¾)",
    ]
    for pattern in frequency_patterns:
        m = re.search(pattern, t, re.IGNORECASE)
        if m: 
            fields["posting_frequency"] = m.group(1).strip()
            break

    return fields

# === NEW: ç„¡æ¨¡å‹æ™‚çš„è‡ªç„¶å›è¦†ï¼ˆåƒè€ƒè³‡æ–™åº«ï¼‰ ===
def natural_fallback_positioning(user_input: str, user_profile: Optional[Dict], memories: List[Dict]) -> str:
    """åœ¨æ²’æœ‰å¤–éƒ¨æ¨¡å‹æ™‚ï¼Œæ ¹æ“šç”¨æˆ¶æª”æ¡ˆèˆ‡è¨˜æ†¶ï¼Œç”Ÿæˆæ¯”è¼ƒè‡ªç„¶çš„å»ºè­°æ–‡æœ¬ã€‚"""
    bp = user_profile or {}
    biz = bp.get("business_type") or "ï¼ˆæœªè¨­å®šï¼‰"
    aud = bp.get("target_audience") or "ï¼ˆæœªè¨­å®šï¼‰"
    voice = bp.get("brand_voice") or "ï¼ˆæœªè¨­å®šï¼‰"
    platform = bp.get("primary_platform") or "ï¼ˆæœªè¨­å®šï¼‰"

    insights_lines = []
    for m in (memories or [])[:3]:
        insights_lines.append(f"- {m.get('content','').strip()}")
    insights_block = "\n".join(insights_lines) if insights_lines else "ï¼ˆæš«ç„¡ï¼‰"

    return (
        "ğŸ” åˆæ­¥åˆ†æï¼ˆæ ¹æ“šå·²çŸ¥æª”æ¡ˆèˆ‡ä½ çš„æè¿°ï¼‰\n\n"
        f"1) æ¥­å‹™é¡å‹ï¼š{biz}\n"
        f"2) ç›®æ¨™å—çœ¾ï¼š{aud}\n"
        f"3) å“ç‰Œèªæ°£å»ºè­°ï¼š{voice if voice!='ï¼ˆæœªè¨­å®šï¼‰' else 'å…ˆä»¥æ¸…æ™°ã€å¯ä¿¡ã€å£èªç‚ºä¸»ï¼Œå¾ŒçºŒå†å¾®èª¿'}\n"
        f"4) å¹³å°ç­–ç•¥ï¼šå„ªå…ˆè€•è€˜ {platform if platform!='ï¼ˆæœªè¨­å®šï¼‰' else 'ä½ æœ€ç†Ÿæ‚‰ä¸”å—çœ¾é›†ä¸­çš„å¹³å°'}ï¼Œå†è¼”ä»¥æ¬¡è¦å¹³å°åšå°æµã€‚\n"
        "5) å…§å®¹æ–¹å‘ï¼šä»¥ç—›é»åˆ‡å…¥ + æ¡ˆä¾‹/ç¤ºç¯„ + æ˜ç¢º CTAã€‚æ¯é€±å›ºå®šæ¬„ç›®ï¼ˆä¾‹å¦‚ï¼šæ•™å­¸/é–‹ç®±/QA/æ¡ˆä¾‹ï¼‰ã€‚\n\n"
        "ğŸ§  è¿‘æœŸæ´å¯Ÿï¼š\n"
        f"{insights_block}\n\n"
        "âœ… ä¸‹ä¸€æ­¥ï¼š\n"
        "- å‘Šè¨´æˆ‘ä½ çš„ç”¢å“/æœå‹™ä¸€å¥è©±ï¼‹ä¸»è¦å—çœ¾ï¼‹å¸Œæœ›é”æˆçš„ç›®æ¨™ï¼ˆä¾‹å¦‚ï¼šè½‰å–®/æ›å…‰/åå–®ï¼‰\n"
        "- æˆ‘æœƒæ“šæ­¤è£œé½Šå®šä½æª”æ¡ˆä¸¦çµ¦ä½  2 ç‰ˆå…§å®¹ç­–ç•¥è‰æ¡ˆ"
    )

# ========= å¼•å°å¼å•ç­” API =========
@app.post("/chat_qa")
async def chat_qa(req: Request):
    data = await req.json()
    session_id = (data.get("session_id") or "qa").strip() or "qa"
    user_msg = (data.get("message") or "").strip()

    # åˆæ¬¡é€²å…¥ï¼šå»ºç«‹ session ä¸¦é€æ­¡è¿ + Q1
    if session_id not in QA_SESSIONS:
        qa_reset(session_id)
        q = qa_next_question(session_id)
        return {
            "session_id": session_id,
            "assistant_message": "å—¨ğŸ‘‹ è®“æˆ‘å€‘ä¸€æ­¥æ­¥ç”Ÿæˆä½ çš„çŸ­å½±éŸ³è…³æœ¬ï¼\n" + (q or ""),
            "segments": [],
            "done": False,
            "error": None
        }

    # æ­£å¸¸æµç¨‹ï¼šè¨˜éŒ„ä¸Šä¸€é¡Œçš„å›ç­”
    qa_record_answer(session_id, user_msg)
    next_q = qa_next_question(session_id)
    if next_q:
        return {
            "session_id": session_id,
            "assistant_message": next_q,
            "segments": [],
            "done": False,
            "error": None
        }

    # å•ç­”å®Œæˆ â†’ çµ„åˆæè¿° + å– KB context â†’ èµ°åŸæœ‰ build_script_prompt
    ans = QA_SESSIONS.get(session_id, {}).get("answers", {})
    brief = compose_brief_from_answers(ans)
    kb_ctx = retrieve_context(brief) or ""
    # å°‡ QA é¸åˆ°çš„ structure/duration å¸¶å…¥
    template_type = (ans.get("structure") or "").strip()[:1].upper() or None
    try:
        duration = int((ans.get("duration") or "").strip())
    except Exception:
        duration = 30

    user_input = f"{brief}\n\nã€KBè¼”åŠ©æ‘˜éŒ„ã€‘\n{kb_ctx}"

    previous_segments = []
    prompt = build_script_prompt(
        user_input,
        previous_segments,
        template_type=template_type,
        duration=duration,
        dialogue_mode="guide",
    )
    try:
        if use_gemini():
            out = gemini_generate_text(prompt)
            j = _ensure_json_block(out)
            segments = parse_segments(j)
        else:
            segments = fallback_segments(user_input, 0, duration=duration)
    except Exception as e:
        print("[chat_qa] error:", e)
        segments = []

    # æ¸…é™¤ session
    QA_SESSIONS.pop(session_id, None)

    return {
        "session_id": session_id,
        "assistant_message": "æˆ‘å·²æ ¹æ“šä½ çš„å›ç­”ç”Ÿæˆç¬¬ä¸€ç‰ˆè…³æœ¬ï¼ˆå¯å†èª¿æ•´ï¼‰ã€‚",
        "segments": segments,
        "done": True,
        "error": None
    }

# ========= /chat_generate =========
@app.post("/chat_generate")
async def chat_generate(req: Request):
    """
    body: {
      user_id?: str,
      session_id?: str,
      messages: [{role, content}],
      previous_segments?: [segment...],
      remember?: bool,
      mode?: "script" | "copy",          # â† ä¿ç•™æ—¢æœ‰ï¼šè…³æœ¬/æ–‡æ¡ˆ
      topic?: str,                        # â† æ–‡æ¡ˆä¸»é¡Œï¼ˆå¯é¸ï¼‰
      dialogue_mode?: "guide" | "free",   # â† æ–°å¢ï¼šå¼•å°/è‡ªç”± å°è©±é¢¨æ ¼ï¼ˆå¯é¸ï¼‰
      template_type?: "A"|"B"|"C"|"D"|"E"|"F",  # â† æ–°å¢
      duration?: 30|60,                   # â† æ–°å¢ï¼šè…³æœ¬æ™‚é•·
      knowledge_hint?: str                # â† æ–°å¢ï¼šæª¢ç´¢æç¤ºè©ï¼ˆå¯é¸ï¼‰
    }
    """
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_id = (data.get("user_id") or "").strip() or "web"
    messages = data.get("messages") or []
    previous_segments = data.get("previous_segments") or []
    topic = (data.get("topic") or "").strip() or None

    explicit_mode = (data.get("mode") or "").strip().lower() or None
    mode = detect_mode(messages, explicit=explicit_mode)

    # NEW: è®€å–æ–°åƒæ•¸ï¼ˆå¾Œç«¯è‹¥æ²’æ”¶åˆ°ä¹Ÿä¸å½±éŸ¿èˆŠè¡Œç‚ºï¼‰
    dialogue_mode = (data.get("dialogue_mode") or "").strip().lower() or None
    template_type = (data.get("template_type") or "").strip().upper() or None
    try:
        duration = int(data.get("duration")) if data.get("duration") is not None else None
    except Exception:
        duration = None
    knowledge_hint = (data.get("knowledge_hint") or "").strip() or None

    user_input = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            user_input = (m.get("content") or "").strip()
            break

    # è‹¥è¼¸å…¥éçŸ­ï¼Œä¹Ÿç›´æ¥å˜—è©¦ç”Ÿæˆï¼ˆé¿å…åˆ¶å¼æç¤ºæ‰“æ–·å°è©±ï¼‰
    hint = SHORT_HINT_COPY if mode == "copy" else SHORT_HINT_SCRIPT
    if len(user_input) < 6:
        user_input = f"ï¼ˆä½¿ç”¨è€…æç¤ºè¼ƒçŸ­ï¼‰è«‹ä¸»å‹•è¿½å•å¿…è¦è³‡è¨Šä¸¦å…ˆçµ¦å‡ºåˆæ­¥å»ºè­°ã€‚\næç¤ºï¼š{user_input or 'è«‹å…ˆå¹«æˆ‘é–‹å§‹'}"

    try:
        if mode == "copy":
            prompt = build_copy_prompt(user_input, topic)
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                copy = parse_copy(j)
            else:
                copy = fallback_copy(user_input, topic)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "æˆ‘å…ˆçµ¦ä½ ç¬¬ä¸€ç‰ˆå®Œæ•´è²¼æ–‡ï¼ˆå¯å†åŠ è¦æ±‚ï¼Œæˆ‘æœƒå¹«ä½ æ”¹å¾—æ›´è²¼è¿‘é¢¨æ ¼ï¼‰ã€‚",
                "segments": [],
                "copy": copy,
                "error": None
            }

        else:  # script
            prompt = build_script_prompt(
                user_input,
                previous_segments,
                template_type=template_type,
                duration=duration,
                dialogue_mode=dialogue_mode,
                knowledge_hint=knowledge_hint,
            )
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                segments = parse_segments(j)
            else:
                segments = fallback_segments(user_input, len(previous_segments or []), duration=duration)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "æˆ‘å…ˆçµ¦ä½ ç¬¬ä¸€ç‰ˆå®Œæ•´è…³æœ¬ï¼ˆå¯å†åŠ è¦æ±‚ï¼Œæˆ‘æœƒå¹«ä½ æ”¹å¾—æ›´è²¼è¿‘é¢¨æ ¼ï¼‰ã€‚",
                "segments": segments,
                "copy": None,
                "error": None
            }

        # DB ç´€éŒ„ï¼ˆä¿ç•™åŸè¡Œç‚ºï¼‰
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "INSERT INTO requests (user_id, user_input, mode, messages_json, previous_segments_json, response_json) VALUES (?, ?, ?, ?, ?, ?)",
                (
                    user_id,
                    user_input, mode,
                    json.dumps(messages, ensure_ascii=False),
                    json.dumps(previous_segments, ensure_ascii=False),
                    json.dumps(resp, ensure_ascii=False),
                ),
            )
            conn.commit()
            conn.close()
        except Exception as e:
            print("[DB] insert failed:", e)

        return resp

    except Exception as e:
        print("[chat_generate] error:", e)
        return JSONResponse(status_code=500, content={
            "session_id": data.get("session_id") or "s",
            "assistant_message": "ä¼ºæœå™¨å¿™ç¢Œï¼Œç¨å¾Œå†è©¦",
            "segments": [],
            "copy": None,
            "error": "internal_server_error"
        })

# ========= èˆŠæµç¨‹ï¼š/generate_script =========
@app.post("/generate_script")
async def generate_script(req: Request):
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_input = (data.get("user_input") or "").strip()
    previous_segments = data.get("previous_segments") or []

    # å‘ä¸‹ç›¸å®¹ï¼šèˆŠç«¯é»è‹¥æƒ³æ”¯æ´ 60s/æ¨¡æ¿ï¼Œä¹Ÿå¯å¸¶å…¥é€™å…©å€‹æ¬„ä½ï¼ˆå¯é¸ï¼‰
    template_type = (data.get("template_type") or "").strip().upper() or None
    try:
        duration = int(data.get("duration")) if data.get("duration") is not None else None
    except Exception:
        duration = None

    if len(user_input) < 6:
        return {"segments": [], "error": SHORT_HINT_SCRIPT}

    try:
        prompt = build_script_prompt(
            user_input,
            previous_segments,
            template_type=template_type,
            duration=duration
        )
        if use_gemini():
            out = gemini_generate_text(prompt)
            j = _ensure_json_block(out)
            segments = parse_segments(j)
        else:
            segments = fallback_segments(user_input, len(previous_segments or []), duration=duration)
        return {"segments": segments, "error": None}
    except Exception as e:
        print("[generate_script] error:", e)
        return JSONResponse(status_code=500, content={"segments": [], "error": "internal_server_error"})

# ========= åŒ¯å‡ºï¼šWord æš«åœ / Excel ä¿ç•™ =========
@app.post("/export/docx")
async def export_docx_disabled():
    return JSONResponse(status_code=501, content={"error": "docx_export_disabled"})

def _ensure_xlsx():
    try:
        import openpyxl  # noqa
        return True
    except Exception:
        return False

@app.post("/export/xlsx")
async def export_xlsx(req: Request):
    if not _ensure_xlsx():
        return JSONResponse(status_code=501, content={"error": "xlsx_not_available"})
    import openpyxl
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = await req.json()
    segments = data.get("segments") or []
    copy = data.get("copy") or None

    wb = openpyxl.Workbook()
    ws1 = wb.active; ws1.title = "è…³æœ¬åˆ†æ®µ"
    ws1.append(["#","type","start_sec","end_sec","camera","dialog","visual","cta"])
    for i, s in enumerate(segments, 1):
        ws1.append([i, s.get("type"), s.get("start_sec"), s.get("end_sec"),
                    s.get("camera"), s.get("dialog"), s.get("visual"), s.get("cta")])

    ws2 = wb.create_sheet("æ–‡æ¡ˆ")
    ws2.append(["ä¸»è²¼æ–‡"]); ws2.append([copy.get("main_copy") if copy else ""])
    ws2.append([]); ws2.append(["å‚™é¸é–‹é ­"])
    for a in (copy.get("alternates") if copy else []) or []: ws2.append([a])
    ws2.append([]); ws2.append(["Hashtags"])
    ws2.append([" ".join(copy.get("hashtags") if copy else [])])
    ws2.append([]); ws2.append(["CTA"])
    ws2.append([copy.get("cta") if copy else ""])
    ws2.append([]); ws2.append(["åœ–ç‰‡å»ºè­°"])
    for idea in (copy.get("image_ideas") if copy else []) or []: ws2.append([idea])

    for ws in (ws1, ws2):
        for col in ws.columns:
            width = max(len(str(c.value)) if c.value else 0 for c in col) + 2
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width, 80)

    bio = BytesIO(); wb.save(bio); bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="export.xlsx"'}
    )

# ========= CSV ä¸‹è¼‰ & Google Sheet é€£å‹• =========
import csv
import json
from fastapi.responses import FileResponse, Response
from io import StringIO

@app.get("/download/requests_export.csv")
def download_requests_csv():
    export_path = "/data/requests_export.csv"
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM requests ORDER BY id DESC")
    rows = cur.fetchall()
    headers = [desc[0] for desc in cur.description]
    conn.close()

    with open(export_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    return FileResponse(
        export_path,
        media_type="text/csv",
        filename="requests_export.csv",
    )


@app.get("/export/google-sheet")
def export_for_google_sheet(limit: int = 100):
    try:
        limit = int(limit)
    except Exception:
        limit = 100
    limit = max(1, min(limit, 2000))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"SELECT id, created_at, user_input, mode FROM requests ORDER BY id DESC LIMIT {limit}"
    )
    rows = cur.fetchall()
    conn.close()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "created_at", "user_input", "mode"])
    for row in rows:
        writer.writerow(row)

    return Response(content=output.getvalue(), media_type="text/csv")


@app.get("/export/google-sheet-flat")
def export_google_sheet_flat(limit: int = 200):
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT id, created_at, user_input, mode, response_json
        FROM requests
        ORDER BY id DESC
        LIMIT {limit}
        """
    )
    rows = cur.fetchall()
    conn.close()

    out = StringIO()
    writer = csv.writer(out)

    headers = [
        "id", "created_at", "mode", "user_input",
        "assistant_message",
        "copy_main_copy",
        "copy_cta",
        "copy_hashtags",
        "copy_alternates_joined",
        "segments_count",
        "seg1_type", "seg1_start_sec", "seg1_end_sec", "seg1_dialog", "seg1_visual", "seg1_cta",
        "seg2_type", "seg2_start_sec", "seg2_end_sec", "seg2_dialog", "seg2_visual", "seg2_cta",
        "seg3_type", "seg3_start_sec", "seg3_end_sec", "seg3_dialog", "seg3_visual", "seg3_cta",
    ]
    writer.writerow(headers)

    for _id, created_at, user_input, mode, resp_json in rows:
        assistant_message = ""
        copy_main = ""
        copy_cta = ""
        copy_hashtags = ""
        copy_alternates_joined = ""
        segments_count = 0

        def empty_seg():
            return ["", "", "", "", "", ""]
        seg1 = empty_seg()
        seg2 = empty_seg()
        seg3 = empty_seg()

        try:
            data = json.loads(resp_json or "{}")
            assistant_message = (data.get("assistant_message") or "")[:500]

            c = data.get("copy") or {}
            if isinstance(c, dict):
                copy_main = c.get("main_copy") or ""
                copy_cta = c.get("cta") or ""
                tags = c.get("hashtags") or []
                if isinstance(tags, list):
                    copy_hashtags = " ".join(map(str, tags))
                alts = c.get("alternates") or c.get("openers") or []
                if isinstance(alts, list):
                    copy_alternates_joined = " | ".join(map(str, alts))

            segs = data.get("segments") or []
            if isinstance(segs, list):
                segments_count = len(segs)

                def to_seg(s):
                    return [
                        s.get("type", ""),
                        s.get("start_sec", ""),
                        s.get("end_sec", ""),
                        s.get("dialog", ""),
                        s.get("visual", ""),
                        s.get("cta", ""),
                    ]

                if len(segs) >= 1: seg1 = to_seg(segs[0])
                if len(segs) >= 2: seg2 = to_seg(segs[1])
                if len(segs) >= 3: seg3 = to_seg(segs[2])

        except Exception as e:
            assistant_message = f"[JSON parse error] {str(e)}"

        writer.writerow([
            _id, created_at, mode, user_input,
            assistant_message,
            copy_main,
            copy_cta,
            copy_hashtags,
            copy_alternates_joined,
            segments_count,
            *seg1, *seg2, *seg3,
        ])

    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "inline; filename=export_flat.csv"},
    )

# ========= Google Sheet æ‰å¹³åŒ–ï¼ˆv2ï¼‰ =========
import csv
import json
from io import StringIO
from fastapi.responses import Response

@app.get("/export/google-sheet-flat-v2")
def export_google_sheet_flat_v2(limit: int = 200):
    """
    æ‰å¹³åŒ– CSVï¼ˆå« copy èˆ‡å‰ 3 å€‹ segmentsï¼‰ï¼Œç¦ç”¨å¿«å–ã€‚
    åœ¨ Google Sheets ä½¿ç”¨ï¼š
      =IMPORTDATA("https://aijobvideobackend.zeabur.app/export/google-sheet-flat-v2?limit=500")
    """
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT id, created_at, user_input, mode, response_json
        FROM requests
        ORDER BY id DESC
        LIMIT {limit}
        """
    )
    rows = cur.fetchall()
    conn.close()

    out = StringIO()
    writer = csv.writer(out)

    headers = [
        "id", "created_at", "mode", "user_input",
        "assistant_message",
        "copy_main_copy", "copy_cta", "copy_hashtags", "copy_alternates_joined",
        "segments_count",
        "seg1_type", "seg1_start_sec", "seg1_end_sec", "seg1_dialog", "seg1_visual", "seg1_cta",
        "seg2_type", "seg2_start_sec", "seg2_end_sec", "seg2_dialog", "seg2_visual", "seg2_cta",
        "seg3_type", "seg3_start_sec", "seg3_end_sec", "seg3_dialog", "seg3_visual", "seg3_cta",
    ]
    writer.writerow(headers)

    def empty_seg():
        return ["", "", "", "", "", ""]

    for _id, created_at, user_input, mode, resp_json in rows:
        assistant_message = ""
        copy_main = ""
        copy_cta = ""
        copy_hashtags = ""
        copy_alternates = ""
        segments_count = 0
        seg1 = empty_seg()
        seg2 = empty_seg()
        seg3 = empty_seg()

        try:
            data = json.loads(resp_json or "{}")
            assistant_message = (data.get("assistant_message") or "")[:500]

            c = data.get("copy") or {}
            if isinstance(c, dict):
                copy_main = c.get("main_copy") or ""
                copy_cta = c.get("cta") or ""
                tags = c.get("hashtags") or []
                if isinstance(tags, list):
                    copy_hashtags = " ".join(map(str, tags))
                alts = c.get("alternates") or c.get("openers") or []
                if isinstance(alts, list):
                    copy_alternates = " | ".join(map(str, alts))

            segs = data.get("segments") or []
            if isinstance(segs, list):
                segments_count = len(segs)

                def to_seg(s):
                    return [
                        s.get("type", ""),
                        s.get("start_sec", ""),
                        s.get("end_sec", ""),
                        s.get("dialog", ""),
                        s.get("visual", ""),
                        s.get("cta", ""),
                    ]

                if len(segs) >= 1: seg1 = to_seg(segs[0])
                if len(segs) >= 2: seg2 = to_seg(segs[1])
                if len(segs) >= 3: seg3 = to_seg(segs[2])

        except Exception as e:
            assistant_message = f"[JSON parse error] {str(e)}"

        writer.writerow([
            _id, created_at, mode, user_input,
            assistant_message,
            copy_main, copy_cta, copy_hashtags, copy_alternates,
            segments_count,
            *seg1, *seg2, *seg3,
        ])

    return Response(
        content=out.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": "inline; filename=export_flat_v2.csv",
            "Cache-Control": "no-store, max-age=0, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )

# ========= Admin APIsï¼ˆç°¡æ˜“ç‹€æ…‹èˆ‡ç”¨æˆ¶åˆ—è¡¨ï¼‰ =========
ADMIN_TOKEN = os.getenv("ADMIN_TOKEN")  # å¯é¸ï¼Œè‹¥æœªè¨­ç½®å‰‡ä¸é©—è­‰

def _check_admin(req: Request):
    # å…ˆçœ‹æ˜¯å¦æœ‰æœ‰æ•ˆ admin session cookie
    adm_cookie = req.cookies.get("admin_session")
    if adm_cookie and verify_admin_session_cookie(adm_cookie):
        return True
    # å…¶æ¬¡å…è¨± tokenï¼ˆè‡ªå‹•åŒ–å·¥å…·/å‚™æ´ï¼‰
    tok = req.headers.get("x-admin-token") or req.query_params.get("token")
    if ADMIN_TOKEN and tok == ADMIN_TOKEN:
        return True
    return False

@app.get("/admin/users")
async def admin_users(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    users = conn.execute("SELECT user_id, email, name, created_at, updated_at, status FROM users ORDER BY created_at DESC LIMIT 500").fetchall()
    auths = conn.execute(
        """
        SELECT user_id, username, email, phone, created_at,
               CASE WHEN password_hash IS NOT NULL AND length(password_hash)>0 THEN 1 ELSE 0 END AS has_password
        FROM users_auth ORDER BY created_at DESC LIMIT 500
        """
    ).fetchall()
    conn.close()
    return {
        "users": [dict(u) for u in users],
        "users_auth": [dict(a) for a in auths],
    }

@app.get("/admin/users_full")
async def admin_users_full(req: Request, limit: int = 500):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 500
    limit = max(1, min(limit, 2000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    users = conn.execute("SELECT * FROM users ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
    auths = conn.execute(
        """
        SELECT user_id, username, email, phone, created_at,
               CASE WHEN password_hash IS NOT NULL AND length(password_hash)>0 THEN 1 ELSE 0 END AS has_password
        FROM users_auth ORDER BY created_at DESC LIMIT ?
        """,
        (limit,)
    ).fetchall()
    credits = conn.execute("SELECT * FROM user_credits").fetchall()
    orders = conn.execute("SELECT * FROM orders ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()
    conn.close()
    return {
        "users": [dict(u) for u in users],
        "users_auth": [dict(a) for a in auths],
        "credits": [dict(c) for c in credits],
        "orders": [dict(o) for o in orders],
    }

@app.get("/admin/usage")
async def admin_usage(req: Request, limit: int = 30):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 30
    limit = max(1, min(limit, 500))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    total_requests = conn.execute("SELECT COUNT(1) AS c FROM requests").fetchone()["c"]
    latest = conn.execute(
        f"SELECT id, created_at, user_input, mode FROM requests ORDER BY id DESC LIMIT {limit}"
    ).fetchall()
    conn.close()
    return {
        "total_requests": total_requests,
        "latest": [dict(r) for r in latest],
    }

@app.get("/admin/users.csv")
async def admin_users_csv(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    users = conn.execute("SELECT user_id, email, name, created_at, updated_at, status FROM users ORDER BY created_at DESC").fetchall()
    auths = conn.execute("SELECT user_id, username, email AS auth_email, phone, created_at AS auth_created FROM users_auth ORDER BY created_at DESC").fetchall()
    conn.close()

    from io import StringIO
    s = StringIO()
    import csv
    w = csv.writer(s)
    w.writerow(["user_id","email","name","created_at","updated_at","status","username","auth_email","phone","auth_created"])
    # ä»¥ user_id é—œè¯ï¼ˆæ­¤è™•ç°¡åŒ–ï¼šé€ç­†åˆä½µï¼Œè‹¥ç„¡å°æ‡‰å‰‡ç•™ç©ºï¼‰
    auth_map = {a["user_id"]: a for a in auths}
    for u in users:
        a = auth_map.get(u["user_id"]) or {}
        w.writerow([
            u["user_id"], u["email"], u["name"], u["created_at"], u["updated_at"], u["status"],
            a.get("username",""), a.get("auth_email",""), a.get("phone",""), a.get("auth_created",""),
        ])
    return Response(content=s.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=users.csv"})

@app.get("/admin/usage.csv")
async def admin_usage_csv(req: Request, limit: int = 1000):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 1000
    limit = max(1, min(limit, 5000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"SELECT id, created_at, user_input, mode FROM requests ORDER BY id DESC LIMIT {limit}"
    ).fetchall()
    conn.close()

    from io import StringIO
    s = StringIO()
    import csv
    w = csv.writer(s)
    w.writerow(["id","created_at","mode","user_input"])
    for r in rows:
        w.writerow([r["id"], r["created_at"], r["mode"], r["user_input"]])
    return Response(content=s.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=usage.csv"})

@app.get("/admin/users_auth")
async def admin_users_auth(req: Request, limit: int = 1000):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 1000
    limit = max(1, min(limit, 5000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    rows = conn.execute(
        f"SELECT id, user_id, username, email, phone, created_at, updated_at FROM users_auth ORDER BY created_at DESC LIMIT {limit}"
    ).fetchall()
    conn.close()
    return {"users_auth": [dict(r) for r in rows]}

@app.post("/admin/user/reset_password")
async def admin_reset_password(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    data = await req.json()
    user_id = (data.get("user_id") or "").strip()
    identifier = (data.get("identifier") or "").strip()  # username æˆ– email
    new_password = (data.get("new_password") or "").strip()
    if not new_password or len(new_password) < 6:
        return JSONResponse(status_code=400, content={"error": "weak_password", "message": "å¯†ç¢¼è‡³å°‘ 6 ç¢¼"})
    if not user_id and not identifier:
        return JSONResponse(status_code=400, content={"error": "missing_identifier"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        if user_id:
            row = conn.execute("SELECT id, user_id, username, email FROM users_auth WHERE user_id=?", (user_id,)).fetchone()
        else:
            row = conn.execute("SELECT id, user_id, username, email FROM users_auth WHERE username=? OR email=?", (identifier, identifier)).fetchone()
        if not row:
            conn.close()
            return JSONResponse(status_code=404, content={"error": "user_not_found"})
        conn.execute("UPDATE users_auth SET password_hash=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (hash_password(new_password), row["id"]))
        # ç¨½æ ¸è¨˜éŒ„
        try:
            _tok = req.headers.get("x-admin-token") or req.query_params.get("token") or ""
            admin_hash = hashlib.sha256(_tok.encode("utf-8")).hexdigest() if _tok else None
            conn.execute(
                "INSERT INTO admin_audit_logs (action, admin_token_hash, target_user_id, details) VALUES (?, ?, ?, ?)",
                ("reset_password", admin_hash, row["user_id"], json.dumps({"username": row["username"], "email": row["email"]}, ensure_ascii=False))
            )
        except Exception as _e:
            print("[audit] write failed:", _e)
        conn.commit(); conn.close()
        return {"ok": True}
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/admin/requests_full")
async def admin_requests_full(req: Request, limit: int = 200, user_id: str | None = None, mode: str | None = None, date_from: str | None = None, date_to: str | None = None):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    conditions = []
    params = []
    if user_id:
        conditions.append("user_id = ?"); params.append(user_id)
    if mode:
        conditions.append("mode = ?"); params.append(mode)
    if date_from:
        conditions.append("date(created_at) >= date(?)"); params.append(date_from)
    if date_to:
        conditions.append("date(created_at) <= date(?)"); params.append(date_to)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    rows = conn.execute(
        f"SELECT id, created_at, user_id, mode, user_input, response_json FROM requests {where} ORDER BY id DESC LIMIT {limit}",
        params,
    ).fetchall()
    conn.close()
    return {"items": [dict(r) for r in rows]}

@app.get("/admin/messages")
async def admin_messages(req: Request, user_id: str | None = None, session_id: str | None = None, limit: int = 200):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    try:
        limit = int(limit)
    except Exception:
        limit = 200
    limit = max(1, min(limit, 2000))
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        conditions = []
        params = []
        if user_id:
            conditions.append("s.user_id = ?"); params.append(user_id)
        if session_id:
            conditions.append("m.session_id = ?"); params.append(session_id)
        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
        rows = conn.execute(
            f"""
            SELECT m.id, m.session_id, s.user_id, s.agent_type, m.role, m.content, m.timestamp
            FROM messages m
            LEFT JOIN sessions s ON s.session_id = m.session_id
            {where}
            ORDER BY m.id DESC
            LIMIT {limit}
            """,
            params,
        ).fetchall()
        conn.close()
        return {"items": [dict(r) for r in rows]}
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/admin/analytics")
async def admin_analytics(req: Request):
    if not _check_admin(req):
        return JSONResponse(status_code=403, content={"error": "forbidden"})
    conn = get_conn(); conn.row_factory = sqlite3.Row
    try:
        total_users = conn.execute("SELECT COUNT(1) AS c FROM users").fetchone()["c"]
        total_requests = conn.execute("SELECT COUNT(1) AS c FROM requests").fetchone()["c"]
        # ä»Šæ—¥è«‹æ±‚
        today = conn.execute("SELECT COUNT(1) AS c FROM requests WHERE date(created_at) = date('now','localtime')").fetchone()["c"]
        # è¿‘7æ—¥
        last7 = conn.execute(
            """
            SELECT strftime('%Y-%m-%d', created_at) AS d, COUNT(1) AS c
            FROM requests
            WHERE date(created_at) >= date('now','localtime','-6 day')
            GROUP BY d ORDER BY d ASC
            """
        ).fetchall()
        last7d = [{"date": r["d"], "count": r["c"]} for r in last7]
        # æ¨¡å¼åˆ†ä½ˆ
        by_mode_rows = conn.execute("SELECT COALESCE(mode,'') AS mode, COUNT(1) AS c FROM requests GROUP BY COALESCE(mode,'')").fetchall()
        by_mode = { (r["mode"] or ""): r["c"] for r in by_mode_rows }
        # agent åˆ†ä½ˆï¼ˆsessionsï¼‰
        by_agent_rows = conn.execute("SELECT agent_type, COUNT(1) AS c FROM sessions GROUP BY agent_type").fetchall()
        by_agent = { r["agent_type"]: r["c"] for r in by_agent_rows }
        # è¿‘7æ—¥ agent ä½¿ç”¨æ¬¡æ•¸ï¼ˆä¾ sessions.created_atï¼‰
        agent_daily_rows = conn.execute(
            """
            SELECT strftime('%Y-%m-%d', created_at) AS d, agent_type, COUNT(1) AS c
            FROM sessions
            WHERE date(created_at) >= date('now','localtime','-6 day')
            GROUP BY d, agent_type
            ORDER BY d ASC
            """
        ).fetchall()
        agent_daily = {}
        for r in agent_daily_rows:
            agent_daily.setdefault(r["d"], {})[r["agent_type"]] = r["c"]
        # è¨Šæ¯ç¸½æ•¸/ä»Šæ—¥
        total_messages = conn.execute("SELECT COUNT(1) AS c FROM messages").fetchone()["c"]
        today_messages = conn.execute("SELECT COUNT(1) AS c FROM messages WHERE date(timestamp) = date('now','localtime')").fetchone()["c"]
        conn.close()
        return {
            "total_users": total_users,
            "total_requests": total_requests,
            "today_requests": today,
            "last7d": last7d,
            "by_mode": by_mode,
            "by_agent": by_agent,
            "total_messages": total_messages,
            "today_messages": today_messages,
            "agent_daily": agent_daily,
        }
    except Exception as e:
        try: conn.close()
        except Exception: pass
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(admin_session: str | None = Cookie(default=None)):
    if not (admin_session and verify_admin_session_cookie(admin_session)):
        # ç°¡æ˜“ç™»å…¥é 
        return HTMLResponse(content="""
<!DOCTYPE html>
<html lang=\"zh-Hant\"><head><meta charset=\"utf-8\"/><meta name=\"viewport\" content=\"width=device-width, initial-scale=1\"/><title>AIJob ç®¡ç†ç™»å…¥</title>
<style>body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,'Noto Sans TC',sans-serif;margin:40px;color:#111;background:#f6f7fb}
.card{max-width:360px;margin:0 auto;border:1px solid #e5e7eb;border-radius:12px;background:#fff;padding:16px}
input,button{width:100%;padding:10px;border:1px solid #cbd5e1;border-radius:8px;margin-top:10px}
button{background:#111;color:#fff}
.muted{color:#6b7280;font-size:12px;margin-top:8px}
</style></head><body>
<div class=\"card\"><h2>AIJob ç®¡ç†ç™»å…¥</h2>
<input id=\"u\" placeholder=\"å¸³è™Ÿ\"><input id=\"p\" placeholder=\"å¯†ç¢¼\" type=\"password\">
<button onclick=\"login()\">ç™»å…¥</button>
<div class=\"muted\">åƒ…é™ç®¡ç†è€…ä½¿ç”¨ã€‚ç™»å…¥å¾Œå°‡å»ºç«‹å®‰å…¨çš„ç®¡ç† Sessionã€‚</div></div>
<script>
async function login(){
  const r = await fetch('/admin/login',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({username:document.getElementById('u').value,password:document.getElementById('p').value})});
  const j = await r.json(); if(j&&j.ok){ location.href='/admin'; } else { alert(j.message||'ç™»å…¥å¤±æ•—'); }
}
</script></body></html>
""", status_code=200)
    return """
<!DOCTYPE html>
<html lang=\"zh-Hant\">\n<head>\n  <meta charset=\"utf-8\" />\n  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />\n  <title>AIJobçŸ­å½±éŸ³æ™ºèƒ½-ç®¡ç†å¾Œå°</title>\n  <style>\n    body{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,Arial,'Noto Sans TC',sans-serif;margin:24px;color:#111;background:#fafafa}\n    .wrap{max-width:1200px;margin:0 auto}\n    .grid{display:grid;grid-template-columns:1fr;gap:16px}\n    @media (min-width:960px){.grid{grid-template-columns:1fr 1fr}}\n    .card{background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:16px}\n    .row{display:flex;gap:12px;align-items:center;flex-wrap:wrap}\n    input,button,select{padding:8px 10px;border:1px solid #cbd5e1;border-radius:8px;background:#fff}\n    button.primary{background:#111;color:#fff;border-color:#111}\n    table{border-collapse:collapse;width:100%}\n    th,td{border:1px solid #e5e7eb;padding:8px;text-align:left;vertical-align:top;font-size:13px}\n    th{background:#f8fafc}\n    .muted{color:#6b7280}\n    .right{margin-left:auto}\n    .mono{font-family:ui-monospace,SFMono-Regular,Menlo,Monaco,Consolas,'Liberation Mono','Courier New',monospace}\n    /* æ—¥/å¤œæ¨¡å¼ */\n    body.dark{background:#0f172a;color:#e5e7eb}\n    .dark .card{background:#0b1220;border-color:#1f2937}\n    .dark input,.dark button,.dark select{background:#111827;color:#e5e7eb;border-color:#374151}\n    .dark th,.dark td{border-color:#1f2937}\n    .dark th{background:#0f172a}\n    .dark .muted{color:#9ca3af}\n    .dark .barWrap{background:#111827;border-color:#374151}\n    /* å„€è¡¨æ¿æ¨£å¼ */\n    .kpiGrid{display:grid;grid-template-columns:1fr 1fr;gap:12px}\n    @media (min-width:960px){.kpiGrid{grid-template-columns:repeat(4,1fr)}}\n    .kpiCard{border:1px solid #e5e7eb;border-radius:12px;padding:12px}\n    .kpi-title{font-size:12px;color:#6b7280;margin-bottom:6px}\n    .kpi-value{font-size:28px;font-weight:700}\n    .barWrap{display:flex;height:12px;border-radius:999px;overflow:hidden;background:#f3f4f6;border:1px solid #e5e7eb;margin-top:8px}\n    .bar{height:100%;background:#2563eb}\n    .bar2{height:100%;background:#10b981}\n    /* å½©è‰² KPI */\n    .kpi-blue{background:linear-gradient(180deg,#e8f0ff,#f5f8ff)}\n    .kpi-green{background:linear-gradient(180deg,#e7f8f1,#f3fbf7)}\n    .kpi-purple{background:linear-gradient(180deg,#efe9ff,#f7f4ff)}\n    .kpi-amber{background:linear-gradient(180deg,#fff3d6,#fff8e8)}\n    .dark .kpi-blue{background:linear-gradient(180deg,#0b1a3a,#0f234d)}\n    .dark .kpi-green{background:linear-gradient(180deg,#0d2a21,#123528)}\n    .dark .kpi-purple{background:linear-gradient(180deg,#1a0f3a,#23124d)}\n    .dark .kpi-amber{background:linear-gradient(180deg,#3a2a0b,#4d3612)}\n  </style>\n</head>\n<body>\n  <div class=\"wrap\">\n    <h1 style=\"margin:0 0 16px 0\">AIJobçŸ­å½±éŸ³æ™ºèƒ½-ç®¡ç†å¾Œå°</h1>\n    <div class=\"card row\">\n      <label>API Baseï¼š<input id=\"apiBase\" placeholder=\"ä¾‹å¦‚ï¼šhttps://aijobvideobackend.zeabur.app\" style=\"width:360px\"></label>\n      <label>Admin Tokenï¼š<input id=\"admTok\" placeholder=\"è‹¥å¾Œç«¯æœªè¨­å®šå¯ç•™ç©º\" style=\"width:260px\"></label>\n      <button class=\"primary\" onclick=\"saveConfig()\">å„²å­˜</button>\n      <button id=\"themeBtn\" title=\"åˆ‡æ›æ—¥/å¤œ\">åˆ‡æ›æ—¥/å¤œ</button>\n      <a id=\"dlUsers\" href=\"#\" class=\"right\">ä¸‹è¼‰ Users CSV</a>\n      <a id=\"dlUsage\" href=\"#\" style=\"margin-left:8px\">ä¸‹è¼‰ Usage CSV</a>\n    </div>\n\n    <div class=\"card\" id=\"dashboard\">\n      <div class=\"row\">\n        <strong>å„€è¡¨æ¿</strong>\n        <span class=\"muted\">å³æ™‚çµ±è¨ˆï¼ˆä¾æœ€æ–°è¼‰å…¥è³‡æ–™ï¼‰</span>\n        <button class=\"right\" onclick=\"loadAll()\">é‡æ–°æ•´ç†</button>\n      </div>\n      <div class=\"kpiGrid\" style=\"margin-top:10px\">\n        <div class=\"kpiCard kpi-blue\">\n          <div class=\"kpi-title\">ç¸½ä½¿ç”¨è€…æ•¸</div>\n          <div class=\"kpi-value\" id=\"kpiUsers\">-</div>\n        </div>\n        <div class=\"kpiCard kpi-green\">\n          <div class=\"kpi-title\">ç¸½è«‹æ±‚æ•¸</div>\n          <div class=\"kpi-value\" id=\"kpiTotalReq\">-</div>\n        </div>\n        <div class=\"kpiCard kpi-purple\">\n          <div class=\"kpi-title\">ä»Šæ—¥è«‹æ±‚æ•¸</div>\n          <div class=\"kpi-value\" id=\"kpiToday\">-</div>\n        </div>\n        <div class=\"kpiCard kpi-amber\">\n          <div class=\"kpi-title\">è¿‘7æ—¥è«‹æ±‚</div>\n          <div class=\"kpi-value\" id=\"kpi7d\">-</div>\n        </div>\n      </div>\n\n      <div class=\"grid\" style=\"margin-top:12px\">\n        <div class=\"card\" style=\"padding:12px\">\n          <div class=\"row\"><strong>æ¨¡å¼èˆ‡æ™ºèƒ½é«”åˆ†ä½ˆ</strong><span class=\"muted\">/admin/analytics</span></div>\n          <div style=\"display:flex;gap:12px;flex-wrap:wrap\">\n            <canvas id=\"modeChart\" width=\"260\" height=\"160\" style=\"background:transparent\"></canvas>\n            <canvas id=\"agentChart\" width=\"260\" height=\"160\" style=\"background:transparent\"></canvas>\n          </div>\n          <div class=\"row\" style=\"margin-top:6px\"><span class=\"muted\">modeï¼š</span><span id=\"lblMode\" class=\"mono\"></span><span class=\"muted\" style=\"margin-left:8px\">agentï¼š</span><span id=\"lblAgent\" class=\"mono\"></span></div>\n        </div>\n        <div class=\"card\" style=\"padding:12px\">\n          <div class=\"row\"><strong>è¿‘7æ—¥è¶¨å‹¢</strong><span class=\"muted\">æ¯æ—¥è«‹æ±‚</span></div>\n          <canvas id=\"trendCanvas\" width=\"500\" height=\"120\" style=\"width:100%;height:120px\"></canvas>\n        </div>\n      </div>\n    </div>\n\n    <div class=\"grid\">\n      <div class=\"card\">\n        <div class=\"row\">\n          <strong>å¸³è™Ÿç®¡ç†</strong>\n          <button class=\"right\" onclick=\"loadUsersAuth()\">é‡æ–°è¼‰å…¥</button>\n        </div>\n        <div class=\"row\" style=\"margin-top:8px;gap:8px\">\n          <input id=\"qAuth\" placeholder=\"æœå°‹ username / email / user_id\" style=\"flex:1\">\n          <input id=\"rpUser\" placeholder=\"user_id æˆ– username/email ç”¨æ–¼é‡è¨­å¯†ç¢¼\" style=\"flex:1\">\n          <input id=\"rpPass\" placeholder=\"æ–°å¯†ç¢¼ï¼ˆè‡³å°‘6ç¢¼ï¼‰\" type=\"password\" style=\"width:220px\">\n          <button onclick=\"resetPassword()\">é‡è¨­å¯†ç¢¼</button>\n        </div>\n        <div style=\"overflow:auto;max-height:300px;margin-top:10px\">\n          <table id=\"tblAuth\"><thead><tr>\n            <th>id</th><th>user_id</th><th>username</th><th>email</th><th>phone</th><th>created_at</th>\n          </tr></thead><tbody></tbody></table>\n        </div>\n      </div>\n\n      <div class=\"card\">\n        <div class=\"row\">\n          <strong>ä½¿ç”¨è€…æ¸…å–®</strong>\n          <input id=\"qUser\" placeholder=\"æœå°‹ email / username / user_id\" class=\"right\" style=\"flex:1\" oninput=\"renderUsers()\">\n          <button onclick=\"loadAll()\">é‡æ–°è¼‰å…¥</button>\n        </div>\n        <div style=\"overflow:auto;max-height:420px;margin-top:10px\">\n          <table id=\"tblUsers\"><thead><tr>\n            <th>user_id</th><th>email</th><th>name</th><th>username</th><th>phone</th><th>has_password</th><th>created_at</th>\n          </tr></thead><tbody></tbody></table>\n        </div>\n      </div>\n\n      <div class=\"card\">\n        <div class=\"row\">\n          <strong>æœ€è¿‘è«‹æ±‚</strong>\n          <label>é¡¯ç¤ºæ•¸é‡ï¼š\n            <select id=\"usageLimit\">\n              <option>30</option>\n              <option selected>100</option>\n              <option>200</option>\n            </select>\n          </label>\n          <input id=\"qReq\" placeholder=\"é—œéµå­—ï¼ˆuser_input / modeï¼‰\" class=\"right\" style=\"flex:1\" oninput=\"renderReq()\">\n          <button onclick=\"loadUsage()\">é‡æ–°è¼‰å…¥</button>\n        </div>\n        <div class=\"muted\" id=\"reqSummary\" style=\"margin:8px 0 0 0\"></div>\n        <div style=\"overflow:auto;max-height:420px;margin-top:10px\">\n          <table id=\"tblReq\"><thead><tr>\n            <th>id</th><th>created_at</th><th>mode</th><th>user_input</th>\n          </tr></thead><tbody></tbody></table>\n        </div>\n      </div>\n\n      <div class=\"card\">\n        <div class=\"row\">\n          <strong>è«‹æ±‚/è¨Šæ¯æª¢è¦–</strong>\n          <input id=\"fUserId\" placeholder=\"user_id\" style=\"margin-left:auto\">\n          <input id=\"fSessionId\" placeholder=\"session_id\">\n          <input id=\"fDateFrom\" placeholder=\"èµ·å§‹æ—¥æœŸ YYYY-MM-DD\">\n          <input id=\"fDateTo\" placeholder=\"çµæŸæ—¥æœŸ YYYY-MM-DD\">\n          <select id=\"fMode\"><option value=\"\">mode(å…¨éƒ¨)</option><option>script</option><option>copy</option></select>\n          <select id=\"fAgent\"><option value=\"\">agent(å…¨éƒ¨)</option><option>positioning</option><option>topics</option><option>script</option></select>\n          <button onclick=\"loadInspect()\">æŸ¥è©¢</button>\n          <button onclick=\"downloadJSON()\">ä¸‹è¼‰JSON</button>\n          <button onclick=\"downloadCSV()\">ä¸‹è¼‰CSV</button>\n        </div>\n        <div class=\"grid\" style=\"margin-top:10px\">\n          <div class=\"card\" style=\"padding:12px\">\n            <div class=\"row\"><strong>Messages</strong></div>\n            <div id=\"msgList\" style=\"max-height:260px;overflow:auto\" class=\"mono\"></div>\n          </div>\n          <div class=\"card\" style=\"padding:12px\">\n            <div class=\"row\"><strong>Requests (full)</strong></div>\n            <div id=\"reqFullList\" style=\"max-height:260px;overflow:auto\" class=\"mono\"></div>\n          </div>\n        </div>\n      </div>\n\n      <div class=\"card\">\n        <div class=\"row\"><strong>ç”¨æˆ¶é»æ•¸ / è¨‚å–®</strong><button class=\"right\" onclick=\"loadUsersFull()\">é‡æ–°è¼‰å…¥</button></div>\n        <div class=\"grid\" style=\"margin-top:8px\">\n          <div class=\"card\" style=\"padding:12px\">\n            <div class=\"row\"><strong>Credits</strong></div>\n            <div id=\"creditsList\" class=\"mono\" style=\"max-height:240px;overflow:auto\"></div>\n          </div>\n          <div class=\"card\" style=\"padding:12px\">\n            <div class=\"row\"><strong>Orders</strong></div>\n            <div id=\"ordersList\" class=\"mono\" style=\"max-height:240px;overflow:auto\"></div>\n          </div>\n        </div>\n      </div>\n    </div>\n\n    <div class=\"card\">\n      <div class=\"row\">\n        <strong>Google Sheet é€£å‹•</strong>\n        <span class=\"muted\">ï¼ˆå¯è¤‡è£½ç‚º `=IMPORTDATA(API_BASE & \"/export/google-sheet-flat-v2?limit=500\")`ï¼‰</span>\n      </div>\n      <div class=\"mono\" id=\"gsExample\" style=\"margin-top:8px;font-size:13px\"></div>\n    </div>\n  </div>\n\n  <script>\n    const $ = s => document.querySelector(s);\n    function applyTheme(dark){\n      if (dark){ document.body.classList.add('dark'); } else { document.body.classList.remove('dark'); }\n      localStorage.setItem('ADMIN_DARK', dark? '1':'0');\n      try{ renderDashboard(); }catch(_){ }\n    }\n    (function initTheme(){\n      const dark = localStorage.getItem('ADMIN_DARK') === '1';\n      applyTheme(dark);\n      const btn = document.getElementById('themeBtn');\n      if (btn){ btn.addEventListener('click', ()=> applyTheme(!document.body.classList.contains('dark'))); }\n    })();\n    function getConfig(){\n      return {\n        base: localStorage.getItem('ADMIN_API_BASE') || (location.hostname==='localhost' ? 'http://localhost:8080' : 'https://aijobvideobackend.zeabur.app'),\n        tok: localStorage.getItem('ADMIN_TOKEN') || ''\n      }\n    }\n    function applyConfig(){\n      const cfg = getConfig();\n      $('#apiBase').value = cfg.base;\n      $('#admTok').value = cfg.tok;\n      const q = cfg.tok ? ('?token='+encodeURIComponent(cfg.tok)) : '';\n      $('#dlUsers').href = cfg.base + '/admin/users.csv' + q;\n      $('#dlUsage').href = cfg.base + '/admin/usage.csv' + q;\n      $('#gsExample').textContent = '=IMPORTDATA("' + cfg.base + '/export/google-sheet-flat-v2?limit=500")';\n    }\n    function saveConfig(){\n      const base = ($('#apiBase').value||'').trim();\n      const tok = ($('#admTok').value||'').trim();\n      if (base){ localStorage.setItem('ADMIN_API_BASE', base) }\n      localStorage.setItem('ADMIN_TOKEN', tok)\n      applyConfig();\n      loadAll();\n    }\n\n    let users=[], authMap={}; let latest=[]; let totalRequests=0; let usersAuth=[]; let analytics=null; let lastInspect={messages:[], requests:[]};\n    let credits=[], orders=[];\n\n    async function loadAll(){\n      await Promise.all([loadUsers(), loadUsage(), loadAnalytics()]);\n    }\n    async function loadUsers(){\n      const {base, tok} = getConfig();\n      const hdr = tok? {'x-admin-token':tok} : {};\n      const res = await fetch(base + '/admin/users', {headers: hdr});\n      const u = await res.json();\n      users = (u.users||[]); const ua = u.users_auth||[]; authMap={}; ua.forEach(a=>authMap[a.user_id]=a);\n      renderUsers(); applyConfig();\n    }\n    async function loadUsage(){\n      const {base, tok} = getConfig();\n      const hdr = tok? {'x-admin-token':tok} : {};\n      const limit = Number($('#usageLimit').value||100);\n      const res = await fetch(base + '/admin/usage?limit=' + encodeURIComponent(limit), {headers: hdr});\n      const g = await res.json();\n      latest = g.latest||[]; totalRequests = g.total_requests||0;\n      $('#reqSummary').textContent = 'ç¸½è«‹æ±‚æ•¸ï¼š' + totalRequests + 'ï¼Œæœ¬æ¬¡é¡¯ç¤ºï¼š' + latest.length;\n      renderReq(); renderDashboard(); applyConfig();\n    }\n\n    async function loadAnalytics(){\n      const {base, tok} = getConfig();\n      const hdr = tok? {'x-admin-token':tok} : {};\n      const res = await fetch(base + '/admin/analytics', {headers: hdr});\n      analytics = await res.json();\n      try{ renderDashboard(); }catch(_){ }\n    }\n\n    async function loadUsersFull(){\n      const {base, tok} = getConfig();\n      const hdr = tok? {'x-admin-token':tok} : {};\n      const data = await fetch(base + '/admin/users_full', {headers: hdr}).then(r=>r.json());\n      credits = data.credits||[]; orders = data.orders||[];\n      renderBilling();\n    }\n\n    function renderUsers(){\n      const q = ($('#qUser').value||'').toLowerCase();\n      const tb = $('#tblUsers tbody'); tb.innerHTML='';\n      users.filter(u=>{\n        const a = authMap[u.user_id]||{};\n        const hay = [u.user_id,u.email,u.name,a.username,a.phone].join(' ').toLowerCase();\n        return !q || hay.includes(q);\n      }).slice(0,500).forEach(u=>{\n        const a = authMap[u.user_id]||{};\n        const tr = document.createElement('tr');\n        tr.innerHTML = `<td>${u.user_id}</td><td>${u.email||''}</td><td>${u.name||''}</td><td>${a.username||''}</td><td>${a.phone||''}</td><td>${a.has_password? 'Yes':'No'}</td><td>${u.created_at||''}</td>`;\n        tb.appendChild(tr);\n      });\n    }\n\n    function renderReq(){\n      const q = ($('#qReq').value||'').toLowerCase();\n      const tb = $('#tblReq tbody'); tb.innerHTML='';\n      latest.filter(r=>{\n        const hay = [r.user_input,r.mode].join(' ').toLowerCase();\n        return !q || hay.includes(q);\n      }).forEach(r=>{\n        const tr = document.createElement('tr');\n        tr.innerHTML = `<td>${r.id}</td><td>${r.created_at}</td><td>${r.mode}</td><td>${(r.user_input||'').slice(0,300)}</td>`;\n        tb.appendChild(tr);\n      });\n    }\n\n    function parseTime(s){\n      if (!s) return new Date();\n      let d = new Date(s);\n      if (isNaN(d)) d = new Date((s||'').replace(' ', 'T'));\n      if (isNaN(d)) {\n        try{\n          const [datePart,timePart] = (s||'').split(' ');\n          const [y,m,dd] = datePart.split('-').map(Number);\n          const [hh,mm,ss] = (timePart||'0:0:0').split(':').map(Number);\n          d = new Date(y, (m||1)-1, dd||1, hh||0, mm||0, ss||0);\n        }catch(_){ d = new Date(); }\n      }\n      return d;\n    }\n\n    function renderDashboard(){\n      const a = analytics||{};\n      // KPI æ•¸å­—ï¼ˆå„ªå…ˆç”¨ /admin/analyticsï¼‰\n      $('#kpiUsers').textContent = String(a.total_users ?? users.length ?? 0);\n      $('#kpiTotalReq').textContent = String(a.total_requests ?? totalRequests ?? 0);\n\n      const now = new Date();\n      const startToday = new Date(now.getFullYear(), now.getMonth(), now.getDate());\n      const start7d = new Date(startToday.getTime() - 6*24*3600*1000);\n\n      const byDay = {};\n      if (a.last7d && Array.isArray(a.last7d)){\n        for (const it of a.last7d){ byDay[it.date] = it.count; }\n      } else {\n        for (const r of latest){\n          const t = parseTime(r.created_at);\n          const key = t.getFullYear()+'-'+String(t.getMonth()+1).padStart(2,'0')+'-'+String(t.getDate()).padStart(2,'0');\n          byDay[key] = (byDay[key]||0)+1;\n        }\n      }\n      $('#kpiToday').textContent = String(a.today_requests ?? 0);\n      $('#kpi7d').textContent = String(Object.values(byDay).reduce((s,v)=>s+v,0));\n\n      // æ¨¡å¼èˆ‡æ™ºèƒ½é«”åˆ†ä½ˆ\n      const modeData = a.by_mode || {}; // e.g. {script:100, copy:60}\n      const agentData = a.by_agent || {}; // e.g. {positioning:20, topics:10, script:30}\n      drawPieOrBar('modeChart', modeData);\n      drawPieOrBar('agentChart', agentData);\n      $('#lblMode').textContent = JSON.stringify(modeData);\n      $('#lblAgent').textContent = JSON.stringify(agentData);\n\n      // è¿‘7æ—¥è¶¨å‹¢\n      const days=[]; for(let i=6;i>=0;i--){ const d=new Date(startToday.getTime()-i*24*3600*1000); const key=d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0'); days.push({key, val: byDay[key]||0}); }\n      drawTrend(days);\n    }\n\n    function drawTrend(points){\n      const c = document.getElementById('trendCanvas'); if (!c) return;\n      const ctx = c.getContext('2d');\n      const W = c.width, H = c.height;\n      ctx.clearRect(0,0,W,H);\n      const dark = document.body.classList.contains('dark');\n      // è»¸ç·š\n      ctx.strokeStyle = dark? '#1f2937' : '#e5e7eb'; ctx.lineWidth = 1;\n      for(let i=0;i<4;i++){ const y = H-10 - i*((H-30)/3); ctx.beginPath(); ctx.moveTo(35, y); ctx.lineTo(W-10, y); ctx.stroke(); }\n      // è³‡æ–™\n      const maxV = Math.max(1, ...points.map(p=>p.val));\n      const dx = (W-60)/(points.length-1);\n      ctx.strokeStyle = dark? '#38bdf8' : '#2563eb'; ctx.lineWidth = 2; ctx.beginPath();\n      points.forEach((p,idx)=>{\n        const x = 35 + idx*dx;\n        const y = H-10 - (p.val/maxV)*(H-30);\n        if(idx===0) ctx.moveTo(x,y); else ctx.lineTo(x,y);\n      });\n      ctx.stroke();\n      // é»èˆ‡æ¨™ç±¤\n      ctx.fillStyle = dark? '#38bdf8' : '#2563eb';\n      points.forEach((p,idx)=>{\n        const x = 35 + idx*dx;\n        const y = H-10 - (p.val/maxV)*(H-30);\n        ctx.beginPath(); ctx.arc(x,y,3,0,Math.PI*2); ctx.fill();\n      });\n    }\n\n    function drawPieOrBar(canvasId, data){\n      const c = document.getElementById(canvasId); if(!c) return;\n      const ctx = c.getContext('2d');\n      const W = c.width, H = c.height; ctx.clearRect(0,0,W,H);\n      const dark = document.body.classList.contains('dark');\n      const keys = Object.keys(data||{}); const vals = keys.map(k=>data[k]||0);\n      const total = vals.reduce((s,v)=>s+v,0) || 1;\n      // è‹¥é …ç›®<=3 ç”¨åœ“é¤…ï¼Œå¦å‰‡ç”¨ç›´æ¢\n      if (keys.length<=3){\n        let start=0; const colors=['#2563eb','#10b981','#f59e0b','#8b5cf6','#ef4444'];\n        keys.forEach((k,i)=>{\n          const ang = (vals[i]/total)*Math.PI*2; ctx.beginPath(); ctx.moveTo(W/2,H/2); ctx.fillStyle=colors[i%colors.length]; ctx.arc(W/2,H/2, Math.min(W,H)/2-8, start, start+ang); ctx.closePath(); ctx.fill(); start+=ang;\n        });\n      } else {\n        const maxV = Math.max(1, ...vals); const bw = Math.max(18, Math.floor((W-20)/keys.length)-6);\n        keys.forEach((k,i)=>{\n          const x = 10 + i*(bw+6); const h = Math.round(((vals[i]/maxV)*(H-30)));\n          ctx.fillStyle = ['#2563eb','#10b981','#f59e0b','#8b5cf6','#ef4444'][i%5];\n          ctx.fillRect(x, H-10-h, bw, h);\n        });\n      }\n    }\n\n    function renderBilling(){\n      const cDiv = document.getElementById('creditsList');\n      const oDiv = document.getElementById('ordersList');\n      if (cDiv){\n        cDiv.innerHTML = (credits||[]).map(x=>`<div>${x.user_id}: balance=${x.balance} (upd:${x.updated_at||''})</div>`).join('');\n      }\n      if (oDiv){\n        oDiv.innerHTML = (orders||[]).map(x=>`<div>#${x.id} ${x.user_id} ${x.order_type} ${x.amount} ${x.plan||''} [${x.status}] ${x.created_at}</div>`).join('');\n      }\n    }\n\n    async function loadUsersAuth(){\n      const {base, tok} = getConfig();\n      const hdr = tok? {'x-admin-token':tok} : {};\n      const res = await fetch(base + '/admin/users_auth?limit=1000', {headers: hdr});\n      const data = await res.json();\n      usersAuth = data.users_auth||[]; renderUsersAuth();\n    }\n    function renderUsersAuth(){\n      const q = ($('#qAuth').value||'').toLowerCase();\n      const tb = document.querySelector('#tblAuth tbody'); if (!tb) return; tb.innerHTML='';\n      usersAuth.filter(a=>{\n        const hay = [a.id,a.user_id,a.username,a.email].join(' ').toLowerCase();\n        return !q || hay.includes(q);\n      }).slice(0,1000).forEach(a=>{\n        const tr = document.createElement('tr');\n        tr.innerHTML = `<td>${a.id}</td><td>${a.user_id}</td><td>${a.username||''}</td><td>${a.email||''}</td><td>${a.phone||''}</td><td>${a.created_at||''}</td>`;\n        tb.appendChild(tr);\n      });\n    }\n    async function resetPassword(){\n      const {base, tok} = getConfig();\n      const idOrUser = ($('#rpUser').value||'').trim();\n      const newPass = ($('#rpPass').value||'').trim();\n      if (!idOrUser || !newPass || newPass.length<6){ alert('è«‹è¼¸å…¥ user_id æˆ– username/emailï¼Œä¸¦è¨­å®šè‡³å°‘6ç¢¼çš„æ–°å¯†ç¢¼'); return; }\n      const hdr = {'Content-Type':'application/json'}; if (tok) hdr['x-admin-token']=tok;\n      const body = JSON.stringify(/@/.test(idOrUser)||/\\D/.test(idOrUser)? {identifier:idOrUser,new_password:newPass} : {user_id:idOrUser,new_password:newPass});\n      const res = await fetch(base + '/admin/user/reset_password',{method:'POST',headers:hdr,body});\n      const j = await res.json(); if (j && j.ok){ alert('é‡è¨­æˆåŠŸ'); $('#rpPass').value=''; } else { alert('é‡è¨­å¤±æ•—ï¼š'+(j.message||j.error||'unknown')); }\n    }\n\n    async function loadInspect(){\n      const {base, tok} = getConfig();\n      const hdr = tok? {'x-admin-token':tok} : {};\n      const uid = ($('#fUserId').value||'').trim();\n      const sid = ($('#fSessionId').value||'').trim();\n      const dFrom = ($('#fDateFrom').value||'').trim();\n      const dTo = ($('#fDateTo').value||'').trim();\n      const mode = ($('#fMode').value||'').trim();\n      const agent = ($('#fAgent').value||'').trim();\n      // messages\n      const murl = new URL(base + '/admin/messages');\n      if (uid) murl.searchParams.set('user_id', uid);\n      if (sid) murl.searchParams.set('session_id', sid);\n      if (agent) murl.searchParams.set('agent_type', agent);\n      if (dFrom) murl.searchParams.set('date_from', dFrom);\n      if (dTo) murl.searchParams.set('date_to', dTo);\n      const mres = await fetch(murl.toString(), {headers: hdr});\n      const mdata = await mres.json();\n      let messages = mdata.items||[];\n\n      // requests_full\n      const rurl = new URL(base + '/admin/requests_full');\n      rurl.searchParams.set('limit','200');\n      if (uid) rurl.searchParams.set('user_id', uid);\n      if (mode) rurl.searchParams.set('mode', mode);\n      if (dFrom) rurl.searchParams.set('date_from', dFrom);\n      if (dTo) rurl.searchParams.set('date_to', dTo);\n      const rres = await fetch(rurl.toString(), {headers: hdr});\n      const rdata = await rres.json();\n      let reqs = rdata.items||[];\n\n      lastInspect = {messages, requests:reqs};\n      renderInspect();\n    }\n    function renderInspect(){\n      const mDiv = $('#msgList'); const rDiv = $('#reqFullList');\n      mDiv.innerHTML = lastInspect.messages.map(m=> `<div>[${m.timestamp}] (${m.agent_type||''}) ${m.user_id} ${m.role}: ${escapeHtml(m.content||'')}</div>`).join('');\n      rDiv.innerHTML = lastInspect.requests.map(r=> `<details><summary>#${r.id} ${r.created_at} ${r.user_id?('('+r.user_id+')'):''} [${r.mode}]</summary><pre>${escapeHtml(r.user_input||'')}</pre><pre>${escapeHtml(r.response_json||'')}</pre></details>`).join('');\n    }\n    function downloadJSON(){\n      const blob = new Blob([JSON.stringify(lastInspect,null,2)], {type:'application/json'});\n      const a = document.createElement('a'); a.href = URL.createObjectURL(blob); a.download = 'inspect.json'; a.click();\n    }\n    async function downloadCSV(){\n      const {base, tok} = getConfig();\n      const hdr = tok? {'x-admin-token':tok} : {};\n      const uid = ($('#fUserId').value||'').trim();\n      const dFrom = ($('#fDateFrom').value||'').trim();\n      const dTo = ($('#fDateTo').value||'').trim();\n      const mode = ($('#fMode').value||'').trim();\n      const url = new URL(base + '/admin/requests_full.csv');\n      if (uid) url.searchParams.set('user_id', uid);\n      if (mode) url.searchParams.set('mode', mode);\n      if (dFrom) url.searchParams.set('date_from', dFrom);\n      if (dTo) url.searchParams.set('date_to', dTo);\n      const res = await fetch(url.toString(), {headers: hdr});\n      const blob = await res.blob();\n      const a = document.createElement('a'); a.href = URL.createObjectURL(blob); a.download = 'requests_full.csv'; a.click();\n    }\n\n    // äºŒæ¬¡ç¢ºèªï¼šé‡è¨­å¯†ç¢¼\n    const _origResetPassword = resetPassword;\n    resetPassword = async function(){\n      const idOrUser = ($('#rpUser').value||'').trim();\n      const newPass = ($('#rpPass').value||'').trim();\n      if (!idOrUser || !newPass){ return _origResetPassword(); }\n      if (!confirm(`ç¢ºèªè¦é‡è¨­å¸³è™Ÿã€Œ${idOrUser}ã€çš„å¯†ç¢¼å—ï¼Ÿ`)) return;\n      await _origResetPassword();\n    }\n    function escapeHtml(s){ return (s||'').replace(/[&<>"]\/g, c=> ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;','\'':'&#39;'}[c])); }\n\n    // åˆå§‹\n    applyConfig();\n    loadAll();\n  </script>\n</body>\n</html>
"""

# === Admin Login/Logout ===
@app.post("/admin/login")
async def admin_login(req: Request):
    try:
        data = await req.json()
    except Exception:
        return JSONResponse(status_code=400, content={"error": "bad_request"})
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not ADMIN_USER or not ADMIN_PASSWORD:
        return JSONResponse(status_code=500, content={"error": "admin_not_configured", "message": "å°šæœªè¨­å®š ADMIN_USER/ADMIN_PASSWORD"})
    if username != ADMIN_USER or password != ADMIN_PASSWORD:
        return JSONResponse(status_code=401, content={"error": "invalid_credentials", "message": "å¸³è™Ÿæˆ–å¯†ç¢¼éŒ¯èª¤"})
    token = create_admin_session_cookie(username)
    resp = JSONResponse({"ok": True})
    # Cookie å±¬æ€§ï¼šHttpOnly+Secure+SameSite=Laxï¼Œå­˜æ´» 12 å°æ™‚
    resp.set_cookie("admin_session", token, httponly=True, secure=True, samesite="none", max_age=12*3600)
    return resp

@app.post("/admin/logout")
async def admin_logout():
    resp = JSONResponse({"ok": True})
    resp.delete_cookie("admin_session")
    return resp

@app.get("/admin/healthz")
async def admin_healthz(req: Request):
    has_session = False
    try:
        c = req.cookies.get("admin_session")
        has_session = bool(c and verify_admin_session_cookie(c))
    except Exception:
        has_session = False
    return {
        "ok": True,
        "admin_ready": bool(ADMIN_USER and ADMIN_PASSWORD),
        "oauth_ready": bool('_OAUTH_READY' in globals() and _OAUTH_READY),
        "has_admin_session": has_session,
    }

# ========= ä¸‰æ™ºèƒ½é«” API ç«¯é» =========
# çµ±ä¸€èŠå¤©ç«¯é»ï¼ˆè‡ªç„¶å°è«‡ + KB + è¨˜æ†¶ + äººè¨­ï¼‰
AGENT_PERSONAS = {
    "positioning": "ä½ æ˜¯å°ˆæ¥­çš„å½±éŸ³å®šä½é¡§å•ï¼Œèªæ°£å°ˆæ¥­ã€æ¸…æ¥šä½†å£èªï¼Œé¿å…åˆ¶å¼é …ç›®æ¸…å–®ã€‚",
    "topics": "ä½ æ˜¯å°ˆæ¥­çš„çˆ†æ¬¾çŸ­å½±éŸ³é¸é¡Œé¡§å•ï¼Œå–„ç”¨ç†±é»èˆ‡ä½¿ç”¨è€…å®šä½ï¼Œçµ¦å…·é«”å¯æ“ä½œå»ºè­°ã€‚",
    "script": "ä½ æ˜¯å°ˆæ¥­çš„AIè…³æœ¬æ’°å¯«å¯«æ‰‹ï¼Œå›è¦†è‡ªç„¶ä¸”å…·é«”ï¼Œå¿…è¦æ™‚ä¸»å‹•è©¢å•è£œå……è³‡è¨Šã€‚",
}

def _mem_agent_key(agent_type: str) -> str:
    if agent_type == "positioning":
        return "positioning"
    if agent_type == "topics":
        return "topic_selection"
    return "script_copy"

@app.post("/chat")
async def chat(req: Request):
    """çµ±ä¸€èŠå¤©ï¼šè‡ªç„¶å°è«‡ï¼Œå¸¶å…¥ç”¨æˆ¶æª”æ¡ˆ/è¨˜æ†¶/çŸ¥è­˜åº«ã€‚"""
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_id = (data.get("user_id") or "").strip()
    agent_type = (data.get("agent_type") or "script").strip()
    messages = data.get("messages") or []
    template_type = (data.get("template_type") or "").strip().upper() or None
    duration = data.get("duration")

    if not user_id:
        raise HTTPException(status_code=400, detail="user_id is required")

    # ç¢ºä¿ç”¨æˆ¶å­˜åœ¨
    create_or_get_user(user_id)

    # è®€å–æª”æ¡ˆèˆ‡è¨˜æ†¶
    user_profile = get_user_profile(user_id)
    memories = get_user_memories(user_id, agent_type=_mem_agent_key(agent_type), limit=8)

    # å»ºæœƒè©±
    session_id = data.get("session_id") or create_session(user_id, agent_type)

    # å°‡æœ€è¿‘ä¸€å‰‡ user è¨Šæ¯åŠ å…¥è¨Šæ¯è¡¨
    last_user = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            last_user = (m.get("content") or "").strip()
            break
    if last_user:
        add_message(session_id, "user", last_user)

    # äººè¨­èˆ‡ KB ground
    persona = AGENT_PERSONAS.get(agent_type, AGENT_PERSONAS["script"])
    kb_ctx = retrieve_context(last_user) if last_user else ""
    kb_all = (EXTRA_KB or "").strip()

    # å¯é¸ï¼šæŠŠæ¨¡æ¿/æ™‚é•·é™„åŠ åˆ°ä¸Šä¸‹æ–‡
    script_hint = ""
    if agent_type == "script":
        if template_type:
            script_hint += f"\nã€æŒ‡å®šæ¨¡æ¿ã€‘{template_type}"
        if duration:
            try:
                script_hint += f"\nã€æŒ‡å®šæ™‚é•·ã€‘{int(duration)} ç§’"
            except Exception:
                pass

    system_ctx = (
        f"{persona}\nè«‹ä»¥è‡ªç„¶ä¸­æ–‡å°è«‡ï¼Œä¸ç”¨åˆ¶å¼æ¸…å–®ã€‚è‹¥èƒ½å¾çŸ¥è­˜åº«æˆ–ç”¨æˆ¶æª”æ¡ˆå¾—åˆ°ç­”æ¡ˆï¼Œè«‹å„ªå…ˆçµåˆã€‚\n\n"
        f"ã€é‡è¦æ ¼å¼è¦æ±‚ã€‘\n"
        f"â€¢ ä½¿ç”¨emojiä½œç‚ºåˆ†é»ç¬¦è™Ÿï¼Œè®“å…§å®¹æ›´æ˜“è®€\n"
        f"â€¢ æ®µè½åˆ†æ˜ï¼Œé‡é»çªå‡º\n"
        f"â€¢ åŸºæ–¼çŸ¥è­˜åº«å…§å®¹æä¾›å°ˆæ¥­å»ºè­°\n"
        f"â€¢ å›æ‡‰çµæ§‹ï¼šğŸ“ ä¸»è¦è§€é» â†’ ğŸ’¡ å…·é«”å»ºè­° â†’ âœ¨ å¯¦ä½œè¦é» â†’ ğŸ¯ è¡Œå‹•æŒ‡å¼•\n\n"
        f"ã€ç”¨æˆ¶æª”æ¡ˆï¼ˆè‹¥ç©ºä»£è¡¨æœªè¨­å®šï¼‰ã€‘\n{json.dumps(user_profile or {}, ensure_ascii=False)}\n\n"
        f"ã€ç›¸é—œè¨˜æ†¶ï¼ˆç¯€é¸ï¼‰ã€‘\n" + "\n".join([f"- {m.get('content','')}" for m in memories[:5]]) + "\n\n"
        f"ã€å…¨åŸŸçŸ¥è­˜æ‘˜è¦ï¼ˆæˆªæ–·ï¼‰ã€‘\n{kb_all[:1200]}\n\n"
        f"ã€KBå‹•æ…‹æ“·å–ã€‘\n{(kb_ctx or '')[:800]}\n" 
        f"{script_hint}\n"
    )

    # ç”¢ç”Ÿå›è¦†
    if use_gemini():
        prompt = (
            system_ctx + "\n---\n" + (last_user or "") + "\n\nè«‹ä»¥å°è«‡å½¢å¼å›è¦†ï¼Œé¿å…é‡è¦†ä½¿ç”¨ç›¸åŒå¥å‹ã€‚ä½¿ç”¨emojiåˆ†æ®µï¼Œè®“å…§å®¹æ›´æ˜“è®€ã€‚"
        )
        ai_response = gemini_generate_text(prompt)
    else:
        # ç„¡æ¨¡å‹çš„è‡ªç„¶å›è¦†ï¼ˆè¼ƒå¿«ï¼‰
        if agent_type == "positioning":
            ai_response = natural_fallback_positioning(last_user, user_profile, memories)
        elif agent_type == "topics":
            base = last_user or "è«‹æä¾›ä»Šæ—¥çš„é¸é¡Œéˆæ„Ÿ"
            ai_response = (
                "ä»¥ä¸‹æ˜¯ä¾ä½ çš„å®šä½èˆ‡è¿‘æœŸæ´å¯Ÿçµ¦çš„é¸é¡Œæ–¹å‘ï¼ˆå¯å›æˆ‘è¦å“ªå€‹å±•é–‹ï¼‰ï¼š\n\n"
                "1) ç†±é»ï¼‹ä½ ç”¢å“çš„é—œè¯åˆ‡å…¥\n"
                "2) å—çœ¾å¸¸è¦‹ç—›é»çš„å¿«é€Ÿè§£æ³•\n"
                "3) ä½¿ç”¨å‰/å¾Œå°æ¯”æ¡ˆä¾‹\n"
                "4) 30 ç§’å¾®æ•™å­¸ + è¡Œå‹•å‘¼ç±²\n"
                "5) è¿·ä½ è¨ªè«‡/QA å›è¦†ç•™è¨€\n\n"
                f"ä½ å‰›æåˆ°ï¼š{base[:80]}â€¦ æˆ‘å»ºè­°å…ˆå¾ 2) æˆ– 4) é–‹å§‹ã€‚"
            )
        else:  # script
            ai_response = (
                "äº†è§£ï¼Œæˆ‘æœƒç”¨è‡ªç„¶å£å»é™ªä½ è¨è«–è…³æœ¬ã€‚å…ˆèªªæ˜ä½ çš„ä¸»é¡Œã€å¹³å°èˆ‡ç›®æ¨™ï¼Œæˆ‘å†çµ¦ä½ ç¬¬ä¸€ç‰ˆçµæ§‹èˆ‡é–‹å ´ã€‚"
            )

    add_message(session_id, "assistant", ai_response)

    # å˜—è©¦æŠ½å–ä¸¦æ›´æ–°å®šä½æª”æ¡ˆï¼ˆåªé‡å°å®šä½ï¼‰
    if agent_type == "positioning":
        try:
            draft = {}
            draft.update(extract_profile_fields(last_user))
            draft.update(extract_profile_fields(ai_response))
            draft = {k:v for k,v in draft.items() if v}
            if draft:
                update_user_profile(user_id, draft)
                user_profile = get_user_profile(user_id)
        except Exception as e:
            print("[/chat] profile extract failed:", e)

    return {
        "session_id": session_id,
        "assistant_message": ai_response,
        "user_profile": user_profile if agent_type == "positioning" else None,
        "error": None
    }

# === NEW: æµå¼èŠå¤©ç«¯é» ===
from fastapi import BackgroundTasks

@app.post("/chat_stream")
async def chat_stream(req: Request):
    try:
        data = await req.json()
    except Exception:
        raise HTTPException(status_code=422, detail="invalid_json")

    user_id = (data.get("user_id") or "").strip()
    agent_type = (data.get("agent_type") or "script").strip()
    messages = data.get("messages") or []
    template_type = (data.get("template_type") or "").strip().upper() or None
    duration = data.get("duration")

    if not user_id:
        raise HTTPException(status_code=400, detail="user_id is required")

    create_or_get_user(user_id)
    user_profile = get_user_profile(user_id)
    memories = get_user_memories(user_id, agent_type=_mem_agent_key(agent_type), limit=8)

    session_id = data.get("session_id") or create_session(user_id, agent_type)

    last_user = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            last_user = (m.get("content") or "").strip()
            break
    if last_user:
        add_message(session_id, "user", last_user)

    persona = AGENT_PERSONAS.get(agent_type, AGENT_PERSONAS["script"])
    kb_ctx = retrieve_context(last_user) if last_user else ""
    kb_all = (EXTRA_KB or "").strip()
    script_hint = ""
    if agent_type == "script":
        if template_type:
            script_hint += f"\nã€æŒ‡å®šæ¨¡æ¿ã€‘{template_type}"
        if duration:
            try:
                script_hint += f"\nã€æŒ‡å®šæ™‚é•·ã€‘{int(duration)} ç§’"
            except Exception:
                pass

    system_ctx = (
        f"{persona}\nè«‹ä»¥è‡ªç„¶ä¸­æ–‡å°è«‡ï¼Œä¸ç”¨åˆ¶å¼æ¸…å–®ã€‚è‹¥èƒ½å¾çŸ¥è­˜åº«æˆ–ç”¨æˆ¶æª”æ¡ˆå¾—åˆ°ç­”æ¡ˆï¼Œè«‹å„ªå…ˆçµåˆã€‚\n" 
        f"ã€ç”¨æˆ¶æª”æ¡ˆï¼ˆè‹¥ç©ºä»£è¡¨æœªè¨­å®šï¼‰ã€‘\n{json.dumps(user_profile or {}, ensure_ascii=False)}\n\n"
        f"ã€ç›¸é—œè¨˜æ†¶ï¼ˆç¯€é¸ï¼‰ã€‘\n" + "\n".join([f"- {m.get('content','')}" for m in memories[:5]]) + "\n\n"
        f"ã€å…¨åŸŸçŸ¥è­˜æ‘˜è¦ï¼ˆæˆªæ–·ï¼‰ã€‘\n{kb_all[:1200]}\n\n"
        f"ã€KBå‹•æ…‹æ“·å–ã€‘\n{(kb_ctx or '')[:800]}\n" 
        f"{script_hint}\n"
    )

    # å–å¾—æœ€è¿‘å°è©±ä»¥å¢å¼·ä¸Šä¸‹æ–‡é€£è²«
    def get_recent_messages(session_id: str, limit: int = 8):
        try:
            conn = get_conn()
            conn.row_factory = sqlite3.Row
            rows = conn.execute(
                "SELECT role, content FROM messages WHERE session_id = ? ORDER BY id DESC LIMIT ?",
                (session_id, limit),
            ).fetchall()
            conn.close()
            return list(reversed([dict(r) for r in rows]))
        except Exception:
            return []

    recent_msgs = get_recent_messages(session_id, 8)

    async def gen():
        # ç°¡æ˜“åˆ‡ç‰‡æµï¼šè‹¥æœ‰æ¨¡å‹å¯é€æ®µé€å‡ºï¼Œå¦å‰‡ä¸€æ¬¡é€å‡ºè‡ªç„¶å›è¦†
        if use_gemini():
            convo = "\n".join([f"{m['role']}: {m['content']}" for m in recent_msgs])
            full = gemini_generate_text(system_ctx + "\n---\n" + (convo or (last_user or "")))
        else:
            if agent_type == "positioning":
                full = natural_fallback_positioning(last_user, user_profile, memories)
            elif agent_type == "topics":
                base = last_user or "è«‹æä¾›ä»Šæ—¥çš„é¸é¡Œéˆæ„Ÿ"
                full = (
                    "ä»¥ä¸‹æ˜¯ä¾ä½ çš„å®šä½èˆ‡è¿‘æœŸæ´å¯Ÿçµ¦çš„é¸é¡Œæ–¹å‘ï¼ˆå¯å›æˆ‘è¦å“ªå€‹å±•é–‹ï¼‰ï¼š\n\n"
                    "1) ç†±é»ï¼‹ä½ ç”¢å“çš„é—œè¯åˆ‡å…¥\n"
                    "2) å—çœ¾å¸¸è¦‹ç—›é»çš„å¿«é€Ÿè§£æ³•\n"
                    "3) ä½¿ç”¨å‰/å¾Œå°æ¯”æ¡ˆä¾‹\n"
                    "4) 30 ç§’å¾®æ•™å­¸ + è¡Œå‹•å‘¼ç±²\n"
                    "5) è¿·ä½ è¨ªè«‡/QA å›è¦†ç•™è¨€\n\n"
                    f"ä½ å‰›æåˆ°ï¼š{base[:80]}â€¦ æˆ‘å»ºè­°å…ˆå¾ 2) æˆ– 4) é–‹å§‹ã€‚"
                )
            else:
                full = "äº†è§£ï¼Œæˆ‘æœƒç”¨è‡ªç„¶å£å»é™ªä½ è¨è«–è…³æœ¬ã€‚å…ˆèªªæ˜ä½ çš„ä¸»é¡Œã€å¹³å°èˆ‡ç›®æ¨™ï¼Œæˆ‘å†çµ¦ä½ ç¬¬ä¸€ç‰ˆçµæ§‹èˆ‡é–‹å ´ã€‚"

        # é€æ®µè¼¸å‡º
        chunk_size = 60
        for i in range(0, len(full), chunk_size):
            yield full[i:i+chunk_size]
        # å®Œæˆå¾Œå¯«å…¥è¨Šæ¯
        add_message(session_id, "assistant", full)

        # å®šä½ï¼šå˜—è©¦æ›´æ–°æª”æ¡ˆä¸¦æŠŠå›è¦†æ‘˜è¦å­˜æˆç­†è¨˜
        if agent_type == "positioning":
            try:
                draft = {}
                draft.update(extract_profile_fields(last_user))
                draft.update(extract_profile_fields(full))
                draft = {k:v for k,v in draft.items() if v}
                if draft:
                    update_user_profile(user_id, draft)
                # å­˜æˆã€Œnoteã€å‹è¨˜æ†¶ï¼Œä¾›å‰ç«¯å³å´ç­†è¨˜æœ¬é¡¯ç¤º
                note = (full or "").strip()
                if note:
                    add_memory(user_id, "positioning", "note", note[:800], importance_score=6)
            except Exception:
                pass
        # é¸é¡Œï¼šæŠŠå›è¦†å­˜æˆç­†è¨˜ä¸¦ä¿å­˜é¸é¡Œå»ºè­°
        elif agent_type == "topics":
            try:
                note = (full or "").strip()
                if note:
                    add_memory(user_id, "topic_selection", "note", note[:800], importance_score=6)
                    
                    # ä¿å­˜é¸é¡Œå»ºè­°åˆ°è³‡æ–™åº«
                    from datetime import date
                    conn = get_conn()
                    conn.execute(
                        """INSERT OR REPLACE INTO topic_suggestions 
                           (user_id, suggested_date, topics, reasoning) 
                           VALUES (?, ?, ?, ?)""",
                        (user_id, date.today().isoformat(), json.dumps({"suggestions": note}), note)
                    )
                    conn.commit()
                    conn.close()
            except Exception as e:
                print(f"[Topics Save Error] {e}")
                pass
        # è…³æœ¬ï¼šæŠŠå›è¦†å­˜æˆç­†è¨˜
        elif agent_type == "script":
            try:
                note = (full or "").strip()
                if note:
                    add_memory(user_id, "script_copy", "note", note[:800], importance_score=6)
            except Exception:
                pass

    return StreamingResponse(gen(), media_type="text/plain")


# å®šä½æ™ºèƒ½é«”
@app.post("/agent/positioning/analyze")
async def positioning_analyze(req: Request):
    """å®šä½æ™ºèƒ½é«”åˆ†æç”¨æˆ¶å®šä½"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        user_input = data.get("user_input", "")
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # ç¢ºä¿ç”¨æˆ¶å­˜åœ¨
        create_or_get_user(user_id)
        
        # ç²å–ç”¨æˆ¶æª”æ¡ˆå’Œç›¸é—œè¨˜æ†¶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="positioning", limit=10)
        
        # å‰µå»ºæœƒè©±
        session_id = create_session(user_id, "positioning")
        add_message(session_id, "user", user_input)
        
        # ç”Ÿæˆåˆ†æ
        analysis_context = positioning_agent_analyze(user_input, user_profile, memories)
        
        # èª¿ç”¨ AI ç”Ÿæˆå›æ‡‰ï¼ˆç„¡æ¨¡å‹æ™‚æä¾›è‡ªç„¶å›è¦†ï¼‰
        if use_gemini():
            ai_response = gemini_generate_text(analysis_context)
        else:
            ai_response = natural_fallback_positioning(user_input, user_profile, memories)
        
        add_message(session_id, "assistant", ai_response)
        
        # æå–é—œéµæ´å¯Ÿä¸¦ä¿å­˜ç‚ºè¨˜æ†¶
        if ai_response and len(ai_response) > 50:
            key_insights = extract_key_insights(ai_response, "positioning")
            for insight in key_insights:
                add_memory(user_id, "positioning", "insight", insight, importance_score=7)

        # å¾ç”¨æˆ¶è¼¸å…¥èˆ‡ AI å›æ‡‰ä¸­å˜—è©¦æ“·å–å®šä½æ¬„ä½ï¼Œæ›´æ–°æª”æ¡ˆï¼ˆè‰ç¨¿ï¼‰
        try:
            draft_fields = {}
            draft_fields.update(extract_profile_fields(user_input))
            draft_fields.update(extract_profile_fields(ai_response))
            # éæ¿¾ç©ºå€¼
            draft_fields = {k:v for k,v in draft_fields.items() if v}
            if draft_fields:
                update_user_profile(user_id, draft_fields)
                # é‡æ–°è®€å–æœ€æ–°æª”æ¡ˆ
                user_profile = get_user_profile(user_id)
        except Exception as _e:
            print("[Positioning] extract_profile_fields failed:", _e)
        
        # ç”Ÿæˆçµæ§‹åŒ–çš„å®šä½æ‘˜è¦ï¼ˆåŒ…å«åŸ·è¡Œå»ºè­°ï¼‰
        positioning_summary = ""
        tone_guidelines = ""
        execution_suggestions = ""
        
        if ai_response:
            # ç°¡å–®è§£æAIå›æ‡‰ï¼Œæå–é—œéµä¿¡æ¯
            lines = ai_response.split('\n')
            for line in lines:
                line = line.strip()
                if 'æ¥­å‹™é¡å‹ï¼š' in line or 'ç›®æ¨™å—çœ¾ï¼š' in line or 'å“ç‰Œèªæ°£ï¼š' in line:
                    positioning_summary += line + "\n"
                elif 'èªæ°£' in line and ('å°ˆæ¥­' in line or 'è¦ªåˆ‡' in line or 'å¹½é»˜' in line or 'æ¬Šå¨' in line):
                    tone_guidelines = line
                elif 'å¯¦ä½œå»ºè­°' in line or 'åŸ·è¡Œ' in line or 'å»ºè­°' in line:
                    execution_suggestions += line + "\n"
        
        # å¦‚æœæ²’æœ‰æå–åˆ°è¶³å¤ ä¿¡æ¯ï¼Œä½¿ç”¨é»˜èªå€¼
        if not positioning_summary:
            positioning_summary = "åŸºæ–¼æ‚¨çš„æè¿°ï¼Œå»ºè­°å»ºç«‹å°ˆæ¥­çš„çŸ­å½±éŸ³å®šä½ç­–ç•¥ã€‚"
        if not tone_guidelines:
            tone_guidelines = "ä½¿ç”¨å°ˆæ¥­è¡“èªï¼Œä¿æŒå®¢è§€ç†æ€§ï¼Œå¼·èª¿æ•¸æ“šå’Œäº‹å¯¦ã€‚"
        if not execution_suggestions:
            execution_suggestions = "å»ºè­°æ¡ç”¨æµé‡å‹èˆ‡è½‰æ›å‹å…§å®¹ 7:3 é…æ¯”ï¼Œæ¯é€±ç™¼å¸ƒ 3-5 æ¬¡ï¼Œå°ˆæ³¨æ–¼ Instagram Reels å¹³å°ã€‚"
        
        return {
            "session_id": session_id,
            "response": ai_response,
            "user_profile": user_profile,
            "positioning_summary": positioning_summary,
            "tone_guidelines": tone_guidelines,
            "execution_suggestions": execution_suggestions,
            "error": None
        }
        
    except Exception as e:
        print(f"[Positioning Agent Error] {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            status_code=500,
            content={
                "error": "internal_server_error", 
                "message": "ä¼ºæœå™¨å…§éƒ¨éŒ¯èª¤ï¼Œè«‹ç¨å¾Œå†è©¦",
                "details": str(e) if "DEBUG" in os.environ else "Internal server error"
            }
        )

@app.put("/agent/positioning/profile")
async def update_positioning_profile(req: Request):
    """æ›´æ–°ç”¨æˆ¶å®šä½æª”æ¡ˆ"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        profile_data = data.get("profile_data", {})
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # ç¢ºä¿ç”¨æˆ¶å­˜åœ¨
        create_or_get_user(user_id)
        
        # æ›´æ–°æª”æ¡ˆ
        success = update_user_profile(user_id, profile_data)
        
        if success:
            # ä¿å­˜æª”æ¡ˆæ›´æ–°ç‚ºè¨˜æ†¶
            add_memory(user_id, "positioning", "profile_update", 
                      f"ç”¨æˆ¶æª”æ¡ˆå·²æ›´æ–°ï¼š{json.dumps(profile_data, ensure_ascii=False)}", 
                      importance_score=8)
        
        return {
            "success": success,
            "message": "æª”æ¡ˆæ›´æ–°æˆåŠŸ" if success else "æª”æ¡ˆæ›´æ–°å¤±æ•—",
            "error": None
        }
        
    except Exception as e:
        print(f"[Profile Update Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# æ–°å¢ï¼šå–å¾—ç”¨æˆ¶å®šä½æª”æ¡ˆèˆ‡ç­†è¨˜ï¼ˆä¾›å‰ç«¯å³å´åŒæ­¥é¡¯ç¤ºï¼‰
@app.get("/agent/positioning/profile")
async def get_positioning_profile(user_id: str, notes_limit: int = 10):
    try:
        profile = get_user_profile(user_id)
        notes = get_user_memories(user_id, agent_type="positioning", memory_type="note", limit=notes_limit)
        return {
            "user_id": user_id,
            "profile": profile or {},
            "notes": notes,
            "error": None
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

# æ–°å¢ï¼šé€šç”¨ç­†è¨˜æŸ¥è©¢ API
@app.get("/agent/notes")
async def get_agent_notes(user_id: str, agent_type: str, memory_type: str = "note", limit: int = 10):
    try:
        notes = get_user_memories(user_id, agent_type=agent_type, memory_type=memory_type, limit=limit)
        return {
            "user_id": user_id,
            "agent_type": agent_type,
            "notes": notes,
            "error": None
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": "internal_server_error", "message": str(e)})

# æ–°å¢ï¼šä¸€éµç”Ÿæˆå®šä½åŠŸèƒ½
@app.post("/agent/positioning/generate")
async def generate_positioning(req: Request):
    """ä¸€éµç”Ÿæˆå®Œæ•´å®šä½æª”æ¡ˆ"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        theme = data.get("theme", "")  # ç”¨æˆ¶æä¾›çš„ä¸»é¡Œ/ç”¢å“/æœå‹™
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        if not theme.strip():
            raise HTTPException(status_code=400, detail="theme is required")
        
        # ç¢ºä¿ç”¨æˆ¶å­˜åœ¨
        create_or_get_user(user_id)
        
        # ç²å–ç¾æœ‰æª”æ¡ˆå’Œè¨˜æ†¶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="positioning", limit=5)
        
        # å‰µå»ºæœƒè©±
        session_id = create_session(user_id, "positioning")
        add_message(session_id, "user", f"ä¸€éµç”Ÿæˆå®šä½ï¼š{theme}")
        
        # æ§‹å»ºä¸€éµç”Ÿæˆæç¤ºè©
        context = f"""ä½ æ˜¯å°ˆæ¥­çš„çŸ­å½±éŸ³å®šä½é¡§å•ï¼Œå°ˆé–€æœå‹™å°ç£å¸‚å ´ï¼Œè«‹æ ¹æ“šç”¨æˆ¶æä¾›çš„ä¸»é¡Œã€Œ{theme}ã€ç”Ÿæˆå®Œæ•´çš„å®šä½æª”æ¡ˆã€‚

è«‹åˆ†æé€™å€‹ä¸»é¡Œä¸¦æä¾›ï¼š
1. æ¥­å‹™é¡å‹ï¼šå…·é«”çš„è¡Œæ¥­åˆ†é¡
2. ç›®æ¨™å—çœ¾ï¼šæ˜ç¢ºçš„å°ç£å—çœ¾ç•«åƒï¼ˆå¹´é½¡ã€è·æ¥­ã€ç—›é»ã€éœ€æ±‚ï¼‰
3. å“ç‰Œèªæ°£ï¼šé©åˆå°ç£ç”¨æˆ¶çš„æºé€šé¢¨æ ¼
4. ä¸»è¦å¹³å°ï¼šå°ç£æœ€é©åˆçš„çŸ­å½±éŸ³å¹³å°ï¼ˆæ¨è–¦ï¼šInstagram Reelsã€TikTokã€YouTube Shortsã€Facebook Reelsï¼‰
5. å…§å®¹ç›®æ¨™ï¼šå…·é«”è¦é”æˆçš„æ•ˆæœ
6. ç™¼æ–‡é »ç‡ï¼šå»ºè­°çš„æ›´æ–°é »ç‡

ã€é‡è¦ã€‘å¹³å°æ¨è–¦è«‹å°ˆæ³¨æ–¼å°ç£ç”¨æˆ¶å¸¸ç”¨çš„å¹³å°ï¼Œé¿å…æ¨è–¦Bç«™ã€å°ç´…æ›¸ç­‰å¤§é™¸å¹³å°ã€‚

è«‹ä»¥çµæ§‹åŒ–æ ¼å¼å›æ‡‰ï¼Œæ¯å€‹æ¬„ä½éƒ½è¦å…·é«”æ˜ç¢ºï¼Œä¾¿æ–¼ç³»çµ±è‡ªå‹•æå–ã€‚

æ ¼å¼ï¼š
æ¥­å‹™é¡å‹ï¼š[å…·é«”åˆ†é¡]
ç›®æ¨™å—çœ¾ï¼š[è©³ç´°æè¿°]
å“ç‰Œèªæ°£ï¼š[é¢¨æ ¼ç‰¹é»]
ä¸»è¦å¹³å°ï¼š[å¹³å°åç¨±]
å…§å®¹ç›®æ¨™ï¼š[å…·é«”ç›®æ¨™]
ç™¼æ–‡é »ç‡ï¼š[é »ç‡å»ºè­°]"""
        
        # èª¿ç”¨ AI ç”Ÿæˆå®šä½
        if use_gemini():
            ai_response = gemini_generate_text(context)
        else:
            # ç„¡æ¨¡å‹æ™‚çš„ç¯„ä¾‹å›è¦†
            ai_response = f"""æ ¹æ“šã€Œ{theme}ã€ä¸»é¡Œï¼Œæˆ‘ç‚ºä½ ç”Ÿæˆä»¥ä¸‹å®šä½ï¼š

æ¥­å‹™é¡å‹ï¼š{theme}ç›¸é—œæœå‹™
ç›®æ¨™å—çœ¾ï¼šå°{theme}æœ‰èˆˆè¶£çš„å°ç£æ½›åœ¨å®¢æˆ¶
å“ç‰Œèªæ°£ï¼šå°ˆæ¥­è¦ªåˆ‡
ä¸»è¦å¹³å°ï¼šInstagram Reels
å…§å®¹ç›®æ¨™ï¼šå»ºç«‹å°ˆæ¥­å½¢è±¡ï¼Œå¸å¼•æ½›åœ¨å®¢æˆ¶
ç™¼æ–‡é »ç‡ï¼šæ¯é€±2-3æ¬¡"""
        
        add_message(session_id, "assistant", ai_response)
        
        # æå–å®šä½æ¬„ä½ä¸¦æ›´æ–°æª”æ¡ˆ
        extracted_fields = extract_profile_fields(ai_response)
        if extracted_fields:
            update_user_profile(user_id, extracted_fields)
            # é‡æ–°è®€å–æœ€æ–°æª”æ¡ˆ
            user_profile = get_user_profile(user_id)
        
        # ä¿å­˜ AI å›æ‡‰ç‚ºç­†è¨˜
        if ai_response and len(ai_response) > 50:
            add_memory(user_id, "positioning", "note", ai_response, importance_score=8)
        
        return {
            "session_id": session_id,
            "response": ai_response,
            "user_profile": user_profile,
            "extracted_fields": extracted_fields,
            "error": None
        }
        
    except Exception as e:
        print(f"[Generate Positioning Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# é¸é¡Œæ™ºèƒ½é«”
@app.post("/agent/topics/suggest")
async def topic_suggest(req: Request):
    """ç²å–é¸é¡Œå»ºè­°"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        target_date = data.get("target_date")  # YYYY-MM-DD æ ¼å¼
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # ç¢ºä¿ç”¨æˆ¶å­˜åœ¨
        create_or_get_user(user_id)
        
        # è§£ææ—¥æœŸ
        if target_date:
            try:
                from datetime import datetime
                target_date = datetime.strptime(target_date, "%Y-%m-%d").date()
            except ValueError:
                from datetime import date
                target_date = date.today()
        else:
            from datetime import date
            target_date = date.today()
        
        # ç²å–ç”¨æˆ¶æª”æ¡ˆå’Œç›¸é—œè¨˜æ†¶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="topic_selection", limit=5)
        
        # ç”Ÿæˆé¸é¡Œå»ºè­°
        suggestion_context = topic_selection_agent_generate(user_profile, memories)
        
        # èª¿ç”¨ AI ç”Ÿæˆé¸é¡Œ
        if use_gemini():
            ai_response = gemini_generate_text(suggestion_context)
        else:
            ai_response = "AIæœå‹™æš«æ™‚ä¸å¯ç”¨ï¼Œè«‹ç¨å¾Œå†è©¦ã€‚"
        
        # ä¿å­˜é¸é¡Œå»ºè­°
        conn = get_conn()
        conn.execute(
            """INSERT OR REPLACE INTO topic_suggestions 
               (user_id, suggested_date, topics, reasoning) 
               VALUES (?, ?, ?, ?)""",
            (user_id, target_date.isoformat(), json.dumps({"suggestions": ai_response}), ai_response)
        )
        conn.commit()
        conn.close()
        
        return {
            "user_id": user_id,
            "suggested_date": target_date.isoformat(),
            "suggestions": ai_response,
            "reasoning": ai_response,
            "error": None
        }
        
    except Exception as e:
        print(f"[Topic Selection Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

@app.get("/agent/topics/history")
async def topic_history(user_id: str, limit: int = 10):
    """ç²å–é¸é¡Œæ­·å²"""
    try:
        conn = get_conn()
        conn.row_factory = sqlite3.Row
        suggestions = conn.execute(
            "SELECT * FROM topic_suggestions WHERE user_id = ? ORDER BY suggested_date DESC LIMIT ?",
            (user_id, limit)
        ).fetchall()
        conn.close()
        
        return {
            "user_id": user_id,
            "suggestions": [dict(s) for s in suggestions],
            "error": None
        }
        
    except Exception as e:
        print(f"[Topic History Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# è…³æœ¬æ–‡æ¡ˆæ™ºèƒ½é«”ï¼ˆå¢å¼·ç‰ˆï¼‰
@app.post("/agent/content/generate")
async def content_generate(req: Request):
    """ç”Ÿæˆè…³æœ¬æˆ–æ–‡æ¡ˆï¼ˆå¢å¼·ç‰ˆï¼Œæ•´åˆè¨˜æ†¶ç³»çµ±ï¼‰"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        user_input = data.get("user_input", "")
        mode = data.get("mode", "script")  # "script" æˆ– "copy"
        template_type = data.get("template_type")
        duration = data.get("duration")
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        # ç¢ºä¿ç”¨æˆ¶å­˜åœ¨
        create_or_get_user(user_id)
        
        # ç²å–ç”¨æˆ¶æª”æ¡ˆå’Œç›¸é—œè¨˜æ†¶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="script_copy", limit=10)
        
        # å‰µå»ºæœƒè©±
        session_id = create_session(user_id, "script_copy")
        add_message(session_id, "user", user_input)
        
        # æ§‹å»ºå¢å¼·çš„æç¤ºè©ï¼ˆæ•´åˆç”¨æˆ¶æª”æ¡ˆå’Œè¨˜æ†¶ï¼‰
        enhanced_input = user_input
        
        if user_profile:
            profile_context = f"""
ã€ç”¨æˆ¶å®šä½æª”æ¡ˆã€‘
- æ¥­å‹™é¡å‹ï¼š{user_profile.get('business_type', 'æœªè¨­å®š')}
- ç›®æ¨™å—çœ¾ï¼š{user_profile.get('target_audience', 'æœªè¨­å®š')}
- å“ç‰Œèªæ°£ï¼š{user_profile.get('brand_voice', 'æœªè¨­å®š')}
- ä¸»è¦å¹³å°ï¼š{user_profile.get('primary_platform', 'æœªè¨­å®š')}
"""
            enhanced_input = f"{profile_context}\n\nç”¨æˆ¶éœ€æ±‚ï¼š{user_input}"
        
        if memories:
            memory_context = "\nã€ç›¸é—œè¨˜æ†¶ã€‘\n"
            for memory in memories[:3]:
                memory_context += f"- {memory['content']}\n"
            enhanced_input = f"{enhanced_input}\n\n{memory_context}"
        
        # ä½¿ç”¨ç¾æœ‰çš„ chat_generate é‚è¼¯ï¼Œä½†å‚³å…¥å¢å¼·å¾Œçš„è¼¸å…¥
        enhanced_data = {
            "user_id": user_id,
            "session_id": session_id,
            "messages": [{"role": "user", "content": enhanced_input}],
            "mode": mode,
            "template_type": template_type,
            "duration": duration
        }
        
        # èª¿ç”¨ç¾æœ‰çš„ç”Ÿæˆé‚è¼¯
        result = await chat_generate_internal(enhanced_data)
        
        # æ·»åŠ è¨˜æ†¶
        if result.get("assistant_message"):
            add_memory(user_id, "script_copy", "generation", 
                      f"ç”Ÿæˆ{mode}ï¼š{user_input[:100]}...", 
                      importance_score=6)
        
        return result
        
    except Exception as e:
        print(f"[Content Generation Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# è¨˜æ†¶ç³»çµ± API
@app.get("/memory/user/{user_id}")
async def get_user_memory(user_id: str, agent_type: str = None, memory_type: str = None, limit: int = 20):
    """ç²å–ç”¨æˆ¶è¨˜æ†¶"""
    try:
        memories = get_user_memories(user_id, agent_type, memory_type, limit)
        
        return {
            "user_id": user_id,
            "memories": memories,
            "count": len(memories),
            "error": None
        }
        
    except Exception as e:
        print(f"[Memory Retrieval Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

@app.post("/memory/add")
async def add_memory_endpoint(req: Request):
    """æ·»åŠ è¨˜æ†¶"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        agent_type = data.get("agent_type")
        memory_type = data.get("memory_type")
        content = data.get("content")
        importance_score = data.get("importance_score", 5)
        tags = data.get("tags", [])
        
        if not all([user_id, agent_type, memory_type, content]):
            raise HTTPException(status_code=400, detail="Missing required fields")
        
        memory_id = add_memory(user_id, agent_type, memory_type, content, importance_score, tags)
        
        return {
            "memory_id": memory_id,
            "message": "è¨˜æ†¶æ·»åŠ æˆåŠŸ",
            "error": None
        }
        
    except Exception as e:
        print(f"[Memory Addition Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# æ–°å¢ï¼šä¸€éµç”Ÿæˆè…³æœ¬åŠŸèƒ½
@app.post("/agent/script/generate")
async def generate_script_one_click(req: Request):
    """ä¸€éµç”Ÿæˆè…³æœ¬åŠŸèƒ½"""
    try:
        data = await req.json()
        user_id = data.get("user_id")
        theme = data.get("theme", "")  # ç”¨æˆ¶æä¾›çš„ä¸»é¡Œ/æ–‡å­—
        template_type = data.get("template_type", "A")  # é è¨­ä¸‰æ®µå¼
        duration = data.get("duration", 30)  # é è¨­30ç§’
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id is required")
        
        if not theme.strip():
            raise HTTPException(status_code=400, detail="theme is required")
        
        # ç¢ºä¿ç”¨æˆ¶å­˜åœ¨
        create_or_get_user(user_id)
        
        # ç²å–ç”¨æˆ¶æª”æ¡ˆå’Œç›¸é—œè¨˜æ†¶
        user_profile = get_user_profile(user_id)
        memories = get_user_memories(user_id, agent_type="script_copy", limit=10)
        
        # å‰µå»ºæœƒè©±
        session_id = create_session(user_id, "script_copy")
        add_message(session_id, "user", f"ä¸€éµç”Ÿæˆè…³æœ¬ï¼š{theme}")
        
        # æ§‹å»ºä¸€éµç”Ÿæˆæç¤ºè©
        context = f"""æ ¹æ“šä¸»é¡Œã€Œ{theme}ã€ç”ŸæˆçŸ­å½±éŸ³è…³æœ¬ã€‚

ğŸ¯ è…³æœ¬åƒæ•¸ï¼š
â€¢ æ¨¡æ¿ï¼š{template_type} - {TEMPLATE_GUIDE.get(template_type, "ä¸‰æ®µå¼")}
â€¢ æ™‚é•·ï¼š{duration} ç§’
â€¢ å¹³å°ï¼šInstagram Reelsã€TikTokã€YouTube Shortsã€Facebook Reels

ğŸ“š çŸ¥è­˜åº«ï¼š
{BUILTIN_KB_SCRIPT}

ğŸ’¡ å°ç£å¸‚å ´ç‰¹è‰²ï¼š
â€¢ å…§å®¹é¢¨æ ¼ï¼šç”Ÿæ´»åŒ–ã€è¦ªåˆ‡ã€å¯¦ç”¨
â€¢ ç¯€å¥è¦æ±‚ï¼š2-3ç§’æ›ç•«é¢ï¼Œç¯€å¥ç·Šæ¹Š
â€¢ HookåŸå‰‡ï¼š0-5ç§’ç›´çµ¦çµè«–ï¼Œç”¨å¤§å­—å¡èˆ‡å¼·æƒ…ç·’è¡¨æƒ…
â€¢ èªæ°£ï¼šå …å®šã€ç›´çµ¦çµè«–ï¼Œé¿å…å£ç™–è´…å­—

ç›´æ¥è¼¸å‡ºJSONæ ¼å¼ï¼Œä¸è¦ä»»ä½•é–‹å ´ç™½æˆ–èªªæ˜æ–‡å­—ï¼š

{{
  "segments":[
    {{"type":"hook","start_sec":0,"end_sec":5,"camera":"CU","dialog":"...","visual":"...","cta":""}},
    {{"type":"value","start_sec":5,"end_sec":25,"camera":"MS","dialog":"...","visual":"...","cta":""}},
    {{"type":"cta","start_sec":25,"end_sec":30,"camera":"WS","dialog":"...","visual":"...","cta":"..."}}
  ]
}}"""
        
        # èª¿ç”¨ AI ç”Ÿæˆè…³æœ¬
        if use_gemini():
            ai_response = gemini_generate_text(context)
        else:
            # ç„¡æ¨¡å‹æ™‚çš„ç¯„ä¾‹å›è¦†
            ai_response = f"""{{
  "segments":[
    {{"type":"hook","start_sec":0,"end_sec":5,"camera":"CU","dialog":"ä½ çŸ¥é“{theme}çš„ç§˜å¯†å—ï¼Ÿ","visual":"å¤§å­—å¡+é©šè¨è¡¨æƒ…","cta":""}},
    {{"type":"value","start_sec":5,"end_sec":25,"camera":"MS","dialog":"ä»Šå¤©æˆ‘è¦åˆ†äº«{theme}çš„å¯¦ç”¨æŠ€å·§ï¼Œè®“ä½ è¼•é¬†æŒæ¡ï¼","visual":"ç¤ºç¯„ç•«é¢","cta":""}},
    {{"type":"cta","start_sec":25,"end_sec":30,"camera":"WS","dialog":"æƒ³è¦æ›´å¤š{theme}æŠ€å·§ï¼Œè¨˜å¾—é—œæ³¨æˆ‘ï¼","visual":"é—œæ³¨æŒ‰éˆ•","cta":"é»é—œæ³¨"}}
  ]
}}"""
        
        add_message(session_id, "assistant", ai_response)
        
        # è§£æè…³æœ¬
        try:
            if use_gemini():
                segments = parse_segments(ai_response)
            else:
                # è§£æç¯„ä¾‹å›è¦†
                import json
                data = json.loads(ai_response)
                segments = data.get("segments", [])
        except Exception as e:
            print(f"[Script Parse Error] {e}")
            segments = []
        
        # ä¿å­˜è…³æœ¬ç”Ÿæˆç‚ºç­†è¨˜
        if ai_response and len(ai_response) > 50:
            add_memory(user_id, "script_copy", "note", ai_response, importance_score=8)
        
        return {
            "session_id": session_id,
            "assistant_message": "ğŸš€ ä¸€éµç”Ÿæˆå®Œæˆï¼æˆ‘ç‚ºä½ ç”Ÿæˆäº†å®Œæ•´çš„è…³æœ¬ã€‚",
            "segments": segments,
            "error": None
        }
        
    except Exception as e:
        print(f"[One-Click Script Generation Error] {e}")
        return JSONResponse(
            status_code=500,
            content={"error": "internal_server_error", "message": str(e)}
        )

# å…§éƒ¨å‡½æ•¸ï¼šchat_generate çš„å…§éƒ¨é‚è¼¯ï¼ˆä¾› content_generate èª¿ç”¨ï¼‰
async def chat_generate_internal(data: dict):
    """chat_generate çš„å…§éƒ¨é‚è¼¯ï¼Œä¾›å…¶ä»–å‡½æ•¸èª¿ç”¨"""
    user_id = (data.get("user_id") or "").strip() or "web"
    messages = data.get("messages") or []
    previous_segments = data.get("previous_segments") or []
    topic = (data.get("topic") or "").strip() or None

    explicit_mode = (data.get("mode") or "").strip().lower() or None
    mode = detect_mode(messages, explicit=explicit_mode)

    dialogue_mode = (data.get("dialogue_mode") or "").strip().lower() or None
    template_type = (data.get("template_type") or "").strip().upper() or None
    try:
        duration = int(data.get("duration")) if data.get("duration") is not None else None
    except Exception:
        duration = None
    knowledge_hint = (data.get("knowledge_hint") or "").strip() or None

    user_input = ""
    for m in reversed(messages):
        if m.get("role") == "user":
            user_input = (m.get("content") or "").strip()
            break

    # è¼¸å…¥éçŸ­æ™‚ï¼Œä»æŒçºŒå°è©±è€Œéå›å‚³åˆ¶å¼æç¤º
    hint = SHORT_HINT_COPY if mode == "copy" else SHORT_HINT_SCRIPT
    if len(user_input) < 6:
        user_input = f"ï¼ˆä½¿ç”¨è€…æç¤ºè¼ƒçŸ­ï¼‰è«‹ä¸»å‹•è¿½å•é—œéµæ¢ä»¶ä¸¦å…ˆçµ¦å‡ºæ–¹å‘æ€§å»ºè­°ã€‚\næç¤ºï¼š{user_input or 'é–‹å§‹'}"

    try:
        if mode == "copy":
            prompt = build_copy_prompt(user_input, topic)
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                copy = parse_copy(j)
            else:
                copy = fallback_copy(user_input, topic)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "æˆ‘å…ˆçµ¦ä½ ç¬¬ä¸€ç‰ˆå®Œæ•´è²¼æ–‡ï¼ˆå¯å†åŠ è¦æ±‚ï¼Œæˆ‘æœƒå¹«ä½ æ”¹å¾—æ›´è²¼è¿‘é¢¨æ ¼ï¼‰ã€‚",
                "segments": [],
                "copy": copy,
                "error": None
            }

        else:  # script
            prompt = build_script_prompt(
                user_input,
                previous_segments,
                template_type=template_type,
                duration=duration,
                dialogue_mode=dialogue_mode,
                knowledge_hint=knowledge_hint,
            )
            if use_gemini():
                out = gemini_generate_text(prompt)
                j = _ensure_json_block(out)
                segments = parse_segments(j)
            else:
                segments = fallback_segments(user_input, len(previous_segments or []), duration=duration)

            resp = {
                "session_id": data.get("session_id") or "s",
                "assistant_message": "æˆ‘å…ˆçµ¦ä½ ç¬¬ä¸€ç‰ˆå®Œæ•´è…³æœ¬ï¼ˆå¯å†åŠ è¦æ±‚ï¼Œæˆ‘æœƒå¹«ä½ æ”¹å¾—æ›´è²¼è¿‘é¢¨æ ¼ï¼‰ã€‚",
                "segments": segments,
                "copy": None,
                "error": None
            }

        return resp

    except Exception as e:
        print("[chat_generate_internal] error:", e)
        return {
            "session_id": data.get("session_id") or "s",
            "assistant_message": "ä¼ºæœå™¨å¿™ç¢Œï¼Œç¨å¾Œå†è©¦",
            "segments": [],
            "copy": None,
            "error": "internal_server_error"
        }

# å•Ÿå‹•æœå‹™å™¨
if __name__ == "__main__":
    import uvicorn
    print("ğŸš€ å•Ÿå‹•ä¸‰æ™ºèƒ½é«”ç³»çµ±...")
    print("ğŸ“ æœ¬åœ°è¨ªå•ï¼šhttp://localhost:8080")
    print("ğŸ“‹ API æ–‡æª”ï¼šhttp://localhost:8080/docs")
    uvicorn.run(app, host="0.0.0.0", port=8080, log_level="info")